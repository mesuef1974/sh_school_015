import csv
from datetime import datetime
from pathlib import Path
from typing import Optional, Iterable

from django.core.management.base import BaseCommand, CommandError

from core.models import Class, Staff, Student

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional import; validated at runtime if .xlsx is used
    load_workbook = None  # type: ignore


DATE_INPUT_FORMATS = (
    "%Y-%m-%d",  # 2024-10-02
    "%d/%m/%Y",  # 02/10/2024
    "%d-%m-%Y",  # 02-10-2024
    "%m/%d/%Y",  # 10/02/2024
)


def parse_date(value: str) -> Optional[datetime.date]:
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    for fmt in DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    # Try ISO-like without padding, e.g. 2024/1/5
    try:
        parts = [int(p) for p in value.replace("\\", "/").replace("-", "/").split("/")]
        if len(parts) == 3 and parts[0] > 31:  # assume Y/M/D
            return datetime(parts[0], parts[1], parts[2]).date()
    except Exception:
        pass
    raise ValueError(f"Unrecognized date format: {value}")


def normalize_header(h: str) -> str:
    return h.strip().lower().replace(" ", "_")


def read_csv_rows(path: Path, delimiter: str = ",") -> Iterable[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        # Normalize headers
        reader.fieldnames = [normalize_header(h) for h in reader.fieldnames or []]
        for row in reader:
            yield {normalize_header(k): (v or "").strip() for k, v in row.items()}


def read_xlsx_first_sheet(path: Path) -> Iterable[dict]:
    if load_workbook is None:
        raise CommandError("openpyxl is required to read .xlsx files. Please install 'openpyxl'.")
    if not path.exists():
        raise CommandError(f"Excel file not found: {path}")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if not wb.worksheets:
        return []
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [normalize_header((h or "")) for h in next(rows)]
    except StopIteration:
        return []
    for r in rows:
        # Pad/truncate to headers length
        values = list(r) if r is not None else []
        values += [None] * (len(headers) - len(values))
        row = {headers[i]: ("" if values[i] is None else str(values[i]).strip()) for i in range(len(headers))}
        # Skip empty lines
        if any(v for v in row.values()):
            yield row


def iter_xlsx_sheets(path: Path) -> Iterable[tuple[str, dict]]:
    """Yield (sheet_name, row_dict) for every non-empty row across worksheets."""
    if load_workbook is None:
        raise CommandError("openpyxl is required to read .xlsx files. Please install 'openpyxl'.")
    if not path.exists():
        raise CommandError(f"Excel file not found: {path}")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [normalize_header((h or "")) for h in next(rows)]
        except StopIteration:
            continue
        for r in rows:
            values = list(r) if r is not None else []
            values += [None] * (len(headers) - len(values))
            row = {headers[i]: ("" if values[i] is None else str(values[i]).strip()) for i in range(len(headers))}
            if any(v for v in row.values()):
                yield ws.title.strip(), row


class Command(BaseCommand):
    help = (
        "Import staff and students from CSV or Excel (.xlsx) files. The command performs upserts "
        "based on unique codes (staff_code, student_code). Works with the configured database, "
        "including PostgreSQL when USE_SQLITE=False and DB_* are set in backend/.env."
    )

    def add_arguments(self, parser):
        parser.add_argument("--staff", type=str, help="Path to staff CSV/XLSX file")
        parser.add_argument("--students", type=str, help="Path to students CSV/XLSX file")
        parser.add_argument(
            "--delimiter",
            type=str,
            default=",",
            help="CSV delimiter (default ',')",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and show what would change without writing to DB",
        )
        parser.add_argument(
            "--create-classes",
            action="store_true",
            help="Auto-create classes referenced by students if missing",
        )

    def handle(self, *args, **options):
        staff_path = options.get("staff")
        students_path = options.get("students")
        delimiter = options.get("delimiter") or ","
        dry_run = bool(options.get("dry_run"))
        create_classes = bool(options.get("create_classes"))

        if not staff_path and not students_path:
            raise CommandError("Provide at least one of --staff or --students file paths (.csv or .xlsx).")

        total_created = 0
        total_updated = 0
        notes: list[str] = []

        if staff_path:
            sp = Path(staff_path)
            if sp.suffix.lower() == ".xlsx":
                created, updated = self._import_staff_xlsx(sp, dry_run)
            else:
                created, updated = self._import_staff(sp, delimiter, dry_run)
            total_created += created
            total_updated += updated

        if students_path:
            sp = Path(students_path)
            if sp.suffix.lower() == ".xlsx":
                created, updated, note_list = self._import_students_xlsx(sp, dry_run, create_classes)
            else:
                created, updated, note_list = self._import_students(sp, delimiter, dry_run, create_classes)
            total_created += created
            total_updated += updated
            notes.extend(note_list)

        msg = (
            f"Completed import. Created: {total_created}, Updated: {total_updated}. "
            + ("DRY-RUN: no changes were written." if dry_run else "")
        )
        if notes:
            for n in notes:
                self.stdout.write(self.style.WARNING(n))
        self.stdout.write(self.style.SUCCESS(msg))

    # ---- helpers ----
    def _import_staff(self, path: Path, delimiter: str, dry_run: bool) -> tuple[int, int]:
        if not path.exists():
            raise CommandError(f"Staff CSV not found: {path}")
        created = 0
        updated = 0
        expected = {"staff_code", "name", "job_title", "email", "phone"}
        for row in read_csv_rows(path, delimiter):
            # header aliases
            staff_code = row.get("staff_code") or row.get("code")
            name = row.get("name") or row.get("full_name")
            job_title = row.get("job_title") or row.get("title") or ""
            email = row.get("email") or ""
            phone = row.get("phone") or row.get("mobile") or ""
            if not staff_code or not name:
                # Skip incomplete rows silently
                continue

            defaults = {
                "name": name,
                "job_title": job_title,
                "email": email,
                "phone": phone,
            }
            obj = None
            if not dry_run:
                obj, is_created = Staff.objects.update_or_create(
                    staff_code=staff_code, defaults=defaults
                )
            else:
                is_created = not Staff.objects.filter(staff_code=staff_code).exists()
            if is_created:
                created += 1
            else:
                updated += 1
        return created, updated

    def _import_students(
        self, path: Path, delimiter: str, dry_run: bool, create_classes: bool
    ) -> tuple[int, int, list[str]]:
        if not path.exists():
            raise CommandError(f"Students CSV not found: {path}")
        created = 0
        updated = 0
        notes: list[str] = []

        for row in read_csv_rows(path, delimiter):
            student_code = row.get("student_code") or row.get("code")
            name = row.get("name") or row.get("full_name")
            dob_raw = row.get("date_of_birth") or row.get("dob") or row.get("birthdate")
            class_name = (
                row.get("class")
                or row.get("class_name")
                or row.get("clazz")
                or row.get("section")
            )

            if not student_code or not name:
                continue

            dob = None
            if dob_raw:
                try:
                    dob = parse_date(dob_raw)
                except Exception as ex:
                    notes.append(f"Row student_code={student_code}: {ex}")

            clazz = None
            if class_name:
                class_name = class_name.strip()
                try:
                    clazz = Class.objects.get(name=class_name)
                except Class.DoesNotExist:
                    if create_classes and not dry_run:
                        clazz = Class.objects.create(name=class_name, grade_level=0)
                    elif create_classes and dry_run:
                        clazz = None
                        # Will appear created in reality; add a note
                        notes.append(
                            f"Would create Class(name='{class_name}', grade_level=0)"
                        )
                    else:
                        notes.append(
                            f"Class '{class_name}' not found; student '{student_code}' will have no class."
                        )

            defaults = {
                "name": name,
                "date_of_birth": dob,
                "clazz": clazz,
            }

            if not dry_run:
                _, is_created = Student.objects.update_or_create(
                    student_code=student_code, defaults=defaults
                )
            else:
                is_created = not Student.objects.filter(student_code=student_code).exists()

            if is_created:
                created += 1
            else:
                updated += 1

        return created, updated, notes

# --- XLSX helpers appended by Junie ---
    def _import_staff_xlsx(self, path: Path, dry_run: bool) -> tuple[int, int]:
        if not path.exists():
            raise CommandError(f"Staff Excel not found: {path}")
        created = 0
        updated = 0
        for row in read_xlsx_first_sheet(path):
            staff_code = row.get("staff_code") or row.get("code")
            name = row.get("name") or row.get("full_name")
            job_title = row.get("job_title") or row.get("title") or ""
            email = row.get("email") or ""
            phone = row.get("phone") or row.get("mobile") or ""
            if not staff_code or not name:
                continue
            defaults = {
                "name": name,
                "job_title": job_title,
                "email": email,
                "phone": phone,
            }
            if not dry_run:
                _, is_created = Staff.objects.update_or_create(
                    staff_code=staff_code, defaults=defaults
                )
            else:
                is_created = not Staff.objects.filter(staff_code=staff_code).exists()
            if is_created:
                created += 1
            else:
                updated += 1
        return created, updated

    def _import_students_xlsx(
        self, path: Path, dry_run: bool, create_classes: bool
    ) -> tuple[int, int, list[str]]:
        if not path.exists():
            raise CommandError(f"Students Excel not found: {path}")
        created = 0
        updated = 0
        notes: list[str] = []

        for sheet_name, row in iter_xlsx_sheets(path):
            student_code = row.get("student_code") or row.get("code")
            name = row.get("name") or row.get("full_name")
            dob_raw = row.get("date_of_birth") or row.get("dob") or row.get("birthdate")
            class_name = (
                row.get("class")
                or row.get("class_name")
                or row.get("clazz")
                or row.get("section")
            )
            # If class not provided in columns, fallback to the worksheet title
            if not class_name:
                class_name = sheet_name

            if not student_code or not name:
                continue

            dob = None
            if dob_raw:
                try:
                    dob = parse_date(dob_raw)
                except Exception as ex:
                    notes.append(f"Row student_code={student_code}: {ex}")

            clazz = None
            if class_name:
                class_name = str(class_name).strip()
                try:
                    clazz = Class.objects.get(name=class_name)
                except Class.DoesNotExist:
                    if create_classes and not dry_run:
                        clazz = Class.objects.create(name=class_name, grade_level=0)
                    elif create_classes and dry_run:
                        clazz = None
                        notes.append(
                            f"Would create Class(name='{class_name}', grade_level=0)"
                        )
                    else:
                        notes.append(
                            f"Class '{class_name}' not found; student '{student_code}' will have no class."
                        )

            defaults = {
                "name": name,
                "date_of_birth": dob,
                "clazz": clazz,
            }

            if not dry_run:
                _, is_created = Student.objects.update_or_create(
                    student_code=student_code, defaults=defaults
                )
            else:
                is_created = not Student.objects.filter(student_code=student_code).exists()

            if is_created:
                created += 1
            else:
                updated += 1

        return created, updated, notes