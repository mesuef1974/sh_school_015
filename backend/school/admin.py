from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django import forms
from django.utils.html import format_html
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from .models import (
    Class,
    Student,
    Staff,
    Subject,
    TeachingAssignment,
    ClassSubject,
    Wing,
    AcademicYear,
    Term,
    PeriodTemplate,
    TemplateSlot,
    TimetableEntry,
    AttendancePolicy,
    AttendanceRecord,
    AttendanceDaily,
    AssessmentPackage,
    SchoolHoliday,
    ExitEvent,
)
from openpyxl import load_workbook
import re


class ClassSubjectInline(admin.TabularInline):
    model = ClassSubject
    extra = 1
    autocomplete_fields = ("subject",)
    fields = ("subject", "weekly_default")


@admin.register(Class)
class ClassAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "grade", "section", "wing", "students_count")
    search_fields = ("name",)
    list_filter = ("grade", "section", "wing")
    ordering = ("grade", "name")
    inlines = [ClassSubjectInline]


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "sid",
        "student_name_one_line",
        "english_name",
        "national_no",
        "needs",
        "class_fk",
        "grade_label",
        "section_label",
        "dob",
        "age",
        "nationality",
        "phone_no",
        "extra_phone_no",
        "email",
        "guardian_name_one_line",
        "parent_relation",
        "parent_national_no",
        "parent_phone",
        "parent_email",
        "active",
    )
    list_display_links = ("sid", "student_name_one_line")
    search_fields = (
        "sid",
        "full_name",
        "national_no",
        "phone_no",
        "parent_phone",
        "parent_name",
        "parent_national_no",
        "email",
        "parent_email",
        "nationality",
        "grade_label",
        "section_label",
    )
    list_filter = (
        "active",
        "class_fk",
        "needs",
        "nationality",
        "grade_label",
        "section_label",
    )
    list_select_related = ("class_fk",)
    list_per_page = 100

    def student_name_one_line(self, obj: Student):
        name = obj.full_name or ""
        return format_html('<span style="white-space:nowrap;">{}</span>', name)

    student_name_one_line.short_description = "اسم الطالب"

    def guardian_name_one_line(self, obj: Student):
        name = obj.parent_name or ""
        return format_html('<span style="white-space:nowrap;">{}</span>', name)

    guardian_name_one_line.short_description = "اسم ولي الأمر"


@admin.register(Staff)
class StaffAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "full_name",
        "get_role_display",
        "job_title",
        "job_no",
        "national_no",
        "email",
        "phone_no",
        "user",
    )
    search_fields = ("full_name", "national_no", "job_no", "email", "phone_no")
    list_filter = (
        "role",
        ("user", admin.EmptyFieldListFilter),
        ("email", admin.EmptyFieldListFilter),
        ("phone_no", admin.EmptyFieldListFilter),
        ("national_no", admin.EmptyFieldListFilter),
        ("job_no", admin.EmptyFieldListFilter),
        ("job_title", admin.EmptyFieldListFilter),
    )


class StaffInline(admin.StackedInline):
    model = Staff
    can_delete = False
    extra = 0
    fk_name = "user"
    fields = (
        "full_name",
        "role",
        "job_title",
        "job_no",
        "national_no",
        "email",
        "phone_no",
        "created_at",
        "updated_at",
    )
    readonly_fields = ("created_at", "updated_at")


# Customize the built-in User admin to include Staff information
if admin.site.is_registered(User):
    admin.site.unregister(User)


@admin.register(User)
class UserAdmin(DjangoUserAdmin):
    inlines = [StaffInline]
    list_select_related = ("staff",)

    # Remove 'Personal info' fieldset and keep others
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        (
            "Permissions",
            {
                "fields": (
                    "is_active",
                    "is_staff",
                    "is_superuser",
                    "groups",
                    "user_permissions",
                )
            },
        ),
        (
            "Important dates",
            {
                "fields": (
                    "last_login",
                    "date_joined",
                )
            },
        ),
    )

    # Extend list display with staff fields
    def staff_full_name(self, obj: User):
        s = getattr(obj, "staff", None)
        return s.full_name if s else "—"

    def staff_role(self, obj: User):
        s = getattr(obj, "staff", None)
        return s.get_role_display() if s else "—"

    def staff_phone(self, obj: User):
        s = getattr(obj, "staff", None)
        return s.phone_no if s else ""

    def staff_job_no(self, obj: User):
        s = getattr(obj, "staff", None)
        return s.job_no if s else ""

    def staff_national_no(self, obj: User):
        s = getattr(obj, "staff", None)
        return s.national_no if s else ""

    staff_full_name.short_description = "اسم الموظف"
    staff_role.short_description = "الصفة/الدور"
    staff_phone.short_description = "جوال الموظف"
    staff_job_no.short_description = "الرقم الوظيفي"
    staff_national_no.short_description = "رقم الهوية"

    # Insert our columns near the start (keeping username, email)
    list_display = (
        "username",
        "email",
        "is_active",
        "is_staff",
        "is_superuser",
        "last_login",
        "staff_full_name",
        "staff_role",
        "staff_phone",
        "staff_job_no",
        "staff_national_no",
    )

    # Allow searching by staff fields
    search_fields = (
        "username",
        "email",
        "first_name",
        "last_name",
        "staff__full_name",
        "staff__phone_no",
        "staff__national_no",
        "staff__job_no",
    )

    # Add filters including whether user has Staff row and by role
    list_filter = DjangoUserAdmin.list_filter + (
        ("staff", admin.EmptyFieldListFilter),
        ("staff__role", admin.AllValuesFieldListFilter),
    )


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ("id", "name_ar", "code", "weekly_default", "is_active")
    search_fields = ("name_ar", "code")
    list_filter = ("is_active",)
    inlines = []  # filled below


class ClassSubjectBySubjectInline(admin.TabularInline):
    model = ClassSubject
    extra = 1
    autocomplete_fields = ("classroom",)
    fields = ("classroom", "weekly_default")


# Attach inline after its definition
SubjectAdmin.inlines = [ClassSubjectBySubjectInline]


class TeachingAssignmentForm(forms.ModelForm):
    class Meta:
        model = TeachingAssignment
        fields = ["teacher", "classroom", "subject", "no_classes_weekly", "notes"]


@admin.register(TeachingAssignment)
class TeachingAssignmentAdmin(admin.ModelAdmin):
    form = TeachingAssignmentForm
    list_display = (
        "id",
        "teacher",
        "classroom",
        "subject",
        "no_classes_weekly",
        "updated_at",
    )
    search_fields = (
        "teacher__full_name",
        "classroom__name",
        "subject__name_ar",
    )
    list_filter = ("classroom__grade", "classroom__section", "subject", "teacher")
    autocomplete_fields = ("teacher", "classroom", "subject")
    actions = ["export_summary"]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "subject":
            classroom_id = None
            try:
                classroom_id = request.GET.get("classroom") or request.POST.get("classroom")
            except Exception:
                classroom_id = None
            # Try resolve object_id when editing existing assignment
            if not classroom_id and hasattr(request, "resolver_match") and request.resolver_match:
                obj_id = (
                    request.resolver_match.kwargs.get("object_id")
                    if hasattr(request.resolver_match, "kwargs")
                    else None
                )
                if obj_id:
                    try:
                        obj = TeachingAssignment.objects.get(pk=obj_id)
                        classroom_id = obj.classroom_id
                    except TeachingAssignment.DoesNotExist:
                        pass
            if classroom_id:
                subject_ids = ClassSubject.objects.filter(classroom_id=classroom_id).values_list(
                    "subject_id", flat=True
                )
                kwargs["queryset"] = Subject.objects.filter(id__in=list(subject_ids))
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel_view),
                name="teachingassignment_import_excel",
            ),
        ]
        return custom + urls

    def export_summary(self, request, queryset):
        from django.http import HttpResponse
        import csv

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="teacher_loads_summary.csv"'
        writer = csv.writer(response)
        writer.writerow(["المعلم", "الصف", "المادة", "حصص/أسبوع"])
        for a in queryset.select_related("teacher", "classroom", "subject"):
            writer.writerow(
                [
                    a.teacher.full_name,
                    a.classroom.name,
                    a.subject.name_ar,
                    a.no_classes_weekly,
                ]
            )
        return response

    export_summary.short_description = "تصدير الملخص CSV"

    # ===== Excel Import =====
    class ImportExcelForm(forms.Form):
        file = forms.FileField(label="ملف Excel للأنصبة (schasual.xlsx)")

    def import_excel_view(self, request):
        if request.method == "POST":
            form = self.ImportExcelForm(request.POST, request.FILES)
            if form.is_valid():
                f = form.cleaned_data["file"]
                try:
                    count, created_subjects = self._import_excel_file(f)
                    messages.success(
                        request,
                        f"تم استيراد {count} سجل. تم إنشاء مواد جديدة: {created_subjects}",
                    )
                    return redirect("..")
                except Exception as e:
                    messages.error(request, f"خطأ أثناء الاستيراد: {e}")
        else:
            form = self.ImportExcelForm()
        context = {
            **self.admin_site.each_context(request),
            "form": form,
            "title": "استيراد الأنصبة من Excel",
        }
        return render(request, "admin/teachingassignment_import.html", context)

    @staticmethod
    def _normalize_ar_text(s: str) -> str:
        if s is None:
            return ""
        s = str(s)
        s = re.sub(r"\s+", " ", s).strip()
        s = s.replace("ـ", "")
        s = re.sub(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]", "", s)
        replacements = {
            "أ": "ا",
            "إ": "ا",
            "آ": "ا",
            "ٱ": "ا",
            "ى": "ي",
            "ئ": "ي",
            "ؤ": "و",
            "ة": "ه",
        }
        for k, v in replacements.items():
            s = s.replace(k, v)
        s = re.sub(r"[^\w\s\u0600-\u06FF]", "", s)
        trans_digits = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
        s = s.translate(trans_digits)
        return s

    def _import_excel_file(self, file_obj):
        wb = load_workbook(filename=file_obj, data_only=True)
        created_subjects = 0
        created_assignments = 0

        # Build teacher index (normalized)
        teacher_index = {}
        for t in Staff.objects.all():
            key = self._normalize_ar_text(t.full_name).replace(" ", "").lower()
            if key:
                teacher_index.setdefault(key, t)

        # Build class index by (grade, section) and by name
        classes = Class.objects.all()
        class_by_grade_section = {(c.grade, (c.section or "").strip()): c for c in classes}
        class_by_name = {self._normalize_ar_text(c.name): c for c in classes}

        # Subject index
        subject_by_norm = {self._normalize_ar_text(s.name_ar): s for s in Subject.objects.all()}

        total_rows = 0
        for ws in wb.worksheets:
            # Try detect headers positions
            # Heuristic columns: teacher_name, grade, section, class(subject), no.classes/week
            header_row = None
            col_map = {
                "teacher": None,
                "grade": None,
                "section": None,
                "subject": None,
                "weekly": None,
            }
            for r in range(1, 11):
                try:
                    vals = [
                        c if c is not None else ""
                        for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
                    ]
                except StopIteration:
                    break
                normed = [self._normalize_ar_text(v).replace(" ", "").lower() for v in vals]
                tokens = {
                    "teacher": {
                        "teachername",
                        "teacher",
                        "اسمالمعلم",
                        "اسمالمعلّم",
                        "اسمالمعلم",
                        "المعلم",
                    },
                    "grade": {"grade", "الصف", "المرحلة", "الصفالدراسي"},
                    "section": {"section", "الشعبة", "الفصل"},
                    "subject": {"class", "subject", "المادة", "المواد"},
                    "weekly": {
                        "no.classesw",
                        "weekly",
                        "weeklyclasses",
                        "الحصصالاسبوعية",
                        "الحصصالأسبوعية",
                        "حصصالاسبوع",
                    },
                }
                for i, v in enumerate(normed):
                    for k, opts in tokens.items():
                        if v in opts and col_map[k] is None:
                            col_map[k] = i
                if any(col_map.values()):
                    header_row = r
                    break
            if header_row is None:
                header_row = 1
                col_map = {
                    "teacher": 0,
                    "grade": 1,
                    "section": 2,
                    "subject": 3,
                    "weekly": 4,
                }

            current_teacher = None
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                total_rows += 1
                if not row:
                    continue
                # Carry-forward teacher if grouped
                tv = (
                    row[col_map["teacher"]]
                    if col_map["teacher"] is not None and col_map["teacher"] < len(row)
                    else None
                )
                teacher_display = str(tv).strip() if tv not in (None, "") else None
                if teacher_display:
                    current_teacher = teacher_display
                if not current_teacher:
                    continue

                gv = (
                    row[col_map["grade"]]
                    if col_map["grade"] is not None and col_map["grade"] < len(row)
                    else None
                )
                sv = (
                    row[col_map["section"]]
                    if col_map["section"] is not None and col_map["section"] < len(row)
                    else None
                )
                subjv = (
                    row[col_map["subject"]]
                    if col_map["subject"] is not None and col_map["subject"] < len(row)
                    else None
                )
                wv = (
                    row[col_map["weekly"]]
                    if col_map["weekly"] is not None and col_map["weekly"] < len(row)
                    else None
                )

                # Normalize
                t_key = self._normalize_ar_text(current_teacher).replace(" ", "").lower()
                teacher = teacher_index.get(t_key)
                if not teacher:
                    # skip if teacher not found
                    continue

                def to_int_safe(x):
                    try:
                        return int(str(x).strip().split()[0])
                    except Exception:
                        return None

                grade = to_int_safe(gv)
                section = None
                if sv not in (None, ""):
                    section = re.sub(r"[^0-9A-Za-zأ-ي]", "", str(sv)).strip()
                subj_norm = self._normalize_ar_text(subjv)
                weekly = to_int_safe(wv) or 0

                # Map class
                classroom = None
                if grade is not None:
                    classroom = class_by_grade_section.get((grade, section or ""))
                if classroom is None and subj_norm:
                    # fallback by name token present in class name
                    # e.g., Class name like "12-1" or "12/1" already set in DB
                    if grade is not None and section is not None:
                        key = f"{grade}{section}"
                        for cname, cobj in class_by_name.items():
                            if key in re.sub(r"\D", "", cname):
                                classroom = cobj
                                break
                if classroom is None:
                    # final fallback: try by normalized name equality if provided in sheet
                    pass

                if classroom is None:
                    continue

                # Subject
                if subj_norm:
                    subj_obj = subject_by_norm.get(subj_norm)
                    if not subj_obj:
                        subj_obj = Subject.objects.create(name_ar=subjv)
                        subject_by_norm[subj_norm] = subj_obj
                        created_subjects += 1
                else:
                    continue

                # Ensure the subject is assigned to the class
                ClassSubject.objects.get_or_create(classroom=classroom, subject=subj_obj)

                obj, created = TeachingAssignment.objects.update_or_create(
                    teacher=teacher,
                    classroom=classroom,
                    subject=subj_obj,
                    defaults={"no_classes_weekly": weekly or subj_obj.weekly_default or 0},
                )
                if created:
                    created_assignments += 1

        return created_assignments, created_subjects


# ===== New admin registrations for attendance and calendar =====
# ==== Wing admin: allow assigning existing classes to a wing (not as notes) ====
class WingAdminForm(forms.ModelForm):
    classes_attach = forms.ModelMultipleChoiceField(
        label="الصفوف",
        queryset=Class.objects.none(),  # set in __init__
        required=False,
        help_text="اختر الصفوف التي تتبع هذا الجناح",
        widget=admin.widgets.FilteredSelectMultiple("الصفوف", is_stacked=False),
    )

    class Meta:
        model = Wing
        fields = ["name", "floor", "notes", "supervisor", "classes_attach"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Show all classes plus preselect those currently attached to this wing
        qs = Class.objects.all().order_by("name")
        self.fields["classes_attach"].queryset = qs
        if self.instance and self.instance.pk:
            self.fields["classes_attach"].initial = list(self.instance.classes.all())


@admin.register(Wing)
class WingAdmin(admin.ModelAdmin):
    form = WingAdminForm
    list_display = ("id", "name", "floor", "supervisor", "classes_count")
    search_fields = ("name", "notes")
    list_filter = ("floor",)
    filter_horizontal = ()  # silence admin js expectations for FilteredSelectMultiple

    def classes_count(self, obj: Wing):
        return obj.classes.count()

    classes_count.short_description = "عدد الصفوف"

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        # Defer handling of classes to save_related where m2m-like widget is processed

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        # Update Class.wing based on selected classes
        wing = form.instance
        if "classes_attach" in form.cleaned_data:
            selected = set(form.cleaned_data["classes_attach"].values_list("id", flat=True))
            # Set the selected ones to this wing
            Class.objects.filter(id__in=selected).update(wing=wing)
            # Detach any previously attached classes that were unselected
            Class.objects.filter(wing=wing).exclude(id__in=selected).update(wing=None)


@admin.register(AcademicYear)
class AcademicYearAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "start_date", "end_date", "is_current")
    list_filter = ("is_current",)
    search_fields = ("name",)


@admin.register(Term)
class TermAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "name",
        "academic_year",
        "start_date",
        "end_date",
        "is_current",
    )
    list_filter = ("academic_year", "is_current")
    search_fields = ("name", "academic_year__name")


class TemplateSlotInline(admin.TabularInline):
    model = TemplateSlot
    extra = 0
    fields = ("number", "start_time", "end_time", "kind")


@admin.register(PeriodTemplate)
class PeriodTemplateAdmin(admin.ModelAdmin):
    list_display = ("id", "code", "name", "day_of_week", "scope")
    list_filter = ("day_of_week", "scope")
    search_fields = ("code", "name", "scope")
    inlines = [TemplateSlotInline]


@admin.register(TimetableEntry)
class TimetableEntryAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "classroom",
        "day_of_week",
        "period_number",
        "subject",
        "teacher",
        "term",
    )
    list_filter = ("day_of_week", "classroom", "teacher", "subject", "term")
    autocomplete_fields = ("classroom", "teacher", "subject", "term")
    search_fields = ("classroom__name", "teacher__full_name", "subject__name_ar")


@admin.register(AttendancePolicy)
class AttendancePolicyAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "term",
        "late_threshold_minutes",
        "late_to_equivalent_period_minutes",
        "first_two_periods_numbers",
        "lesson_lock_after_minutes",
        "daily_lock_time",
    )
    list_filter = ("term",)


@admin.register(AttendanceRecord)
class AttendanceRecordAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "student",
        "classroom",
        "subject",
        "teacher",
        "term",
        "date",
        "day_of_week",
        "period_number",
        "status",
        "note",
        "late_minutes",
        "early_minutes",
        "locked",
        "updated_at",
    )
    list_filter = (
        "date",
        "day_of_week",
        "status",
        "classroom",
        "teacher",
        "subject",
        "term",
    )
    search_fields = (
        "student__full_name",
        "classroom__name",
        "subject__name_ar",
        "teacher__full_name",
        "note",
    )
    autocomplete_fields = ("student", "classroom", "subject", "teacher", "term")


@admin.register(AttendanceDaily)
class AttendanceDailyAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "student",
        "date",
        "school_class",
        "wing",
        "term",
        "present_periods",
        "absent_periods",
        "runaway_periods",
        "excused_periods",
        "late_minutes",
        "early_minutes",
        "daily_absent_unexcused",
        "daily_excused",
        "daily_excused_partial",
        "locked",
    )
    list_filter = ("date", "school_class", "wing", "term", "daily_absent_unexcused")
    search_fields = ("student__full_name", "school_class__name")
    autocomplete_fields = ("student", "school_class", "wing", "term")


@admin.register(AssessmentPackage)
class AssessmentPackageAdmin(admin.ModelAdmin):
    list_display = ("id", "term", "name", "type", "start_date", "end_date")
    list_filter = ("term", "type")
    search_fields = ("name", "term__name")


@admin.register(SchoolHoliday)
class SchoolHolidayAdmin(admin.ModelAdmin):
    list_display = ("id", "title", "start", "end")
    list_filter = ("start", "end")
    search_fields = ("title",)


@admin.register(ExitEvent)
class ExitEventAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "student",
        "classroom",
        "date",
        "period_number",
        "reason",
        "started_at",
        "returned_at",
        "duration_seconds",
        "started_by",
        "returned_by",
    )
    list_filter = (
        "date",
        "reason",
        "classroom",
        ("returned_at", admin.EmptyFieldListFilter),
    )
    search_fields = (
        "student__full_name",
        "student__sid",
        "classroom__name",
        "note",
    )
    autocomplete_fields = (
        "student",
        "classroom",
        "started_by",
        "returned_by",
        "attendance_record",
    )
    readonly_fields = ("started_at", "duration_seconds")
    list_select_related = ("student", "classroom", "started_by", "returned_by")
    date_hierarchy = "date"

    def get_queryset(self, request):
        # Optimize queries
        qs = super().get_queryset(request)
        return qs.select_related("student", "classroom", "started_by", "returned_by")

    # Custom display for duration in human-readable format
    def get_duration_display(self, obj):
        if obj.duration_seconds:
            hours = obj.duration_seconds // 3600
            minutes = (obj.duration_seconds % 3600) // 60
            seconds = obj.duration_seconds % 60
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        return "—"

    get_duration_display.short_description = "المدة"
