from django.core.management import BaseCommand
from openpyxl import load_workbook
from school.models import Student, Class


class Command(BaseCommand):
    help = "Import students from Excel (supports multiple sheets)"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str)

    def handle(self, *args, **opts):
        wb = load_workbook(opts["file"], data_only=True)
        total = 0
        sheets_processed = 0
        for ws in wb.worksheets:
            sheets_processed += 1
            # Read header row (row 1)
            header_row = [
                str(c).strip() if c is not None else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            ]

            def norm(s):
                return str(s).strip().lower().replace(" ", "").replace("_", "")

            header_map = {norm(v): idx for idx, v in enumerate(header_row) if v}

            def col(*names):
                for n in names:
                    k = norm(n)
                    if k in header_map:
                        return header_map[k]
                return None

            idx_sid = col(
                "sid", "student_id", "national_no", "nationalid", "nationalnumber"
            )
            idx_name = col("full_name", "fullname", "student_name", "name")
            idx_classname = col("class_name", "classname", "class")
            idx_grade = col("grade", "gradelevel")
            idx_section = col("section", "sec")
            idx_dob = col("dob", "date_of_birth", "dateofbirth", "birthdate")
            idx_nationality = col("nationality", "country")
            idx_age = col("age")

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                sid = (
                    str(row[idx_sid]).strip()
                    if (idx_sid is not None and row[idx_sid])
                    else None
                )
                full_name = (
                    str(row[idx_name]).strip()
                    if (idx_name is not None and row[idx_name])
                    else None
                )
                class_name = (
                    str(row[idx_classname]).strip()
                    if (idx_classname is not None and row[idx_classname])
                    else None
                )
                grade_val = (
                    row[idx_grade]
                    if (idx_grade is not None and idx_grade < len(row))
                    else None
                )
                section_val = (
                    row[idx_section]
                    if (idx_section is not None and idx_section < len(row))
                    else None
                )
                dob_val = (
                    row[idx_dob]
                    if (idx_dob is not None and idx_dob < len(row))
                    else None
                )

                if not sid:
                    continue

                # Determine class name
                if not class_name:
                    g = None
                    try:
                        g = (
                            int(grade_val)
                            if grade_val is not None and str(grade_val).strip() != ""
                            else None
                        )
                    except Exception:
                        g = None
                    s = str(section_val).strip() if section_val is not None else ""
                    if g is not None:
                        class_name = f"{g}-{s or 'A'}"
                    else:
                        class_name = str(ws.title).strip() or "A"

                # Parse grade number from class_name if possible
                defaults = {"grade": 1, "section": "A"}
                parts = str(class_name).split("-")
                try:
                    if len(parts) >= 1 and str(parts[0]).isdigit():
                        defaults["grade"] = int(parts[0])
                    if len(parts) >= 2 and parts[1]:
                        defaults["section"] = str(parts[1])
                except Exception:
                    pass

                cls, _ = Class.objects.get_or_create(
                    name=class_name,
                    defaults=defaults,
                )

                # Prepare dob if a datetime/date string
                from datetime import date, datetime

                dob_clean = None
                if dob_val:
                    if isinstance(dob_val, (date, datetime)):
                        dob_clean = (
                            dob_val.date() if isinstance(dob_val, datetime) else dob_val
                        )
                    else:
                        try:
                            # Try parse as string DD/MM/YYYY or similar
                            import datetime as _dt

                            dob_clean = _dt.datetime.fromisoformat(str(dob_val)).date()
                        except Exception:
                            dob_clean = None

                # Prepare nationality and age
                nationality_val = None
                if idx_nationality is not None and idx_nationality < len(row):
                    nationality_cell = row[idx_nationality]
                    if nationality_cell is not None:
                        nationality_val = str(nationality_cell).strip()
                age_val = None
                if idx_age is not None and idx_age < len(row):
                    try:
                        age_cell = row[idx_age]
                        if age_cell is not None and str(age_cell).strip() != "":
                            age_val = int(float(age_cell))
                    except Exception:
                        age_val = None

                Student.objects.update_or_create(
                    sid=str(sid),
                    defaults={
                        "full_name": (full_name or ""),
                        "class_fk": cls,
                        "dob": dob_clean,
                        "nationality": nationality_val or "",
                        "age": age_val,
                    },
                )
                total += 1
        self.stdout.write(
            self.style.SUCCESS(
                f"Imported/updated: {total} rows across {sheets_processed} sheet(s)"
            )
        )
