from django.core.management import BaseCommand
from openpyxl import load_workbook

from ...models import Staff


class Command(BaseCommand):
    help = "Import staff from Excel (supports multiple sheets)"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str)

    def handle(self, *args, **opts):
        wb = load_workbook(opts["file"], data_only=True)
        total = 0
        sheets_processed = 0

        def norm(s):
            return str(s).strip().lower().replace(" ", "").replace("_", "") if s is not None else ""

        # Known header tokens (normalized) to detect the header row robustly
        expected_any = {
            # English
            "full_name",
            "fullname",
            "name",
            "staff_name",
            "employee_name",
            "role",
            "job_title",
            "jobtitle",
            "title",
            "national_no",
            "nationalno",
            "nationalid",
            "idnumber",
            "sid",
            "nid",
            "job_no",
            "jobno",
            "employee_no",
            "employeeno",
            "empno",
            "email",
            "mail",
            "phone_no",
            "phoneno",
            "mobile",
            "phone",
            "mobile_no",
            # Arabic (normalized without spaces/underscores)
            "الاسم",
            "اسم",
            "اسمالموظف",
            "اسمالموظفة",
            "اسمالمعلم",
            "اسمالمعلّم",
            "الدور",
            "الصفة",
            "المجموعة",
            "المسمىالوظيفي",
            "المسمىالوظيفى",
            "الوظيفة",
            "المهنة",
            "رقمالهوية",
            "رقمالهويةالوطنية",
            "الرقمالقومي",
            "السجل",
            "رقماسجل",
            "رقمالاقامة",
            "الهوية",
            "رقموالوظيفة",
            "رقموالوظيفي",
            "الرقمالوظيفي",
            "البريد",
            "البريدالالكتروني",
            "البريدالإلكتروني",
            "إيميل",
            "الجوال",
            "الهاتف",
            "رقمالجوال",
            "رقمالهاتف",
            "رقمتواصل",
        }

        for ws in wb.worksheets:
            sheets_processed += 1

            # Find header row within first 10 rows having any expected token
            header_row_idx = None
            header_cells = None
            for r in range(1, 11):
                try:
                    cells = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
                except StopIteration:
                    break
                vals = [c if c is not None else "" for c in cells]
                if any(norm(v) in expected_any for v in vals):
                    header_row_idx = r
                    header_cells = vals
                    break
            if header_row_idx is None:
                # Fallback to row 1 if nothing matched
                try:
                    header_cells = [
                        c if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    ]
                    header_row_idx = 1
                except StopIteration:
                    continue

            header_map = {norm(v): i for i, v in enumerate(header_cells) if v}

            def col(*names):
                for n in names:
                    k = norm(n)
                    if k in header_map:
                        return header_map[k]
                return None

            # Column indices with English + Arabic aliases
            idx_full_name = col(
                "full_name",
                "fullname",
                "name",
                "staff_name",
                "employee_name",
                "stuff_name",
                "stuffname",
                "stuff _name",
                "الاسم",
                "اسم",
                "اسمالموظف",
                "اسمالموظفة",
                "اسمالمعلم",
                "اسمالمعلّم",
            )
            idx_role = col("role", "group", "الدور", "الصفة", "المجموعة", "الوظيفة")
            idx_job_title = col(
                "job_title",
                "jobtitle",
                "title",
                "المسمى الوظيفي",
                "المسمىالوظيفي",
                "المسمىالوظيفى",
                "الوظيفة",
                "المهنة",
            )
            idx_national_no = col(
                "national_no",
                "nationalno",
                "nationalid",
                "idnumber",
                "sid",
                "nid",
                "رقم الهوية",
                "رقمالهوية",
                "رقمالهويةالوطنية",
                "الرقمالقومي",
                "السجل",
                "رقماسجل",
                "رقمالاقامة",
                "الهوية",
            )
            idx_job_no = col(
                "job_no",
                "jobno",
                "employee_no",
                "employeeno",
                "empno",
                "رقم الوظيفة",
                "رقموالوظيفة",
                "رقموالوظيفي",
                "الرقمالوظيفي",
            )
            idx_email = col(
                "email",
                "mail",
                "البريد",
                "البريد الالكتروني",
                "البريدالإلكتروني",
                "إيميل",
            )
            idx_phone = col(
                "phone_no",
                "phoneno",
                "mobile",
                "phone",
                "mobile_no",
                "الجوال",
                "الهاتف",
                "رقم الجوال",
                "رقمالجوال",
                "رقمالهاتف",
                "رقمتواصل",
            )

            for i, row in enumerate(
                ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
                start=header_row_idx + 1,
            ):
                if not row:
                    continue

                def val_at(idx):
                    if idx is None:
                        return None
                    if idx >= len(row):
                        return None
                    return row[idx]

                # Build full_name.
                # Supports composing from first/last names (Arabic/English) if header uses variants.
                full_name_cell = val_at(idx_full_name)
                full_name = str(full_name_cell).strip() if full_name_cell else None
                if not full_name:
                    # Try composing from first/last names if headers are unconventional
                    first_idx = col("first_name", "firstname", "الاسمالاول", "الاسم الأول")
                    last_idx = col(
                        "last_name",
                        "lastname",
                        "اسم_العائلة",
                        "الاسم الاخير",
                        "الاسم الأخير",
                    )
                    fn = str(val_at(first_idx)).strip() if val_at(first_idx) else ""
                    ln = str(val_at(last_idx)).strip() if val_at(last_idx) else ""
                    composed = (fn + " " + ln).strip()
                    full_name = composed if composed else None
                if not full_name:
                    continue

                role_cell = val_at(idx_role)
                role_val = str(role_cell).strip().lower() if role_cell else ""
                # Infer role from job title when role column missing
                job_title_cell = val_at(idx_job_title)
                job_title_val = str(job_title_cell).strip() if job_title_cell else ""
                if not role_val:
                    jt_norm = job_title_val.lower()
                    if "teach" in jt_norm or "معلم" in jt_norm or "معلّم" in jt_norm:
                        role_val = "teacher"
                    elif "admin" in jt_norm or "مشرف" in jt_norm or "رئيس" in jt_norm or "مدير" in jt_norm:
                        role_val = "admin"
                    else:
                        role_val = "staff"

                national_no_cell = val_at(idx_national_no)
                national_no_val = str(national_no_cell).strip() if national_no_cell not in (None, "") else ""
                job_no_cell = val_at(idx_job_no)
                job_no_val = str(job_no_cell).strip() if job_no_cell not in (None, "") else ""
                email_cell = val_at(idx_email)
                email_val = str(email_cell).strip() if email_cell else ""
                phone_cell = val_at(idx_phone)
                phone_val = str(phone_cell).strip() if phone_cell not in (None, "") else ""

                # Use national_no as natural key if present; otherwise job_no+name
                lookup = (
                    {"national_no": national_no_val}
                    if national_no_val
                    else {"job_no": job_no_val, "full_name": full_name}
                )

                Staff.objects.update_or_create(
                    **lookup,
                    defaults={
                        "full_name": full_name,
                        "role": (role_val if role_val in {"teacher", "admin", "staff"} else "staff"),
                        "national_no": national_no_val,
                        "job_title": job_title_val,
                        "job_no": job_no_val,
                        "email": email_val,
                        "phone_no": phone_val,
                    },
                )
                total += 1

        self.stdout.write(
            self.style.SUCCESS(f"Imported/updated: {total} staff rows across {sheets_processed} sheet(s)")
        )
