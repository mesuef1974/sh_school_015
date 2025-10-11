import re
from typing import Dict, Any
from openpyxl import load_workbook
from django.db import transaction
from school.models import Class, Staff, Subject, TeachingAssignment, ClassSubject


def normalize_ar_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("ـ", "")
    s = re.sub(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]", "", s)
    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ٱ": "ا",
        "ى": "ي",
        "ئ": "ي",
        "ؤ": "و",
        "ة": "ه",
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    s = re.sub(r"[^\w\s\u0600-\u06FF]", "", s)
    trans_digits = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    s = s.translate(trans_digits)
    return s


def import_teacher_loads(file_obj, *, dry_run: bool = False) -> Dict[str, Any]:
    """Parse Excel and create/update teaching assignments.

    Returns a summary dict with counts. When dry_run=True, all DB changes are
    rolled back to a savepoint at the end.
    """
    wb = load_workbook(filename=file_obj, data_only=True)

    # Build teacher index (normalized)
    teacher_index = {}
    for t in Staff.objects.all():
        key = normalize_ar_text(t.full_name).replace(" ", "").lower()
        if key:
            teacher_index.setdefault(key, t)

    # Build class index by (grade, section) and by name
    classes = Class.objects.all()
    class_by_grade_section = {(c.grade, (c.section or "").strip()): c for c in classes}
    class_by_name = {normalize_ar_text(c.name): c for c in classes}

    # Subject index
    subject_by_norm = {normalize_ar_text(s.name_ar): s for s in Subject.objects.all()}

    summary = {
        "subjects_created": 0,
        "classsubjects_created": 0,
        "assignments_created": 0,
        "assignments_updated": 0,
        "skipped_no_teacher": 0,
        "skipped_no_class": 0,
        "skipped_no_subject": 0,
        "worksheets": 0,
        "rows_processed": 0,
    }

    sid = transaction.savepoint()
    try:
        for ws in wb.worksheets:
            summary["worksheets"] += 1
            # Try detect headers positions
            header_row = None
            col_map = {
                "teacher": None,
                "grade": None,
                "section": None,
                "subject": None,
                "weekly": None,
            }
            for r in range(1, 11):
                try:
                    vals = [
                        c if c is not None else ""
                        for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
                    ]
                except StopIteration:
                    break
                normed = [normalize_ar_text(v).replace(" ", "").lower() for v in vals]
                tokens = {
                    "teacher": {
                        "teachername",
                        "teacher",
                        "اسمالمعلم",
                        "اسمالمعلّم",
                        "المعلم",
                    },
                    "grade": {"grade", "الصف", "المرحلة", "الصفالدراسي"},
                    "section": {"section", "الشعبة", "الفصل"},
                    "subject": {"class", "subject", "المادة", "المواد"},
                    "weekly": {
                        "no.classesw",
                        "weekly",
                        "weeklyclasses",
                        "الحصصالاسبوعية",
                        "الحصصالأسبوعية",
                        "حصصالاسبوع",
                    },
                }
                for i, v in enumerate(normed):
                    for k, opts in tokens.items():
                        if v in opts and col_map[k] is None:
                            col_map[k] = i
                if any(col_map.values()):
                    header_row = r
                    break
            if header_row is None:
                header_row = 1
                col_map = {
                    "teacher": 0,
                    "grade": 1,
                    "section": 2,
                    "subject": 3,
                    "weekly": 4,
                }

            current_teacher = None
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row:
                    continue
                summary["rows_processed"] += 1
                tv = (
                    row[col_map["teacher"]]
                    if col_map["teacher"] is not None and col_map["teacher"] < len(row)
                    else None
                )
                teacher_display = str(tv).strip() if tv not in (None, "") else None
                if teacher_display:
                    current_teacher = teacher_display
                if not current_teacher:
                    summary["skipped_no_teacher"] += 1
                    continue

                gv = (
                    row[col_map["grade"]]
                    if col_map["grade"] is not None and col_map["grade"] < len(row)
                    else None
                )
                sv = (
                    row[col_map["section"]]
                    if col_map["section"] is not None and col_map["section"] < len(row)
                    else None
                )
                subjv = (
                    row[col_map["subject"]]
                    if col_map["subject"] is not None and col_map["subject"] < len(row)
                    else None
                )
                wv = (
                    row[col_map["weekly"]]
                    if col_map["weekly"] is not None and col_map["weekly"] < len(row)
                    else None
                )

                t_key = normalize_ar_text(current_teacher).replace(" ", "").lower()
                teacher = teacher_index.get(t_key)
                if not teacher:
                    summary["skipped_no_teacher"] += 1
                    continue

                def to_int_safe(x):
                    try:
                        return int(str(x).strip().split()[0])
                    except Exception:
                        return None

                grade = to_int_safe(gv)
                section = None
                if sv not in (None, ""):
                    section = re.sub(r"[^0-9A-Za-zأ-ي]", "", str(sv)).strip()
                subj_norm = normalize_ar_text(subjv)
                weekly = to_int_safe(wv) or 0

                classroom = None
                if grade is not None:
                    classroom = class_by_grade_section.get((grade, section or ""))
                if classroom is None and subj_norm:
                    if grade is not None and section is not None:
                        key = f"{grade}{section}"
                        for cname, cobj in class_by_name.items():
                            if key in re.sub(r"\D", "", cname):
                                classroom = cobj
                                break
                if classroom is None:
                    summary["skipped_no_class"] += 1
                    continue

                if subj_norm:
                    subj_obj = subject_by_norm.get(subj_norm)
                    if not subj_obj:
                        subj_obj = Subject.objects.create(name_ar=subjv)
                        subject_by_norm[subj_norm] = subj_obj
                        summary["subjects_created"] += 1
                else:
                    summary["skipped_no_subject"] += 1
                    continue

                _, cs_created = ClassSubject.objects.get_or_create(
                    classroom=classroom, subject=subj_obj
                )
                if cs_created:
                    summary["classsubjects_created"] += 1

                _, created = TeachingAssignment.objects.update_or_create(
                    teacher=teacher,
                    classroom=classroom,
                    subject=subj_obj,
                    defaults={"no_classes_weekly": weekly or subj_obj.weekly_default or 0},
                )
                if created:
                    summary["assignments_created"] += 1
                else:
                    summary["assignments_updated"] += 1

        if dry_run:
            transaction.savepoint_rollback(sid)
        else:
            transaction.savepoint_commit(sid)
    except Exception:
        transaction.savepoint_rollback(sid)
        raise

    return summary
