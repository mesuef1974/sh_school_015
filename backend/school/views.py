import io
import os
import re
from datetime import date as _date
from datetime import time as _time

from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db import IntegrityError, connection, transaction
from django.db.models import Case, Count, F, IntegerField, Min, OuterRef, Q, Subquery, Sum, Value, When
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.http import Http404, HttpResponse, HttpResponseNotModified, JsonResponse, StreamingHttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook
from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import ModelViewSet

from .models import (
    AttendanceRecord,
    Class,
    ClassSubject,
    Staff,
    Student,
    Subject,
    TeachingAssignment,
    Term,
    TimetableEntry,
    Wing,
)
from .serializers import (
    ClassSerializer,
    ClassSubjectSerializer,
    StaffSerializer,
    StudentSerializer,
    SubjectSerializer,
    TeachingAssignmentSerializer,
)
from .services.imports import import_teacher_loads
from .services.ocr_table_parser import parse_ocr_raw_to_csv
from .services.timetable_import import import_timetable_csv
from .services.timetable_ocr import try_extract_csv_from_image, try_extract_csv_from_pdf

try:
    from backend.common.day_utils import iso_to_school_dow
except Exception:
    from common.day_utils import iso_to_school_dow  # type: ignore


class ClassViewSet(ModelViewSet):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    permission_classes = [IsAuthenticated]


class StudentViewSet(ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]


class StaffViewSet(ModelViewSet):
    queryset = Staff.objects.all()
    serializer_class = StaffSerializer
    permission_classes = [IsAuthenticated]


class SubjectViewSet(ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated]


class ClassSubjectViewSet(ModelViewSet):
    queryset = ClassSubject.objects.select_related("classroom", "subject").all()
    serializer_class = ClassSubjectSerializer
    permission_classes = [IsAuthenticated]


class TeachingAssignmentViewSet(ModelViewSet):
    queryset = TeachingAssignment.objects.select_related("teacher", "classroom", "subject").all()
    serializer_class = TeachingAssignmentSerializer
    permission_classes = [IsAuthenticated]


class TeachingAssignmentForm(forms.ModelForm):
    class Meta:
        model = TeachingAssignment
        fields = ["teacher", "classroom", "subject", "no_classes_weekly", "notes"]


# Generic CRUD forms for editable tables
class ClassForm(forms.ModelForm):
    class Meta:
        model = Class
        # Include common fields; adjust as per actual model fields in project
        fields = [f.name for f in Class._meta.fields if f.editable and f.name != Class._meta.pk.name]


class StudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = [f.name for f in Student._meta.fields if f.editable and f.name != Student._meta.pk.name]


class StaffForm(forms.ModelForm):
    class Meta:
        model = Staff
        fields = [f.name for f in Staff._meta.fields if f.editable and f.name != Staff._meta.pk.name]


class SubjectForm(forms.ModelForm):
    class Meta:
        model = Subject
        fields = [f.name for f in Subject._meta.fields if f.editable and f.name != Subject._meta.pk.name]


class ClassSubjectForm(forms.ModelForm):
    class Meta:
        model = ClassSubject
        fields = [f.name for f in ClassSubject._meta.fields if f.editable and f.name != ClassSubject._meta.pk.name]


class ExcelUploadForm(forms.Form):
    file = forms.FileField(label="ملف Excel للأنصبة (schasual.xlsx)")


def _subject_category_order_expr(field_lookup: str):
    """Return a CASE expression that maps subject Arabic names to an order index.
    field_lookup is a dotted path to Subject.name_ar, e.g.,
    - "subject__name_ar" when annotating TeachingAssignment
    - "assignments__subject__name_ar" when annotating Staff over related assignments
    """
    # Ordered specialties from top to bottom per requirements
    patterns = [
        (1, ["تربية اسلامية", "التربية الاسلامية", "إسلامية", "اسلامية"]),
        (2, ["لغة عربية", "اللغة العربية", "عربي", "العربية"]),
        (3, ["رياضيات", "رياضيات"]),
        (4, ["أحياء", "احياء", "علوم عامة"]),
        (5, ["فيزياء"]),
        (6, ["كيمياء"]),
        (7, ["علوم"]),  # ملاحظة: "علوم عامة" التقطت في فئة (4)
        (8, ["لغة انجليزية", "اللغة الانجليزية", "انجليزي", "إنجليزي", "إنجليزية"]),
        (9, ["جغرافيا", "تاريخ", "اجتماع", "اجتماعية", "علوم اجتماعية"]),
        (10, ["ادارة اعمال", "إدارة أعمال", "ادارة", "إدارة"]),
        (11, ["مهارات حياتية", "حياتية", "مهارات"]),
        (
            12,
            [
                "تكنولوجيا",
                "تكنولوجيا معلومات",
                "تقنية معلومات",
                "علوم حاسب",
                "علوم حاسوب",
                "حاسوب",
                "حاسب",
                "كمبيوتر",
                "معلومات",
            ],
        ),
        (13, ["فنون بصرية", "فنون", "رسم"]),
        (14, ["تربية بدنية", "تربية رياضية", "بدنية"]),
    ]
    whens = []
    for order, keys in patterns:
        for k in keys:
            whens.append(When(**{f"{field_lookup}__icontains": k}, then=Value(order)))
    return Case(*whens, default=Value(999), output_field=IntegerField())


def _job_title_category_order_expr(field_lookup: str):
    """Map Staff.job_title Arabic text to the same category indices used for subjects."""
    patterns = [
        (1, ["اسلام", "التربية الاسلامية", "دين", "شرعية"]),
        (2, ["لغة عربية", "عربي", "العربية"]),
        (3, ["رياضيات", "رياضي"]),
        (4, ["أحياء", "احياء", "علوم عامة", "علوم الحياة"]),
        (5, ["فيزياء", "فيزي"]),
        (6, ["كيمياء", "كيميائ"]),
        (7, ["علوم"]),
        (8, ["لغة انجليزية", "انجليزي", "إنجليزي", "English"]),
        (9, ["جغرافيا", "تاريخ", "اجتماع", "اجتماعية", "علوم اجتماعية"]),
        (10, ["ادارة", "إدارة", "اعمال", "أعمال", "إدارة أعمال"]),
        (11, ["مهارات حياتية", "حياتية", "مهارات"]),
        (12, ["تكنولوجيا", "تقنية معلومات", "حاسوب", "حاسب", "كمبيوتر", "معلومات"]),
        (13, ["فنون", "رسم", "فنون بصرية"]),
        (14, ["تربية بدنية", "رياضة", "بدنية", "رياضية"]),
    ]
    whens = []
    for order, keys in patterns:
        for k in keys:
            whens.append(When(**{f"{field_lookup}__icontains": k}, then=Value(order)))
    return Case(*whens, default=Value(999), output_field=IntegerField())


@login_required
def teacher_loads_dashboard(request):
    from django.conf import settings
    from django.db.models import Count, F, Q, Sum
    from django.utils import timezone

    # Handle manual add
    if request.method == "POST" and request.POST.get("action") == "add":
        form = TeachingAssignmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse("teacher_loads_dashboard"))
    else:
        form = TeachingAssignmentForm()

    # Handle Excel upload (gated by feature flag)
    imports_disabled = getattr(settings, "DISABLE_IMPORTS", False)
    if request.method == "POST" and request.POST.get("action") == "upload_async":
        # Enqueue background import via RQ
        if imports_disabled:
            messages.warning(request, "تم تعطيل الاستيراد مؤقتًا بواسطة إعداد النظام.")
            return redirect(reverse("teacher_loads_dashboard"))
        up_form = ExcelUploadForm(request.POST, request.FILES)
        if up_form.is_valid():
            from django_rq import get_queue

            from .tasks.jobs_rq import enqueue_import_teacher_loads

            dry_run_flag = (request.POST.get("dry_run") or request.GET.get("dry_run") or "").strip() in (
                "1",
                "true",
                "True",
            )
            file_obj = up_form.cleaned_data["file"]
            file_bytes = file_obj.read()
            q = get_queue("default")
            job = enqueue_import_teacher_loads(q, file_bytes, dry_run=dry_run_flag)
            messages.info(
                request,
                ("[Dry-run] " if dry_run_flag else "") + f"تم إرسال مهمة الاستيراد إلى الخلفية. رقم المهمة: {job.id}",
            )
            return redirect(reverse("job_status", kwargs={"job_id": job.id}))
    elif request.method == "POST" and request.POST.get("action") == "upload":
        if imports_disabled:
            messages.warning(request, "تم تعطيل الاستيراد مؤقتًا بواسطة إعداد النظام.")
            return redirect(reverse("teacher_loads_dashboard"))
        up_form = ExcelUploadForm(request.POST, request.FILES)
        if up_form.is_valid():
            dry_run_flag = (request.POST.get("dry_run") or request.GET.get("dry_run") or "").strip() in (
                "1",
                "true",
                "True",
            )
            summary = import_teacher_loads(up_form.cleaned_data["file"], dry_run=dry_run_flag)
            messages.info(
                request,
                (
                    ("[Dry-run] " if dry_run_flag else "")
                    + (
                        f"تمت معالجة {summary['rows_processed']} صفاً عبر {summary['worksheets']} ورقة. "
                        f"مواد جديدة: {summary['subjects_created']}, مواد-صفوف جديدة: {summary['classsubjects_created']}, "
                        f"أنصبة مضافة: {summary['assignments_created']}, أنصبة محدّثة: {summary['assignments_updated']}. "
                        f"تخطّي: معلّم غير معروف={summary['skipped_no_teacher']}, صف غير معروف={summary['skipped_no_class']}, مادة غير معروفة={summary['skipped_no_subject']}."
                    )
                ),
            )
            return redirect(reverse("teacher_loads_dashboard"))
    elif request.method == "POST" and request.POST.get("action") == "upload_default":
        if imports_disabled:
            messages.warning(request, "تم تعطيل الاستيراد مؤقتًا بواسطة إعداد النظام.")
            return redirect(reverse("teacher_loads_dashboard"))
        # Import from the known default path on the server
        from pathlib import Path

        repo_root = Path(__file__).resolve().parents[2]
        default_path = repo_root / "DOC" / "school_DATA" / "schasual.xlsx"
        dry_run_flag = (request.POST.get("dry_run") or request.GET.get("dry_run") or "").strip() in (
            "1",
            "true",
            "True",
        )
        try:
            with open(default_path, "rb") as f:
                summary = import_teacher_loads(f, dry_run=dry_run_flag)
                messages.info(
                    request,
                    (
                        ("[Dry-run] " if dry_run_flag else "")
                        + (
                            f"تمت معالجة {summary['rows_processed']} صفاً عبر {summary['worksheets']} ورقة. "
                            f"مواد جديدة: {summary['subjects_created']}, مواد-صفوف جديدة: {summary['classsubjects_created']}, "
                            f"أنصبة مضافة: {summary['assignments_created']}, أنصبة محدّثة: {summary['assignments_updated']}. "
                            f"تخطّي: معلّم غير معروف={summary['skipped_no_teacher']}, صف غير معروف={summary['skipped_no_class']}, مادة غير معروفة={summary['skipped_no_subject']}."
                        )
                    ),
                )
        except FileNotFoundError:
            messages.error(request, f"لم يتم العثور على الملف: {default_path}")
        return redirect(reverse("teacher_loads_dashboard"))
    else:
        up_form = ExcelUploadForm()

    # Filters and search
    teacher_q = request.GET.get("teacher")
    grade_q = request.GET.get("grade")
    section_q = request.GET.get("section")
    subject_q = request.GET.get("subject")
    q_text = (request.GET.get("q") or "").strip()

    # Build base query string for sorting/links
    from urllib.parse import urlencode

    params = {}
    if teacher_q:
        params["teacher"] = teacher_q
    if grade_q:
        params["grade"] = grade_q
    if section_q:
        params["section"] = section_q
    if subject_q:
        params["subject"] = subject_q
    if q_text:
        params["q"] = q_text
    base_query = urlencode(params, doseq=True)
    base_prefix = "?" + base_query + ("&" if base_query else "")

    qs = TeachingAssignment.objects.select_related("teacher", "classroom", "subject").all()
    if teacher_q:
        qs = qs.filter(teacher_id=teacher_q)
    if grade_q:
        qs = qs.filter(classroom__grade=grade_q)
    if section_q:
        qs = qs.filter(classroom__section=section_q)
    if subject_q:
        qs = qs.filter(subject_id=subject_q)
    if q_text:
        qs = qs.filter(
            Q(teacher__full_name__icontains=q_text)
            | Q(subject__name_ar__icontains=q_text)
            | Q(classroom__name__icontains=q_text)
            | Q(classroom__section__icontains=q_text)
            | Q(classroom__grade__icontains=q_text)
        )

    # Sorting (safe mapping)
    sort_key = (request.GET.get("sort") or "").strip()
    sort_dir = (request.GET.get("dir") or "asc").lower()
    allowed = {
        "teacher": "teacher__full_name",
        "grade": "classroom__grade",
        "section": "classroom__section",
        "subject": "subject__name_ar",
        "weekly": "no_classes_weekly",
    }
    if sort_key in allowed:
        field = allowed[sort_key]
        if sort_dir == "desc":
            qs = qs.order_by(f"-{field}")
        else:
            qs = qs.order_by(field)
    else:
        # Default order: by teacher specialty category, then teacher name, then class
        teacher_min_category = Subquery(
            Staff.objects.filter(pk=OuterRef("teacher_id"))
            .annotate(
                life_skills_total=Sum(
                    "assignments__no_classes_weekly",
                    filter=Q(assignments__subject__name_ar__icontains="مهارات"),
                ),
                non_ls_min=Min(
                    Case(
                        When(
                            assignments__subject__name_ar__icontains="مهارات",
                            then=Value(999),
                        ),
                        default=_subject_category_order_expr("assignments__subject__name_ar"),
                        output_field=IntegerField(),
                    )
                ),
                job_title_cat=_job_title_category_order_expr("job_title"),
                all_min=Min(_subject_category_order_expr("assignments__subject__name_ar")),
            )
            .annotate(
                # Detect presence of Social Studies/History/Geography subjects for this teacher
                has_social=Sum(
                    Case(
                        When(
                            Q(assignments__subject__name_ar__icontains="جغرافيا")
                            | Q(assignments__subject__name_ar__icontains="تاريخ")
                            | Q(assignments__subject__name_ar__icontains="اجتماع")
                            | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                            then=Value(1),
                        ),
                        default=Value(0),
                        output_field=IntegerField(),
                    )
                ),
                override_cat=Case(
                    When(full_name__icontains="وجدي", then=Value(14)),
                    When(full_name__icontains="منير", then=Value(12)),
                    When(
                        Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                        then=Value(12),
                    ),
                    When(full_name__icontains="السيد محمد", then=Value(9)),
                    When(full_name__icontains="محمد عدوان", then=Value(9)),
                    # Specific placement for Social Studies/History/Geography teachers
                    When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                    When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                    When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                    When(
                        full_name__icontains="علاء محمد عبد الهادي القضاه",
                        then=Value(9),
                    ),
                    default=Value(None),
                    output_field=IntegerField(),
                ),
                # Prefer Social Studies group when any such subject exists, regardless of Science also present
                prefer_social=Case(
                    When(has_social__gt=0, then=Value(9)),
                    default=Value(None),
                    output_field=IntegerField(),
                ),
                min_category_smart=Case(
                    When(
                        Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                        then=Coalesce(
                            F("override_cat"),
                            F("prefer_social"),
                            F("job_title_cat"),
                            F("all_min"),
                            Value(999),
                        ),
                    ),
                    default=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("all_min"),
                        F("non_ls_min"),
                        F("job_title_cat"),
                        Value(999),
                    ),
                    output_field=IntegerField(),
                ),
            )
            .values("min_category_smart")[:1]
        )
        qs = qs.annotate(teacher_min_category=teacher_min_category).order_by(
            "teacher_min_category",
            "teacher__full_name",
            "classroom__grade",
            "classroom__section",
        )

    # KPI stats
    today = timezone.localdate()
    total_assignments = TeachingAssignment.objects.count()
    todays_assignments = TeachingAssignment.objects.filter(created_at__date=today).count()

    # Determine latest update timestamp across assignments
    latest_dt = (
        TeachingAssignment.objects.order_by("-updated_at").values_list("updated_at", flat=True).first()
        or TeachingAssignment.objects.order_by("-created_at").values_list("created_at", flat=True).first()
    )
    latest_date = latest_dt.date() if latest_dt else None

    # Top teachers by weekly load (sum no_classes_weekly) for latest snapshot if available
    ta_qs_for_charts = TeachingAssignment.objects.all()

    agg = ta_qs_for_charts.values("teacher__full_name").annotate(total=Sum("no_classes_weekly")).order_by("-total")[:10]
    top_teachers_labels = [a["teacher__full_name"] for a in agg]
    top_teachers_values = [a["total"] or 0 for a in agg]

    # Coverage: ClassSubjects that have at least one assignment vs gaps with none
    cs = ClassSubject.objects.values("classroom_id", "subject_id", "classroom__grade").annotate(
        assignments_count=Count(
            "classroom__assignments",
            filter=Q(classroom__assignments__subject=F("subject_id")),
        )
    )
    covered = 0
    gaps = 0
    # Per-grade coverage tallies
    per_grade = {}
    for x in cs:
        cnt = x.get("assignments_count") or 0
        grade = x.get("classroom__grade")
        if grade is None:
            continue
        if grade not in per_grade:
            per_grade[grade] = {"covered": 0, "gaps": 0}
        if cnt > 0:
            covered += 1
            per_grade[grade]["covered"] += 1
        else:
            gaps += 1
            per_grade[grade]["gaps"] += 1

    # Prepare arrays for stacked bar chart by grade
    grades_sorted = sorted(per_grade.keys())
    coverage_by_grade_labels = [str(g) for g in grades_sorted]
    coverage_by_grade_covered = [per_grade[g]["covered"] for g in grades_sorted]
    coverage_by_grade_gaps = [per_grade[g]["gaps"] for g in grades_sorted]

    # Compute KPI: total weekly assignments (sum of no_classes_weekly)
    weekly_total = TeachingAssignment.objects.aggregate(total=Sum("no_classes_weekly")).get("total") or 0
    gaps_count = gaps  # reuse computed total gaps from coverage aggregation

    # Build a concrete list of gaps (ClassSubject with no assignment) for display box
    gaps_qs = (
        ClassSubject.objects.select_related("classroom", "subject")
        .annotate(
            assignments_count=Count(
                "classroom__assignments",
                filter=Q(classroom__assignments__subject=F("subject_id")),
            )
        )
        .filter(assignments_count=0)
        .order_by("classroom__grade", "classroom__section", "subject__name_ar", "id")
    )
    # Keep it compact: cap at first 100 rows for the dashboard box
    gaps_items = [
        {
            "grade": cs.classroom.grade,
            "section": (cs.classroom.section or ""),
            "subject": cs.subject.name_ar,
        }
        for cs in gaps_qs[:100]
    ]

    # Compute next sort direction per column to avoid template boolean expressions
    _current_sort = sort_key or ""
    _current_dir = sort_dir or "asc"

    def _next_dir(col: str) -> str:
        return "desc" if (_current_sort == col and _current_dir == "asc") else "asc"

    # Prepare ordered teachers list for filters (teachers without assignments go last)
    teachers_ordered = (
        Staff.objects.all()
        .annotate(
            life_skills_total=Sum(
                "assignments__no_classes_weekly",
                filter=Q(assignments__subject__name_ar__icontains="مهارات"),
            ),
            non_ls_min=Min(
                Case(
                    When(
                        assignments__subject__name_ar__icontains="مهارات",
                        then=Value(999),
                    ),
                    default=_subject_category_order_expr("assignments__subject__name_ar"),
                    output_field=IntegerField(),
                )
            ),
            job_title_cat=_job_title_category_order_expr("job_title"),
            all_min=Min(_subject_category_order_expr("assignments__subject__name_ar")),
        )
        .annotate(
            # Detect Social Studies subjects for preference
            has_social=Sum(
                Case(
                    When(
                        Q(assignments__subject__name_ar__icontains="جغرافيا")
                        | Q(assignments__subject__name_ar__icontains="تاريخ")
                        | Q(assignments__subject__name_ar__icontains="اجتماع")
                        | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                        then=Value(1),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            override_cat=Case(
                When(full_name__icontains="وجدي", then=Value(14)),
                When(full_name__icontains="منير", then=Value(12)),
                When(
                    Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                    then=Value(12),
                ),
                When(full_name__icontains="السيد محمد", then=Value(9)),
                When(full_name__icontains="محمد عدوان", then=Value(9)),
                # Specific placement for Social Studies/History/Geography teachers
                When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                When(full_name__icontains="علاء محمد عبد الهادي القضاه", then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            prefer_social=Case(
                When(has_social__gt=0, then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            min_category=Case(
                When(
                    Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                    then=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("job_title_cat"),
                        F("all_min"),
                        Value(999),
                    ),
                ),
                default=Coalesce(
                    F("override_cat"),
                    F("prefer_social"),
                    F("all_min"),
                    F("non_ls_min"),
                    F("job_title_cat"),
                    Value(999),
                ),
                output_field=IntegerField(),
            ),
        )
        .order_by("min_category", "full_name", "id")
    )

    context = {
        "form": form,
        "upload_form": up_form,
        "assignments": qs,
        "teachers": teachers_ordered,
        "classes": Class.objects.all(),
        "subjects": Subject.objects.all(),
        "grades": sorted({c.grade for c in Class.objects.all()}),
        "sections": sorted({(c.section or "") for c in Class.objects.all()}),
        "title": "لوحة الأنصبة — إدخال ورفع Excel",
        "show_imports": not imports_disabled,
        # current filters/search/sort
        "q_text": q_text,
        "teacher_q": teacher_q,
        "grade_q": grade_q,
        "section_q": section_q,
        "subject_q": subject_q,
        "sort_key": sort_key,
        "sort_dir": sort_dir,
        "base_query": base_query,
        "base_prefix": base_prefix,
        # precomputed next sort dirs
        "ndir_teacher": _next_dir("teacher"),
        "ndir_grade": _next_dir("grade"),
        "ndir_section": _next_dir("section"),
        "ndir_subject": _next_dir("subject"),
        "ndir_weekly": _next_dir("weekly"),
        # KPIs
        "kpi_total_assignments": total_assignments,
        "kpi_todays_assignments": todays_assignments,
        "kpi_weekly_total": weekly_total,
        "kpi_gaps_count": gaps_count,
        # latest snapshot info for charts
        "charts_latest_date": latest_date,
        # Charts data
        "chart_top_teachers_labels": top_teachers_labels,
        "chart_top_teachers_values": top_teachers_values,
        "chart_coverage": {"covered": covered, "gaps": gaps},
        "chart_coverage_by_grade_labels": coverage_by_grade_labels,
        "chart_coverage_by_grade_covered": coverage_by_grade_covered,
        "chart_coverage_by_grade_gaps": coverage_by_grade_gaps,
        # gaps box data
        "gaps_items": gaps_items,
    }
    # HTMX partial update for assignments table
    if request.headers.get("HX-Request"):
        return render(request, "school/_assignments_table.html", context)
    return render(request, "school/teacher_loads_dashboard.html", context)


@login_required
def teacher_class_matrix(request):
    """Matrix: rows=teachers, columns=classes, cells=sum of weekly lessons teacher gives to that class.
    Professional RTL table with sticky headers. Now supports inline editing for single-subject cells
    and hover tooltips listing subjects.
    """
    # Handle inline update POST (AJAX)
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            teacher_id = int(request.POST.get("teacher_id"))
            class_id = int(request.POST.get("class_id"))
            new_val = int(request.POST.get("value"))
        except Exception:
            return JsonResponse({"ok": False, "error": "بيانات غير صالحة"}, status=400)

        # If subject_id explicitly provided (from modal), create/update that assignment directly
        subj_id_str = request.POST.get("subject_id")
        if subj_id_str:
            try:
                subject_id = int(subj_id_str)
            except Exception:
                return JsonResponse({"ok": False, "error": "المادة غير صالحة"}, status=400)
            # Ensure the subject is linked to the class; create ClassSubject if missing
            try:
                ClassSubject.objects.get_or_create(classroom_id=class_id, subject_id=subject_id)
                ta, created = TeachingAssignment.objects.get_or_create(
                    teacher_id=teacher_id,
                    classroom_id=class_id,
                    subject_id=subject_id,
                    defaults={"no_classes_weekly": max(0, new_val)},
                )
                if not created:
                    ta.no_classes_weekly = max(0, new_val)
                    ta.save(update_fields=["no_classes_weekly", "updated_at"])  # type: ignore
            except Exception as e:
                return JsonResponse({"ok": False, "error": str(e)}, status=400)
            return JsonResponse({"ok": True, "value": ta.no_classes_weekly})

        # Update if exactly one TeachingAssignment exists, otherwise allow create when none and class has a single subject
        qs = TeachingAssignment.objects.select_related("subject").filter(teacher_id=teacher_id, classroom_id=class_id)
        count = qs.count()
        if count == 1:
            ta = qs.first()
            ta.no_classes_weekly = max(0, new_val)
            try:
                ta.save(update_fields=["no_classes_weekly", "updated_at"])  # type: ignore
            except Exception as e:
                return JsonResponse({"ok": False, "error": str(e)}, status=400)
            return JsonResponse({"ok": True, "value": ta.no_classes_weekly})
        elif count == 0:
            # If the class has exactly one subject configured, create a new assignment for that subject
            from django.db.models import Count

            single_cs = (
                ClassSubject.objects.filter(classroom_id=class_id)
                .values("classroom_id")
                .annotate(cnt=Count("id"))
                .filter(cnt=1)
            )
            if single_cs.exists():
                cs = ClassSubject.objects.filter(classroom_id=class_id).select_related("subject").first()
                try:
                    ta, created = TeachingAssignment.objects.get_or_create(
                        teacher_id=teacher_id,
                        classroom_id=class_id,
                        subject=cs.subject,
                        defaults={"no_classes_weekly": max(0, new_val)},
                    )
                    if not created:
                        ta.no_classes_weekly = max(0, new_val)
                        ta.save(update_fields=["no_classes_weekly", "updated_at"])  # type: ignore
                except Exception as e:
                    return JsonResponse({"ok": False, "error": str(e)}, status=400)
                return JsonResponse({"ok": True, "value": ta.no_classes_weekly})
            else:
                return JsonResponse(
                    {
                        "ok": False,
                        "error": "غير قابل للتعديل: لا توجد مادة محدّدة لهذا الصف أو توجد عدة مواد. الرجاء تحديد المادة أولاً.",
                        "count": count,
                    },
                    status=400,
                )
        else:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "غير قابل للتعديل: توجد عدة مواد لهذا المعلّم مع هذا الصف.",
                    "count": count,
                },
                status=400,
            )

    # Get classes (columns) sorted by grade then section then id
    classes = Class.objects.all().order_by("grade", "section", "id")
    class_list = list(classes)

    # Build aggregated columns by label (grade-section) to avoid duplicate columns for same label
    def _cls_label(c: Class) -> str:  # type: ignore[name-defined]
        sec = (c.section or "").strip()
        return f"{c.grade}-{sec}" if sec else str(c.grade)

    columns_map = {}
    columns_order = []
    for c in class_list:
        label = _cls_label(c)
        if label not in columns_map:
            columns_map[label] = {"label": label, "ids": [c.id], "title": c.name}
            columns_order.append(label)
        else:
            columns_map[label]["ids"].append(c.id)
            # If multiple classes share the same label, concatenate names for tooltip/title clarity
            columns_map[label]["title"] = columns_map[label]["title"] + ", " + c.name

    columns = [columns_map[lbl] for lbl in columns_order]

    # Get teachers (rows) who have at least one assignment
    # Order teachers by requested specialty categories then by name
    cat_expr_for_staff = _subject_category_order_expr("assignments__subject__name_ar")
    teachers = (
        Staff.objects.filter(assignments__isnull=False)
        .distinct()
        .annotate(
            life_skills_total=Sum(
                "assignments__no_classes_weekly",
                filter=Q(assignments__subject__name_ar__icontains="مهارات"),
            ),
            non_ls_min=Min(
                Case(
                    When(
                        assignments__subject__name_ar__icontains="مهارات",
                        then=Value(999),
                    ),
                    default=cat_expr_for_staff,
                    output_field=IntegerField(),
                )
            ),
            job_title_cat=_job_title_category_order_expr("job_title"),
            all_min=Min(cat_expr_for_staff),
        )
        .annotate(
            has_social=Sum(
                Case(
                    When(
                        Q(assignments__subject__name_ar__icontains="جغرافيا")
                        | Q(assignments__subject__name_ar__icontains="تاريخ")
                        | Q(assignments__subject__name_ar__icontains="اجتماع")
                        | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                        then=Value(1),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            override_cat=Case(
                When(full_name__icontains="وجدي", then=Value(14)),
                When(full_name__icontains="منير", then=Value(12)),
                When(
                    Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                    then=Value(12),
                ),
                When(full_name__icontains="السيد محمد", then=Value(9)),
                When(full_name__icontains="محمد عدوان", then=Value(9)),
                # Specific placement for Social Studies/History/Geography teachers
                When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                When(full_name__icontains="علاء محمد عبد الهادي القضاه", then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            prefer_social=Case(
                When(has_social__gt=0, then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            min_category=Case(
                When(
                    Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                    then=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("job_title_cat"),
                        F("all_min"),
                        Value(999),
                    ),
                ),
                default=Coalesce(
                    F("override_cat"),
                    F("prefer_social"),
                    F("all_min"),
                    F("non_ls_min"),
                    F("job_title_cat"),
                    Value(999),
                ),
                output_field=IntegerField(),
            ),
        )
        .order_by("min_category", "full_name", "id")
    )
    teacher_list = list(teachers)

    # Build mapping (teacher_id, class_id) -> total weekly lessons

    agg = TeachingAssignment.objects.values("teacher_id", "classroom_id").annotate(total=Sum("no_classes_weekly"))
    matrix = {(a["teacher_id"], a["classroom_id"]): (a["total"] or 0) for a in agg}

    # Preload subjects per cell
    subs_map = {}
    for ta in (
        TeachingAssignment.objects.select_related("subject")
        .filter(
            teacher_id__in=[t.id for t in teacher_list],
            classroom_id__in=[c.id for c in class_list],
        )
        .order_by("subject__name_ar")
    ):
        key = (ta.teacher_id, ta.classroom_id)
        subs_map.setdefault(key, []).append(
            {
                "id": ta.id,
                "subject": ta.subject.name_ar,
                "weekly": int(ta.no_classes_weekly),
            }
        )

    # Determine subject count per class to allow safe creation when a class has a single subject
    from django.db.models import Count as _Count

    cnt_map = {
        r["classroom_id"]: r["cnt"] for r in ClassSubject.objects.values("classroom_id").annotate(cnt=_Count("id"))
    }
    single_subj_name_map = {}

    # Precompute rows (cells ordered by aggregated columns) and row totals
    cells_by_teacher = {}
    row_totals = {}
    for t in teacher_list:
        row = []
        s = 0
        for col in columns:
            ids = col["ids"]
            # Sum across all classes sharing the same label
            v = 0
            combined_subs = []
            for cid in ids:
                v += int(matrix.get((t.id, cid), 0) or 0)
                sl = subs_map.get((t.id, cid), [])
                if sl:
                    # Tag with class id for clarity when multiple
                    for x in sl:
                        combined_subs.append(
                            {
                                "subject": x["subject"],
                                "weekly": x["weekly"],
                                "class_id": cid,
                            }
                        )
            # Editable only when a single underlying class and the same rules as before
            if len(ids) == 1:
                cid = ids[0]
                sublist = subs_map.get((t.id, cid), [])
                cs_cnt = cnt_map.get(cid, 0)
                editable = (len(sublist) == 1) or (len(sublist) == 0 and cs_cnt == 1)
                # Build tooltip label; if empty but class has one subject, show that subject name
                if sublist:
                    subjects_label = ", ".join([f"{x['subject']} ({x['weekly']})" for x in sublist])
                else:
                    subjects_label = ""
                    if cs_cnt == 1:
                        if cid not in single_subj_name_map:
                            cs_obj = ClassSubject.objects.filter(classroom_id=cid).select_related("subject").first()
                            single_subj_name_map[cid] = cs_obj.subject.name_ar if cs_obj and cs_obj.subject else ""
                        subjects_label = single_subj_name_map.get(cid, "")
                class_id_for_cell = cid
            else:
                editable = False
                # Build combined tooltip
                if combined_subs:
                    subjects_label = "; ".join([f"{x['subject']} ({x['weekly']})" for x in combined_subs])
                else:
                    subjects_label = ""
                class_id_for_cell = 0  # aggregated; not directly editable
            row.append(
                {
                    "value": v,
                    "teacher_id": t.id,
                    "class_id": class_id_for_cell,
                    "subjects": subjects_label,
                    "editable": editable,
                }
            )
            s += v
        cells_by_teacher[t.id] = row
        row_totals[t.id] = s

    # Column totals in columns order
    col_totals_list = []
    for idx, _col in enumerate(columns):
        s = 0
        for t in teacher_list:
            s += cells_by_teacher[t.id][idx]["value"]
        col_totals_list.append(s)

    grand_total = sum(row_totals.values())

    # Build rows for easier templating
    rows = []
    for t in teacher_list:
        rows.append(
            {
                "teacher": t,
                "cells": cells_by_teacher[t.id],
                "total": row_totals[t.id],
            }
        )

    # Coverage gaps across ClassSubject pairs: count pairs with no assignment
    from django.db.models import Count
    from django.db.models import Q as _Q

    cs = ClassSubject.objects.values("classroom_id", "subject_id").annotate(
        assignments_count=Count(
            "classroom__assignments",
            filter=_Q(classroom__assignments__subject=F("subject_id")),
        )
    )
    gaps_count = 0
    for x in cs:
        if not (x.get("assignments_count") or 0):
            gaps_count += 1

    # Provide all subjects for the modal selector (include all subjects taught in the school)
    subjects_all = [{"id": s.id, "name": s.name_ar} for s in Subject.objects.all().order_by("name_ar")]

    context = {
        "title": "مصفوفة الأنصبة — معلّم × صف",
        "rows": rows,
        "columns": columns,
        "col_totals_list": col_totals_list,
        "grand_total": grand_total,
        "gaps_count": gaps_count,
        "subjects_all": subjects_all,
    }

    return render(request, "school/teacher_class_matrix.html", context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def teacher_week_matrix(request):
    """Interactive weekly timetable matrix.
    - For admins/managers: full matrix with edit capabilities (via APIs).
    - For teachers: show ONLY their own timetable and in read-only mode (no edit/move UI).
    """
    from .models import Term, TimetableEntry  # lazy import to avoid circulars

    term = Term.objects.filter(is_current=True).first()
    if not term:
        messages.error(request, "لا يوجد فصل دراسي حالي.")
        return redirect("data_overview")

    # Detect staff and role
    staff = Staff.objects.filter(user=request.user).first()
    is_teacher = bool(staff and staff.role == "teacher")
    # Determine management capability via roles
    try:
        role_names = set(request.user.groups.values_list("name", flat=True))
    except Exception:
        role_names = set()
    can_manage_timetable = bool(
        getattr(request.user, "is_superuser", False)
        or role_names.intersection({"principal", "academic_deputy", "timetable_manager"})
    )

    # Teachers to show (requirement: teachers can view the full general timetable read-only)
    teacher_ids_with_entries = set(TimetableEntry.objects.filter(term=term).values_list("teacher_id", flat=True))
    teacher_qs = (
        Staff.objects.filter(Q(assignments__isnull=False) | Q(id__in=teacher_ids_with_entries))
        .annotate(
            life_skills_total=Sum(
                "assignments__no_classes_weekly",
                filter=Q(assignments__subject__name_ar__icontains="مهارات"),
            ),
            non_ls_min=Min(
                Case(
                    When(
                        assignments__subject__name_ar__icontains="مهارات",
                        then=Value(999),
                    ),
                    default=_subject_category_order_expr("assignments__subject__name_ar"),
                    output_field=IntegerField(),
                )
            ),
            job_title_cat=_job_title_category_order_expr("job_title"),
            all_min=Min(_subject_category_order_expr("assignments__subject__name_ar")),
        )
        .annotate(
            has_social=Sum(
                Case(
                    When(
                        Q(assignments__subject__name_ar__icontains="جغرافيا")
                        | Q(assignments__subject__name_ar__icontains="تاريخ")
                        | Q(assignments__subject__name_ar__icontains="اجتماع")
                        | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                        then=Value(1),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            override_cat=Case(
                When(full_name__icontains="وجدي", then=Value(14)),
                When(full_name__icontains="منير", then=Value(12)),
                When(
                    Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                    then=Value(12),
                ),
                When(full_name__icontains="السيد محمد", then=Value(9)),
                When(full_name__icontains="محمد عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                When(full_name__icontains="علاء محمد عبد الهادي القضاه", then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            prefer_social=Case(
                When(has_social__gt=0, then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            min_category=Case(
                When(
                    Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                    then=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("job_title_cat"),
                        F("all_min"),
                        Value(999),
                    ),
                ),
                default=Coalesce(
                    F("override_cat"),
                    F("prefer_social"),
                    F("all_min"),
                    F("non_ls_min"),
                    F("job_title_cat"),
                    Value(999),
                ),
                output_field=IntegerField(),
            ),
        )
        .distinct()
        .order_by("min_category", "full_name", "id")
    )
    teachers = list(teacher_qs)

    # Build grid per teacher: [day 1..5][period 1..7] -> entry or None
    DAYS = [1, 2, 3, 4, 5]
    PERIODS = [1, 2, 3, 4, 5, 6, 7]

    # Load entries for term (limited to current teacher when is_teacher)
    e_qs = TimetableEntry.objects.filter(term=term)
    if is_teacher:
        e_qs = e_qs.filter(teacher_id=staff.id)
    entries = list(e_qs.select_related("classroom", "subject", "teacher"))

    # Index by teacher/day/period
    grid = {t.id: {d: {p: None for p in PERIODS} for d in DAYS} for t in teachers}
    for e in entries:
        if e.teacher_id in grid and e.day_of_week in DAYS and e.period_number in PERIODS:
            grid[e.teacher_id][e.day_of_week][e.period_number] = e

    # Teacher assignments options: per teacher list of (class_id, subject_id, label)
    asg_opts = {}
    for t in teachers:
        opts = (
            TeachingAssignment.objects.filter(teacher=t)
            .select_related("classroom", "subject")
            .order_by("classroom__grade", "classroom__section", "subject__name_ar")
        )
        asg_opts[t.id] = [
            {
                "class_id": a.classroom_id,
                "subject_id": a.subject_id,
                "label": f"{a.classroom.name} – {a.subject.name_ar}",
            }
            for a in opts
        ]

    # Colors for subjects (stable per subject id)
    all_subjects = Subject.objects.all().order_by("name_ar")
    subject_colors = {}
    for s in all_subjects:
        h = (s.id * 47) % 360
        subject_colors[s.id] = f"hsl({h}, 60%, 85%)"

    context = {
        "title": "الجدول الأسبوعي",
        "teachers": teachers,
        "grid": grid,
        "DAYS": DAYS,
        "DAYS_RTL": list(reversed(DAYS)),
        "DAY_NAMES": {
            1: "الأحد",
            2: "الاثنين",
            3: "الثلاثاء",
            4: "الأربعاء",
            5: "الخميس",
        },
        # Present periods in ascending order in UI (1 → 7)
        "PERIODS_DESC": PERIODS,
        "PERIODS": PERIODS,
        "subject_colors": subject_colors,
        "assign_opts": asg_opts,
        "term": term,
        "read_only": (not can_manage_timetable),
    }
    return render(request, "school/teacher_week_matrix.html", context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def teacher_week_compact(request):
    """Compact, print-friendly weekly timetable matching the provided sheet layout.
    Rows = teachers. Columns = grouped by day (Sun..Thu), each group has 7 period columns.
    Cells show the classroom label and are color-coded by subject. Admin-only, with add/move.
    """
    from .models import Term, TimetableEntry  # lazy import

    term = Term.objects.filter(is_current=True).first()
    if not term:
        messages.error(request, "لا يوجد فصل دراسي حالي.")
        return redirect("data_overview")

    # Teachers to show
    teacher_ids_with_entries = set(TimetableEntry.objects.filter(term=term).values_list("teacher_id", flat=True))
    teacher_qs = (
        Staff.objects.filter(Q(assignments__isnull=False) | Q(id__in=teacher_ids_with_entries))
        .exclude(full_name="يوسف يعقوب")
        .annotate(
            life_skills_total=Sum(
                "assignments__no_classes_weekly",
                filter=Q(assignments__subject__name_ar__icontains="مهارات"),
            ),
            non_ls_min=Min(
                Case(
                    When(
                        assignments__subject__name_ar__icontains="مهارات",
                        then=Value(999),
                    ),
                    default=_subject_category_order_expr("assignments__subject__name_ar"),
                    output_field=IntegerField(),
                )
            ),
            job_title_cat=_job_title_category_order_expr("job_title"),
            all_min=Min(_subject_category_order_expr("assignments__subject__name_ar")),
        )
        .annotate(
            has_social=Sum(
                Case(
                    When(
                        Q(assignments__subject__name_ar__icontains="جغرافيا")
                        | Q(assignments__subject__name_ar__icontains="تاريخ")
                        | Q(assignments__subject__name_ar__icontains="اجتماع")
                        | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                        then=Value(1),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            override_cat=Case(
                When(full_name__icontains="وجدي", then=Value(14)),
                When(full_name__icontains="منير", then=Value(12)),
                When(
                    Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                    then=Value(12),
                ),
                When(full_name__icontains="السيد محمد", then=Value(9)),
                When(full_name__icontains="محمد عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                When(full_name__icontains="علاء محمد عبد الهادي القضاه", then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            prefer_social=Case(
                When(has_social__gt=0, then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            min_category=Case(
                When(
                    Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                    then=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("job_title_cat"),
                        F("all_min"),
                        Value(999),
                    ),
                ),
                default=Coalesce(
                    F("override_cat"),
                    F("prefer_social"),
                    F("all_min"),
                    F("non_ls_min"),
                    F("job_title_cat"),
                    Value(999),
                ),
                output_field=IntegerField(),
            ),
        )
        .distinct()
        .order_by("min_category", "full_name", "id")
    )
    teachers = list(teacher_qs)

    # Grid: teacher -> (day,period) -> entry
    DAYS = [
        (1, "الأحد"),
        (2, "الاثنين"),
        (3, "الثلاثاء"),
        (4, "الأربعاء"),
        (5, "الخميس"),
    ]
    PERIODS = [1, 2, 3, 4, 5, 6, 7]

    entries = list(TimetableEntry.objects.filter(term=term).select_related("classroom", "subject", "teacher"))

    # Nested grid: teacher -> day -> period -> entry
    grid = {t.id: {d: {p: None for p in PERIODS} for d in [1, 2, 3, 4, 5]} for t in teachers}
    for e in entries:
        if e.teacher_id in grid and e.day_of_week in [1, 2, 3, 4, 5] and e.period_number in PERIODS:
            grid[e.teacher_id][e.day_of_week][e.period_number] = e

    # Assignment options for add modal (same as matrix view)
    asg_opts = {}
    for t in teachers:
        opts = (
            TeachingAssignment.objects.filter(teacher=t)
            .select_related("classroom", "subject")
            .order_by("classroom__grade", "classroom__section", "subject__name_ar")
        )
        asg_opts[t.id] = [
            {
                "class_id": a.classroom_id,
                "subject_id": a.subject_id,
                "label": f"{a.classroom.name} – {a.subject.name_ar}",
            }
            for a in opts
        ]

    # Subject colors
    all_subjects = Subject.objects.all().order_by("name_ar")
    subject_colors = {}
    for s in all_subjects:
        h = (s.id * 47) % 360
        subject_colors[s.id] = f"hsl({h}, 60%, 85%)"

    total_cols = 1 + (len(DAYS) * 7)
    context = {
        "title": "الجدول العام للمعلمين — عرض مضغوط",
        "teachers": teachers,
        "grid": grid,
        "DAYS": DAYS,
        "DAYS_RTL": list(reversed(DAYS)),
        # Period headers ascending (1..7) per new requirement
        "PERIODS_DESC": PERIODS,
        "PERIODS": PERIODS,
        "subject_colors": subject_colors,
        "assign_opts": asg_opts,
        "term": term,
        "total_cols": total_cols,
    }
    return render(request, "school/teacher_week_compact.html", context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def timetable_source_image(request):
    # Support multiple possible filenames per user uploads (prioritize Arabic legacy name)
    candidate_names = [
        "الجدول العام.jpg",
        "الجدول العام.jpeg",
        "الجدول العام.png",
        "time_table.png",
        "time_table.jpg",
        "time_table.jpeg",
    ]
    base_dir = os.path.abspath(os.path.join(settings.BASE_DIR, "..", "DOC", "school_DATA"))
    path = None
    for name in candidate_names:
        p = os.path.join(base_dir, name)
        if os.path.exists(p):
            path = p
            break
    if not path:
        raise Http404("الصورة غير موجودة")

    # Guess content type by extension
    import mimetypes

    content_type = mimetypes.guess_type(path)[0] or "image/jpeg"

    def file_iter():
        with open(path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                yield chunk

    return StreamingHttpResponse(file_iter(), content_type=content_type)


@login_required
@user_passes_test(lambda u: u.is_staff)
def timetable_source_pdf(request):
    # Support both the Arabic legacy name and the new file name (prioritize Arabic)
    candidate_names = [
        "الجدول العام.pdf",
        "time_table.pdf",
    ]
    base_dir = os.path.abspath(os.path.join(settings.BASE_DIR, "..", "DOC", "school_DATA"))
    path = None
    for name in candidate_names:
        p = os.path.join(base_dir, name)
        if os.path.exists(p):
            path = p
            break
    if not path:
        raise Http404("الملف غير موجود")

    def file_iter():
        with open(path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                yield chunk

    return StreamingHttpResponse(file_iter(), content_type="application/pdf")


class TimetableImageImportForm(forms.Form):
    csv_text = forms.CharField(
        label="بيانات CSV (انسخ من الجدول)",
        widget=forms.Textarea(
            attrs={
                "rows": 18,
                "dir": "rtl",
                "style": "font-family:monospace; direction: rtl; unicode-bidi: plaintext; text-align: start;",
            }
        ),
    )


@login_required
@user_passes_test(lambda u: u.is_staff)
def timetable_import_from_image(request):
    # Prefill from template if exists
    template_path = os.path.abspath(
        os.path.join(settings.BASE_DIR, "..", "DOC", "school_DATA", "timetable_template.csv")
    )
    template_text = "teacher,class,subject,day,period\n"  # fallback header
    try:
        if os.path.exists(template_path):
            with open(template_path, "r", encoding="utf-8-sig") as f:
                template_text = f.read()
    except Exception:
        pass

    if request.method == "POST":
        action = request.POST.get("action", "import")
        if action == "auto":
            # Auto-extract via OCR/PDF
            # existing logic below
            # Try to auto-extract CSV from PDF first, then image as fallback
            base_dir = os.path.abspath(os.path.join(settings.BASE_DIR, "..", "DOC", "school_DATA"))
            pdf_candidates = [
                "الجدول العام.pdf",
                "time_table.pdf",
            ]
            img_candidates = [
                "الجدول العام.jpg",
                "الجدول العام.jpeg",
                "الجدول العام.png",
                "time_table.png",
                "time_table.jpg",
                "time_table.jpeg",
            ]
            pdf_path = next(
                (os.path.join(base_dir, n) for n in pdf_candidates if os.path.exists(os.path.join(base_dir, n))),
                None,
            )
            img_path = next(
                (os.path.join(base_dir, n) for n in img_candidates if os.path.exists(os.path.join(base_dir, n))),
                None,
            )
            csv_text = ""
            warn = ""
            if pdf_path:
                csv_text, warn = try_extract_csv_from_pdf(pdf_path)
            if not csv_text and img_path:
                csv_text, warn = try_extract_csv_from_image(img_path)
            if csv_text:
                form = TimetableImageImportForm(initial={"csv_text": csv_text})
                messages.info(
                    request,
                    "تمت محاولة الاستخراج الآلي. يرجى المراجعة ثم الضغط على استيراد.",
                )
                if warn:
                    messages.warning(request, warn)
            else:
                form = TimetableImageImportForm(initial={"csv_text": template_text})
                messages.error(request, warn or "تعذر الاستخراج الآلي. الرجاء إدخال CSV يدويًا.")
        elif action == "parse_raw":
            raw_text = request.POST.get("raw_text", "")
            if not raw_text.strip():
                form = TimetableImageImportForm(initial={"csv_text": template_text})
                messages.error(request, "الرجاء لصق النص الخام الملتقط أولاً.")
            else:
                try:
                    csv_text, warns = parse_ocr_raw_to_csv(raw_text)
                    form = TimetableImageImportForm(initial={"csv_text": csv_text})
                    messages.success(request, "تم تحويل النص الخام إلى CSV — راجع ثم اضغط استيراد.")
                    for w in warns[:20]:
                        messages.warning(request, w)
                    if len(warns) > 20:
                        messages.info(request, f"إشعارات إضافية: {len(warns)-20} …")
                except Exception as e:
                    form = TimetableImageImportForm(initial={"csv_text": template_text})
                    messages.error(request, f"تعذر التحويل: {e}")
        else:
            form = TimetableImageImportForm(request.POST)
            if form.is_valid():
                csv_text = form.cleaned_data["csv_text"] or ""
                if not csv_text.strip():
                    messages.error(request, "الرجاء لصق بيانات CSV أولاً.")
                else:
                    try:
                        summary = import_timetable_csv(io.StringIO(csv_text), dry_run=False)
                        messages.success(
                            request,
                            f"تم الاستيراد: مضاف {summary.get('created', 0)}, مستبدل {summary.get('replaced', 0)}, متخطى {summary.get('skipped', 0)}",
                        )
                        if summary.get("errors"):
                            messages.warning(request, "\n".join(summary["errors"]))
                        return redirect("teacher_week_compact")
                    except Exception as e:
                        messages.error(request, f"فشل الاستيراد: {e}")
    else:
        form = TimetableImageImportForm(initial={"csv_text": template_text})

    # Detect which source files are available to show helpful captions
    base_dir = os.path.abspath(os.path.join(settings.BASE_DIR, "..", "DOC", "school_DATA"))
    img_candidates = [
        "الجدول العام.jpg",
        "الجدول العام.jpeg",
        "الجدول العام.png",
        "time_table.png",
        "time_table.jpg",
        "time_table.jpeg",
    ]
    pdf_candidates = [
        "الجدول العام.pdf",
        "time_table.pdf",
    ]
    source_image_name = next((n for n in img_candidates if os.path.exists(os.path.join(base_dir, n))), None)
    source_pdf_name = next((n for n in pdf_candidates if os.path.exists(os.path.join(base_dir, n))), None)

    context = {
        "title": "استيراد الجدول من صورة/بي دي إف",
        "form": form,
        "source_image_name": source_image_name,
        "source_pdf_name": source_pdf_name,
    }
    return render(request, "school/timetable_import_from_image.html", context)


@login_required
@user_passes_test(lambda u: u.is_staff)
def api_timetable_add(request):
    """Add/replace a TimetableEntry for a teacher at (day, period).
    Admin-only (teachers are forbidden).
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST فقط"}, status=405)
    from .models import Term, TimetableEntry

    # Forbid teachers from editing the timetable
    staff = Staff.objects.filter(user=request.user).first()
    if staff and staff.role == "teacher":
        return JsonResponse({"ok": False, "error": "ممنوع للمعلمين"}, status=403)

    term = Term.objects.filter(is_current=True).first()
    if not term:
        return JsonResponse({"ok": False, "error": "لا يوجد فصل حالي"}, status=400)
    try:
        teacher_id = int(request.POST.get("teacher_id") or "0")
        classroom_id = int(request.POST.get("classroom_id") or "0")
        subject_id = int(request.POST.get("subject_id") or "0")
        day = int(request.POST.get("day") or "0")
        period = int(request.POST.get("period") or "0")
    except Exception:
        return JsonResponse({"ok": False, "error": "بيانات غير صالحة"}, status=400)
    if not (teacher_id and classroom_id and subject_id and day and period):
        return JsonResponse({"ok": False, "error": "حقول ناقصة"}, status=400)

    # Replace conflicts: any entry for same teacher or same classroom at (day,period)
    with transaction.atomic():
        TimetableEntry.objects.filter(term=term, day_of_week=day, period_number=period, teacher_id=teacher_id).delete()
        TimetableEntry.objects.filter(
            term=term, day_of_week=day, period_number=period, classroom_id=classroom_id
        ).delete()
        e = TimetableEntry.objects.create(
            classroom_id=classroom_id,
            subject_id=subject_id,
            teacher_id=teacher_id,
            day_of_week=day,
            period_number=period,
            term=term,
        )
    return JsonResponse(
        {
            "ok": True,
            "entry": {
                "id": e.id,
                "teacher_id": teacher_id,
                "classroom": e.classroom.name,
                "classroom_label": f"{e.classroom.grade}-{e.classroom.section}",
                "classroom_id": classroom_id,
                "subject": e.subject.name_ar,
                "subject_id": subject_id,
                "day": day,
                "period": period,
            },
        }
    )


@login_required
@user_passes_test(lambda u: u.is_staff)
def api_timetable_move(request):
    """Move/reassign an existing TimetableEntry to another slot (and optionally another teacher).
    Admin-only (teachers are forbidden).
    POST: entry_id, target_teacher_id (optional), day, period. Replaces conflicts at target.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST فقط"}, status=405)
    from .models import Term, TimetableEntry

    # Forbid teachers from editing the timetable
    staff = Staff.objects.filter(user=request.user).first()
    if staff and staff.role == "teacher":
        return JsonResponse({"ok": False, "error": "ممنوع للمعلمين"}, status=403)

    term = Term.objects.filter(is_current=True).first()
    if not term:
        return JsonResponse({"ok": False, "error": "لا يوجد فصل حالي"}, status=400)
    try:
        entry_id = int(request.POST.get("entry_id") or "0")
        day = int(request.POST.get("day") or "0")
        period = int(request.POST.get("period") or "0")
        target_teacher_id = request.POST.get("target_teacher_id")
        target_teacher_id = int(target_teacher_id) if target_teacher_id else None
    except Exception:
        return JsonResponse({"ok": False, "error": "بيانات غير صالحة"}, status=400)
    if not (entry_id and day and period):
        return JsonResponse({"ok": False, "error": "حقول ناقصة"}, status=400)

    try:
        e = TimetableEntry.objects.select_for_update().get(id=entry_id, term=term)
    except TimetableEntry.DoesNotExist:
        return JsonResponse({"ok": False, "error": "الحصة غير موجودة"}, status=404)

    with transaction.atomic():
        new_teacher_id = target_teacher_id or e.teacher_id
        # Replace conflicts at destination for same teacher or same class
        TimetableEntry.objects.filter(
            term=term, day_of_week=day, period_number=period, teacher_id=new_teacher_id
        ).exclude(id=e.id).delete()
        TimetableEntry.objects.filter(
            term=term,
            day_of_week=day,
            period_number=period,
            classroom_id=e.classroom_id,
        ).exclude(id=e.id).delete()
        # Apply move
        e.teacher_id = new_teacher_id
        e.day_of_week = day
        e.period_number = period
        e.save(update_fields=["teacher_id", "day_of_week", "period_number"])  # type: ignore

    return JsonResponse(
        {
            "ok": True,
            "entry": {
                "id": e.id,
                "teacher_id": e.teacher_id,
                "classroom": e.classroom.name,
                "classroom_label": f"{e.classroom.grade}-{e.classroom.section}",
                "classroom_id": e.classroom_id,
                "subject": e.subject.name_ar,
                "subject_id": e.subject_id,
                "day": e.day_of_week,
                "period": e.period_number,
            },
        }
    )


def _normalize_ar_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("ـ", "")
    s = re.sub(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]", "", s)
    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ٱ": "ا",
        "ى": "ي",
        "ئ": "ي",
        "ؤ": "و",
        "ة": "ه",
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    s = re.sub(r"[^\w\s\u0600-\u06FF]", "", s)
    trans_digits = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    s = s.translate(trans_digits)
    return s


def _import_excel_file(file_obj, dry_run: bool = False):
    wb = load_workbook(filename=file_obj, data_only=True)

    # Build teacher index (normalized)
    teacher_index = {}
    for t in Staff.objects.all():
        key = _normalize_ar_text(t.full_name).replace(" ", "").lower()
        if key:
            teacher_index.setdefault(key, t)

    # Build class index by (grade, section) and by name
    classes = Class.objects.all()
    class_by_grade_section = {(c.grade, (c.section or "").strip()): c for c in classes}
    class_by_name = {_normalize_ar_text(c.name): c for c in classes}

    # Subject index
    subject_by_norm = {_normalize_ar_text(s.name_ar): s for s in Subject.objects.all()}

    summary = {
        "subjects_created": 0,
        "classsubjects_created": 0,
        "assignments_created": 0,
        "assignments_updated": 0,
        "skipped_no_teacher": 0,
        "skipped_no_class": 0,
        "skipped_no_subject": 0,
        "worksheets": 0,
        "rows_processed": 0,
    }

    # Use a savepoint so we can rollback for dry_run without affecting outer transactions
    sid = transaction.savepoint()
    try:
        for ws in wb.worksheets:
            summary["worksheets"] += 1
            # Try detect headers positions
            header_row = None
            col_map = {
                "teacher": None,
                "grade": None,
                "section": None,
                "subject": None,
                "weekly": None,
            }
            for r in range(1, 11):
                try:
                    vals = [
                        c if c is not None else "" for c in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
                    ]
                except StopIteration:
                    break
                normed = [_normalize_ar_text(v).replace(" ", "").lower() for v in vals]
                tokens = {
                    "teacher": {
                        "teachername",
                        "teacher",
                        "اسمالمعلم",
                        "اسمالمعلّم",
                        "المعلم",
                    },
                    "grade": {"grade", "الصف", "المرحلة", "الصفالدراسي"},
                    "section": {"section", "الشعبة", "الفصل"},
                    "subject": {"class", "subject", "المادة", "المواد"},
                    "weekly": {
                        "no.classesw",
                        "weekly",
                        "weeklyclasses",
                        "الحصصالاسبوعية",
                        "الحصصالأسبوعية",
                        "حصصالاسبوع",
                    },
                }
                for i, v in enumerate(normed):
                    for k, opts in tokens.items():
                        if v in opts and col_map[k] is None:
                            col_map[k] = i
                if any(col_map.values()):
                    header_row = r
                    break
            if header_row is None:
                header_row = 1
                col_map = {
                    "teacher": 0,
                    "grade": 1,
                    "section": 2,
                    "subject": 3,
                    "weekly": 4,
                }

            current_teacher = None
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row:
                    continue
                summary["rows_processed"] += 1
                # Carry-forward teacher if grouped
                tv = (
                    row[col_map["teacher"]]
                    if col_map["teacher"] is not None and col_map["teacher"] < len(row)
                    else None
                )
                teacher_display = str(tv).strip() if tv not in (None, "") else None
                if teacher_display:
                    current_teacher = teacher_display
                if not current_teacher:
                    summary["skipped_no_teacher"] += 1
                    continue

                gv = row[col_map["grade"]] if col_map["grade"] is not None and col_map["grade"] < len(row) else None
                sv = (
                    row[col_map["section"]]
                    if col_map["section"] is not None and col_map["section"] < len(row)
                    else None
                )
                subjv = (
                    row[col_map["subject"]]
                    if col_map["subject"] is not None and col_map["subject"] < len(row)
                    else None
                )
                wv = row[col_map["weekly"]] if col_map["weekly"] is not None and col_map["weekly"] < len(row) else None

                # Normalize
                t_key = _normalize_ar_text(current_teacher).replace(" ", "").lower()
                teacher = teacher_index.get(t_key)
                if not teacher:
                    summary["skipped_no_teacher"] += 1
                    continue

                def to_int_safe(x):
                    try:
                        return int(str(x).strip().split()[0])
                    except Exception:
                        return None

                grade = to_int_safe(gv)
                section = None
                if sv not in (None, ""):
                    section = re.sub(r"[^0-9A-Za-zأ-ي]", "", str(sv)).strip()
                subj_norm = _normalize_ar_text(subjv)
                weekly = to_int_safe(wv) or 0

                # Map class
                classroom = None
                if grade is not None:
                    classroom = class_by_grade_section.get((grade, section or ""))
                if classroom is None and subj_norm:
                    if grade is not None and section is not None:
                        key = f"{grade}{section}"
                        for cname, cobj in class_by_name.items():
                            if key in re.sub(r"\D", "", cname):
                                classroom = cobj
                                break
                if classroom is None:
                    summary["skipped_no_class"] += 1
                    continue

                # Subject
                if subj_norm:
                    subj_obj = subject_by_norm.get(subj_norm)
                    if not subj_obj:
                        subj_obj = Subject.objects.create(name_ar=subjv)
                        subject_by_norm[subj_norm] = subj_obj
                        summary["subjects_created"] += 1
                else:
                    summary["skipped_no_subject"] += 1
                    continue

                # Ensure the subject is assigned to the class
                _, cs_created = ClassSubject.objects.get_or_create(classroom=classroom, subject=subj_obj)
                if cs_created:
                    summary["classsubjects_created"] += 1

                # Create/update teaching assignment
                _, created = TeachingAssignment.objects.update_or_create(
                    teacher=teacher,
                    classroom=classroom,
                    subject=subj_obj,
                    defaults={"no_classes_weekly": weekly or subj_obj.weekly_default or 0},
                )
                if created:
                    summary["assignments_created"] += 1
                else:
                    summary["assignments_updated"] += 1

        # Rollback changes if dry_run requested
        if dry_run:
            transaction.savepoint_rollback(sid)
        else:
            transaction.savepoint_commit(sid)
    except Exception:
        # In case of unexpected error, rollback to be safe
        transaction.savepoint_rollback(sid)
        raise

    return summary


def _staff_only(user):
    return user.is_authenticated and user.is_staff


@login_required
def portal_home(request):
    """بوابة أمامية مهنية بحسب الدور — مدير/معلم/مشرف طلابي..."""
    staff = Staff.objects.filter(user=request.user).first() if request.user.is_authenticated else None
    # إذا لم يكن مرتبطاً بكادر، أظهر لوحة عامة بسيطة
    if not staff:
        ctx = {"title": "الواجهة الرئيسية"}
        return render(request, "portal/dashboard_generic.html", ctx)

    role = staff.role or ""
    # توجيه حسب الدور
    if role == "teacher":
        return render(
            request,
            "portal/dashboard_teacher.html",
            {"title": "واجهة المعلم", "staff": staff},
        )
    elif role in {
        "school_principal",
        "academic_vice",
        "administrative_vice",
        "admin",
        "developer",
    }:
        return render(
            request,
            "portal/dashboard_manager.html",
            {"title": "واجهة الإدارة", "staff": staff},
        )
    elif role in {"student_observer", "supervisor", "coordinator", "academic_advisor"}:
        return render(
            request,
            "portal/dashboard_supervisor.html",
            {"title": "واجهة الإشراف الطلابي", "staff": staff},
        )
    else:
        # افتراضي: لوحة عامة
        return render(
            request,
            "portal/dashboard_generic.html",
            {"title": "لوحة المستخدم", "staff": staff},
        )


@login_required
def job_status(request, job_id: str):
    """Simple job status page for background imports."""
    try:
        import django_rq
        from rq.job import Job
    except Exception:
        raise Http404("RQ not available")

    conn = django_rq.get_connection("default")
    try:
        job = Job.fetch(job_id, connection=conn)
    except Exception:
        raise Http404("Job not found")

    state = job.get_status(refresh=True)
    summary = (job.meta or {}).get("summary") if hasattr(job, "meta") else None
    context = {
        "title": f"حالة المهمة {job_id}",
        "job_id": job_id,
        "state": state,
        "summary": summary,
    }
    return render(request, "school/job_status.html", context)


@user_passes_test(_staff_only)
@login_required
def data_overview(request):
    # Redirect teachers to the attendance entry page by default
    try:
        staff = Staff.objects.filter(user=request.user, role="teacher").first()
        if staff:
            return redirect("teacher_attendance_page")
    except Exception:
        pass

    # Denylist sensitive tables by prefix or exact names
    deny_exact = {"django_session"}
    deny_prefixes = ("auth_", "django_", "admin_")

    with connection.cursor() as cursor:
        introspection = connection.introspection
        tables = introspection.table_names()

    visible_tables = [t for t in tables if t not in deny_exact and not any(t.startswith(p) for p in deny_prefixes)]

    # Get row counts efficiently
    table_stats = []
    with connection.cursor() as cursor:
        for t in visible_tables:
            try:
                cursor.execute(f'SELECT COUNT(*) FROM "{t}"')
                cnt = cursor.fetchone()[0]
            except Exception:
                cnt = None
            table_stats.append({"name": t, "count": cnt})

    context = {
        "title": "معاينة البيانات — كل الجداول",
        "tables": table_stats,
    }
    return render(request, "data/overview.html", context)


@user_passes_test(_staff_only)
@login_required
def data_icons_catalog(request):
    """Professional catalog of platform icons with descriptions, tasks, and role permissions.
    Accessible under /data/icons/ and linked from /data/.
    """
    # Define catalog grouped by role with items: id, title, desc, route, roles, perms, icon (lordicon json url), bi (bootstrap icon class)
    catalog = [
        {
            "role": "المعلم",
            "role_key": "teacher",
            "color": "#1565c0",
            "items": [
                {
                    "id": "teacher_dashboard",
                    "title": "لوحة التحكم",
                    "desc": "نظرة عامة على مهام اليوم والتنبيهات.",
                    "route": "/",
                    "roles": ["teacher"],
                    "perms": [],
                    "icon": "https://cdn.lordicon.com/vspbqszr.json",
                    "bi": "bi-speedometer2",
                },
                {
                    "id": "classes",
                    "title": "صفوفي/أقسام",
                    "desc": "إدارة الطلبة والأنشطة لكل صف.",
                    "route": "/students",
                    "roles": ["teacher"],
                    "perms": ["academics.classes.read"],
                    "icon": "https://cdn.lordicon.com/gqdnbnwt.json",
                    "bi": "bi-people",
                },
                {
                    "id": "timetable",
                    "title": "جدولي",
                    "desc": "الجدول اليومي/الأسبوعي مع أوقات الحصص.",
                    "route": "/timetable/teacher",
                    "roles": ["teacher"],
                    "perms": ["academics.timetable.read"],
                    "icon": "https://cdn.lordicon.com/abwrkdvl.json",
                    "bi": "bi-calendar-week",
                },
                {
                    "id": "attendance",
                    "title": "تسجيل الحضور",
                    "desc": "تسجيل حضور/غياب الحصة الحالية.",
                    "route": "/attendance/teacher",
                    "roles": ["teacher"],
                    "perms": ["attendance.session.write"],
                    "icon": "https://cdn.lordicon.com/dfxesbyu.json",
                    "bi": "bi-clipboard-check",
                },
                {
                    "id": "attendance_history",
                    "title": "سجل الغياب",
                    "desc": "عرض وتصدير سجلات الغياب.",
                    "route": "/attendance/teacher/history",
                    "roles": ["teacher"],
                    "perms": ["attendance.history.read"],
                    "icon": "https://cdn.lordicon.com/wxnxiano.json",
                    "bi": "bi-clock-history",
                },
                {
                    "id": "announcements",
                    "title": "الإعلانات",
                    "desc": "نشر التنبيهات للطلبة أو الصفوف.",
                    "route": "/communications/announcements",
                    "roles": ["teacher"],
                    "perms": ["communications.announcements.write"],
                    "icon": "https://cdn.lordicon.com/megfpqpw.json",
                    "bi": "bi-megaphone",
                },
                {
                    "id": "assignments",
                    "title": "الواجبات",
                    "desc": "إنشاء وإدارة الواجبات.",
                    "route": "/teacher/assignments",
                    "roles": ["teacher"],
                    "perms": ["academics.assignments"],
                    "icon": "https://cdn.lordicon.com/ylvuooxd.json",
                    "bi": "bi-clipboard-data",
                },
                {
                    "id": "quizzes",
                    "title": "الاختبارات القصيرة",
                    "desc": "إنشاء الكويز وتصحيحه.",
                    "route": "/teacher/quizzes",
                    "roles": ["teacher"],
                    "perms": ["academics.quizzes"],
                    "icon": "https://cdn.lordicon.com/zzcjjxew.json",
                    "bi": "bi-ui-checks",
                },
                {
                    "id": "gradebook",
                    "title": "دفتر العلامات",
                    "desc": "إدخال الدرجات وتتبعها.",
                    "route": "/teacher/gradebook",
                    "roles": ["teacher"],
                    "perms": ["academics.grades"],
                    "icon": "https://cdn.lordicon.com/krmfspeu.json",
                    "bi": "bi-bar-chart",
                },
                {
                    "id": "lesson_plans",
                    "title": "تحضير الدروس",
                    "desc": "خطة الدرس والموارد.",
                    "route": "/teacher/lesson-plans",
                    "roles": ["teacher"],
                    "perms": ["academics.lesson_plans"],
                    "icon": "https://cdn.lordicon.com/kvdgkbcy.json",
                    "bi": "bi-journal-text",
                },
                {
                    "id": "resources",
                    "title": "الموارد",
                    "desc": "ملفات ومجلدات الدروس.",
                    "route": "/teacher/resources",
                    "roles": ["teacher"],
                    "perms": ["academics.resources"],
                    "icon": "https://cdn.lordicon.com/svbmmyue.json",
                    "bi": "bi-folder2-open",
                },
                {
                    "id": "messages",
                    "title": "الرسائل",
                    "desc": "التواصل مع الطلبة والأهالي.",
                    "route": "/messages",
                    "roles": ["teacher"],
                    "perms": ["communications.messages"],
                    "icon": "https://cdn.lordicon.com/kthelypq.json",
                    "bi": "bi-chat-dots",
                },
            ],
        },
        {
            "role": "مشرف الجناح",
            "role_key": "wing_supervisor",
            "color": "#8a1538",
            "items": [
                {
                    "id": "wing_dashboard",
                    "title": "لوحة الجناح",
                    "desc": "مؤشرات اليوم للحضور والانضباط.",
                    "route": "/wing/dashboard",
                    "roles": ["wing_supervisor"],
                    "perms": ["attendance.wing.read"],
                    "icon": "https://cdn.lordicon.com/gqzfzudq.json",
                    "bi": "bi-grid",
                },
                {
                    "id": "wing_daily",
                    "title": "الغياب اليومي",
                    "desc": "متابعة حضور اليوم لكل الصفوف.",
                    "route": "/wing/attendance/daily",
                    "roles": ["wing_supervisor"],
                    "perms": ["attendance.daily.read"],
                    "icon": "https://cdn.lordicon.com/akuwjdzh.json",
                    "bi": "bi-calendar2-day",
                },
                {
                    "id": "wing_missing",
                    "title": "حصص بلا إدخال",
                    "desc": "رصد الحصص التي لم تُسجل.",
                    "route": "/wing/attendance/missing",
                    "roles": ["wing_supervisor"],
                    "perms": ["attendance.missing.read"],
                    "icon": "https://cdn.lordicon.com/lznlxwtc.json",
                    "bi": "bi-exclamation-triangle-fill",
                },
                {
                    "id": "wing_exits",
                    "title": "أذونات الخروج",
                    "desc": "بدء/إيقاف وتتبع إذن الخروج.",
                    "route": "/wing/exits",
                    "roles": ["wing_supervisor"],
                    "perms": ["attendance.exits.manage"],
                    "icon": "https://cdn.lordicon.com/qhviklyi.json",
                    "bi": "bi-box-arrow-up-right",
                },
                {
                    "id": "wing_incidents",
                    "title": "البلاغات",
                    "desc": "تسجيل ومتابعة الحوادث السلوكية.",
                    "route": "/wing/incidents",
                    "roles": ["wing_supervisor"],
                    "perms": ["behavior.incidents.manage"],
                    "icon": "https://cdn.lordicon.com/iqaqnsks.json",
                    "bi": "bi-shield-exclamation",
                },
                {
                    "id": "wing_timetable",
                    "title": "جدول 7×7",
                    "desc": "شبكة أسبوعية 7×7 للحصص.",
                    "route": "/wing/timetable",
                    "roles": ["wing_supervisor"],
                    "perms": ["academics.timetable.read"],
                    "icon": "https://cdn.lordicon.com/abwrkdvl.json",
                    "bi": "bi-table",
                },
                {
                    "id": "wing_reports",
                    "title": "تقارير الجناح",
                    "desc": "تصدير ومؤشرات أسبوعية/شهرية.",
                    "route": "/wing/reports",
                    "roles": ["wing_supervisor"],
                    "perms": ["reports.wing"],
                    "icon": "https://cdn.lordicon.com/lywgqtim.json",
                    "bi": "bi-graph-up-arrow",
                },
            ],
        },
        {
            "role": "منسق المادة",
            "role_key": "subject_coordinator",
            "color": "#2e7d32",
            "items": [
                {
                    "id": "subject_dashboard",
                    "title": "لوحة المنسق",
                    "desc": "تحليل أداء المادة ومتابعة الخطط.",
                    "route": "/subject/dashboard",
                    "roles": ["subject_coordinator"],
                    "perms": ["analytics.subject.read"],
                    "icon": "https://cdn.lordicon.com/oezixobx.json",
                    "bi": "bi-graph-up",
                },
                {
                    "id": "question_bank",
                    "title": "بنك الأسئلة",
                    "desc": "إدارة بنك الأسئلة المشترك.",
                    "route": "/subject/qbank",
                    "roles": ["subject_coordinator"],
                    "perms": ["academics.question_bank"],
                    "icon": "https://cdn.lordicon.com/svbmmyue.json",
                    "bi": "bi-database",
                },
                {
                    "id": "rubrics",
                    "title": "قوالب التقييم (روبرك)",
                    "desc": "إنشاء ومراجعة الروبركات.",
                    "route": "/subject/rubrics",
                    "roles": ["subject_coordinator"],
                    "perms": ["academics.rubrics"],
                    "icon": "https://cdn.lordicon.com/qtqvorle.json",
                    "bi": "bi-list-check",
                },
            ],
        },
        {
            "role": "المدير",
            "role_key": "principal",
            "color": "#6a1b9a",
            "items": [
                {
                    "id": "principal_dashboard",
                    "title": "لوحة القيادة",
                    "desc": "مؤشرات KPI شاملة: حضور، تحصيل، انضباط.",
                    "route": "/principal/dashboard",
                    "roles": ["principal"],
                    "perms": ["analytics.kpi"],
                    "icon": "https://cdn.lordicon.com/vspbqszr.json",
                    "bi": "bi-speedometer",
                },
                {
                    "id": "reports_official",
                    "title": "التقارير الرسمية",
                    "desc": "تقارير وتصدير PDF/Excel.",
                    "route": "/reports/official",
                    "roles": ["principal"],
                    "perms": ["reports.official"],
                    "icon": "https://cdn.lordicon.com/kbtmbyzy.json",
                    "bi": "bi-filetype-pdf",
                },
                {
                    "id": "users_roles",
                    "title": "المستخدمون والأدوار",
                    "desc": "إدارة المستخدمين والصلاحيات.",
                    "route": "/admin/users",
                    "roles": ["principal"],
                    "perms": ["it.users"],
                    "icon": "https://cdn.lordicon.com/iltqorsz.json",
                    "bi": "bi-person-gear",
                },
            ],
        },
        {
            "role": "نائب أكاديمي",
            "role_key": "academic_deputy",
            "color": "#1e88e5",
            "items": [
                {
                    "id": "timetable_master",
                    "title": "الجداول الأكاديمية",
                    "desc": "إدارة الجداول وقوالب الحصص.",
                    "route": "/academic/dashboard",
                    "roles": ["academic_deputy"],
                    "perms": ["academics.timetable.manage"],
                    "icon": "https://cdn.lordicon.com/abwrkdvl.json",
                    "bi": "bi-calendar3",
                },
                {
                    "id": "assessments",
                    "title": "الاختبارات المركزية",
                    "desc": "ضبط الجداول والأسئلة الموحدة.",
                    "route": "/academic/assessments",
                    "roles": ["academic_deputy"],
                    "perms": ["academics.assessments.manage"],
                    "icon": "https://cdn.lordicon.com/zzcjjxew.json",
                    "bi": "bi-ui-checks-grid",
                },
                {
                    "id": "curricula",
                    "title": "المناهج والخطط",
                    "desc": "تخطيط وتوزيع المنهاج.",
                    "route": "/academic/curricula",
                    "roles": ["academic_deputy"],
                    "perms": ["academics.curricula"],
                    "icon": "https://cdn.lordicon.com/wxnxiano.json",
                    "bi": "bi-journal-richtext",
                },
            ],
        },
        {
            "role": "نائب إداري",
            "role_key": "admin_deputy",
            "color": "#ef6c00",
            "items": [
                {
                    "id": "hr",
                    "title": "الموارد البشرية",
                    "desc": "ملفات العاملين والإجازات.",
                    "route": "/ops/hr",
                    "roles": ["admin_deputy"],
                    "perms": ["admin_ops.hr"],
                    "icon": "https://cdn.lordicon.com/iltqorsz.json",
                    "bi": "bi-person-gear",
                },
                {
                    "id": "finance",
                    "title": "المالية والرسوم",
                    "desc": "الفوترة والتحصيل.",
                    "route": "/ops/finance",
                    "roles": ["admin_deputy"],
                    "perms": ["admin_ops.finance"],
                    "icon": "https://cdn.lordicon.com/vduvxizq.json",
                    "bi": "bi-wallet2",
                },
                {
                    "id": "procurement",
                    "title": "المشتريات",
                    "desc": "أوامر الشراء والموردون.",
                    "route": "/ops/procurement",
                    "roles": ["admin_deputy"],
                    "perms": ["admin_ops.procurement"],
                    "icon": "https://cdn.lordicon.com/kbtmbyzy.json",
                    "bi": "bi-cart-check",
                },
                {
                    "id": "assets",
                    "title": "الأصول والعهد",
                    "desc": "تتبع الأصول والجرد.",
                    "route": "/ops/assets",
                    "roles": ["admin_deputy"],
                    "perms": ["admin_ops.assets"],
                    "icon": "https://cdn.lordicon.com/qmcsqnle.json",
                    "bi": "bi-box-seam",
                },
                {
                    "id": "facilities",
                    "title": "المرافق والصيانة",
                    "desc": "طلبات الصيانة ومتابعتها.",
                    "route": "/ops/facilities",
                    "roles": ["admin_deputy"],
                    "perms": ["admin_ops.facilities"],
                    "icon": "https://cdn.lordicon.com/nqtddedc.json",
                    "bi": "bi-wrench",
                },
                {
                    "id": "transport",
                    "title": "النقل المدرسي",
                    "desc": "الحافلات والمسارات.",
                    "route": "/ops/transport",
                    "roles": ["admin_deputy"],
                    "perms": ["admin_ops.transport"],
                    "icon": "https://cdn.lordicon.com/vduvxizq.json",
                    "bi": "bi-bus-front",
                },
            ],
        },
        # Additional functional areas
        {
            "role": "شؤون الطلاب",
            "role_key": "student_affairs",
            "color": "#0d6efd",
            "items": [
                {
                    "id": "admissions",
                    "title": "القبول والتحويلات",
                    "desc": "طلبات القبول والتحويل.",
                    "route": "/affairs/admissions",
                    "roles": ["student_affairs"],
                    "perms": ["student_affairs.admissions"],
                    "icon": "https://cdn.lordicon.com/qhviklyi.json",
                    "bi": "bi-door-open",
                },
                {
                    "id": "records",
                    "title": "السجلات والملفات",
                    "desc": "ملفات الطلاب والوثائق.",
                    "route": "/affairs/records",
                    "roles": ["student_affairs"],
                    "perms": ["student_affairs.records"],
                    "icon": "https://cdn.lordicon.com/svbmmyue.json",
                    "bi": "bi-file-earmark-text",
                },
            ],
        },
        {
            "role": "المالية",
            "role_key": "finance",
            "color": "#795548",
            "items": [
                {
                    "id": "billing",
                    "title": "الفوترة والتحصيل",
                    "desc": "إدارة الرسوم والمدفوعات.",
                    "route": "/finance/billing",
                    "roles": ["finance"],
                    "perms": ["finance.billing"],
                    "icon": "https://cdn.lordicon.com/vduvxizq.json",
                    "bi": "bi-receipt",
                }
            ],
        },
        {
            "role": "المكتبة",
            "role_key": "library",
            "color": "#009688",
            "items": [
                {
                    "id": "catalog",
                    "title": "الفهرس والإعارة",
                    "desc": "عناوين وإعارات المكتبة.",
                    "route": "/library/catalog",
                    "roles": ["librarian"],
                    "perms": ["library.catalog"],
                    "icon": "https://cdn.lordicon.com/fgxwhgys.json",
                    "bi": "bi-book",
                }
            ],
        },
        {
            "role": "تقنية المعلومات",
            "role_key": "it",
            "color": "#455a64",
            "items": [
                {
                    "id": "it_users",
                    "title": "المستخدمون والأدوار",
                    "desc": "RBAC وإدارة الوصول.",
                    "route": "/it/users",
                    "roles": ["it"],
                    "perms": ["it.users"],
                    "icon": "https://cdn.lordicon.com/iltqorsz.json",
                    "bi": "bi-people-fill",
                },
                {
                    "id": "integrations",
                    "title": "التكاملات وAPI",
                    "desc": "واجهات وربط الأنظمة.",
                    "route": "/it/integrations",
                    "roles": ["it"],
                    "perms": ["it.integrations"],
                    "icon": "https://cdn.lordicon.com/uvextprq.json",
                    "bi": "bi-braces",
                },
            ],
        },
        {
            "role": "العيادة",
            "role_key": "health",
            "color": "#e53935",
            "items": [
                {
                    "id": "clinic_records",
                    "title": "السجلات الصحية",
                    "desc": "زيارات ولقاحات وحوادث.",
                    "route": "/health/clinic",
                    "roles": ["nurse"],
                    "perms": ["health.clinic"],
                    "icon": "https://cdn.lordicon.com/hjbsbdhw.json",
                    "bi": "bi-heart-pulse",
                }
            ],
        },
        {
            "role": "الإرشاد",
            "role_key": "guidance",
            "color": "#6d4c41",
            "items": [
                {
                    "id": "counseling",
                    "title": "الملفات الإرشادية",
                    "desc": "جلسات ومتابعات وخطط تدخل.",
                    "route": "/guidance/cases",
                    "roles": ["counselor"],
                    "perms": ["guidance.cases"],
                    "icon": "https://cdn.lordicon.com/kjkiqtxg.json",
                    "bi": "bi-people",
                }
            ],
        },
    ]
    # Extend with additional operational roles and catalogs (secretary, canteen, exams, safety, registrar, transport, attendance office, discipline, substitutions)
    # Additional operational roles and catalogs (secretary, canteen, exams, safety, registrar, transport, attendance office, discipline, substitutions)
    extra_sections = [
        {
            "role": "سكرتير المدرسة",
            "role_key": "secretary",
            "color": "#6c757d",
            "items": [
                {
                    "id": "front_office",
                    "title": "المكتب الأمامي",
                    "desc": "استقبال المراجعين، المراسلات، الأختام.",
                    "route": "/office/front",
                    "roles": ["secretary"],
                    "perms": ["office.front"],
                    "icon": "https://cdn.lordicon.com/qhviklyi.json",
                    "bi": "bi-envelope-paper",
                },
                {
                    "id": "letters",
                    "title": "الكتب الرسمية",
                    "desc": "إصدار واستقبال الكتب ومتابعتها.",
                    "route": "/office/letters",
                    "roles": ["secretary"],
                    "perms": ["office.letters"],
                    "icon": "https://cdn.lordicon.com/kbtmbyzy.json",
                    "bi": "bi-mailbox",
                },
            ],
        },
        {
            "role": "مسؤول الامتحانات",
            "role_key": "exams_officer",
            "color": "#1f6feb",
            "items": [
                {
                    "id": "exam_calendar",
                    "title": "تقويم الامتحانات",
                    "desc": "جداول وتوزيع اللجان.",
                    "route": "/exams/calendar",
                    "roles": ["exams_officer"],
                    "perms": ["exams.manage"],
                    "icon": "https://cdn.lordicon.com/zzcjjxew.json",
                    "bi": "bi-calendar2-range",
                },
                {
                    "id": "exam_invigilation",
                    "title": "المراقبات",
                    "desc": "إسناد المراقبات والتبديل.",
                    "route": "/exams/invigilation",
                    "roles": ["exams_officer"],
                    "perms": ["exams.invigilation"],
                    "icon": "https://cdn.lordicon.com/kthelypq.json",
                    "bi": "bi-clipboard2-check",
                },
            ],
        },
        {
            "role": "مسؤول المناوبات/الاحتياط",
            "role_key": "substitutions_officer",
            "color": "#9c27b0",
            "items": [
                {
                    "id": "subs_assign",
                    "title": "إسناد الاحتياط",
                    "desc": "تغطيات فورية للحصص.",
                    "route": "/subs/assign",
                    "roles": ["substitutions_officer"],
                    "perms": ["staffing.substitution"],
                    "icon": "https://cdn.lordicon.com/egiwmiit.json",
                    "bi": "bi-people",
                }
            ],
        },
        {
            "role": "مسؤول الانضباط",
            "role_key": "discipline_officer",
            "color": "#c62828",
            "items": [
                {
                    "id": "behavior_cases",
                    "title": "حالات السلوك",
                    "desc": "فتح ومتابعة القضايا.",
                    "route": "/behavior/cases",
                    "roles": ["discipline_officer"],
                    "perms": ["behavior.incidents.manage"],
                    "icon": "https://cdn.lordicon.com/iqaqnsks.json",
                    "bi": "bi-shield-exclamation",
                }
            ],
        },
        {
            "role": "مسؤول الحضور",
            "role_key": "attendance_office",
            "color": "#2e7d32",
            "items": [
                {
                    "id": "attendance_office_daily",
                    "title": "الحضور اليومي",
                    "desc": "إغلاق اليوم، معالجة الأعذار.",
                    "route": "/attendance/office/daily",
                    "roles": ["attendance_office"],
                    "perms": ["attendance.daily.manage"],
                    "icon": "https://cdn.lordicon.com/akuwjdzh.json",
                    "bi": "bi-calendar-check",
                }
            ],
        },
        {
            "role": "مسؤول السلامة",
            "role_key": "safety_officer",
            "color": "#f57f17",
            "items": [
                {
                    "id": "safety_incidents",
                    "title": "السلامة والحوادث",
                    "desc": "الإبلاغ والمتابعة.",
                    "route": "/safety/incidents",
                    "roles": ["safety_officer"],
                    "perms": ["safety.incidents"],
                    "icon": "https://cdn.lordicon.com/lywgqtim.json",
                    "bi": "bi-alarm",
                }
            ],
        },
        {
            "role": "مسؤول السجلات",
            "role_key": "registrar",
            "color": "#37474f",
            "items": [
                {
                    "id": "registry",
                    "title": "السجل الأكاديمي",
                    "desc": "الملفات والشهادات.",
                    "route": "/registrar/records",
                    "roles": ["registrar"],
                    "perms": ["student_affairs.records"],
                    "icon": "https://cdn.lordicon.com/svbmmyue.json",
                    "bi": "bi-archive",
                }
            ],
        },
        {
            "role": "مسؤول النقل",
            "role_key": "transport_officer",
            "color": "#1565c0",
            "items": [
                {
                    "id": "transport_routes",
                    "title": "مسارات الحافلات",
                    "desc": "الجداول والطلاب.",
                    "route": "/transport/routes",
                    "roles": ["transport_officer"],
                    "perms": ["admin_ops.transport"],
                    "icon": "https://cdn.lordicon.com/vduvxizq.json",
                    "bi": "bi-bus-front",
                }
            ],
        },
        {
            "role": "المقصف المدرسي",
            "role_key": "canteen_manager",
            "color": "#8d6e63",
            "items": [
                {
                    "id": "canteen",
                    "title": "إدارة المقصف",
                    "desc": "القوائم والحسابات.",
                    "route": "/canteen",
                    "roles": ["canteen_manager"],
                    "perms": ["ops.canteen"],
                    "icon": "https://cdn.lordicon.com/kbtmbyzy.json",
                    "bi": "bi-basket3",
                }
            ],
        },
        {
            "role": "المرشد الاجتماعي",
            "role_key": "social_counselor",
            "color": "#6d4c41",
            "items": [
                {
                    "id": "guidance_cases",
                    "title": "ملفات الإرشاد الاجتماعي",
                    "desc": "حالات سلوكية وجلسات ومتابعات.",
                    "route": "/guidance/social/cases",
                    "roles": ["social_counselor"],
                    "perms": ["guidance.cases"],
                    "icon": "https://cdn.lordicon.com/kjkiqtxg.json",
                    "bi": "bi-people",
                },
                {
                    "id": "interventions",
                    "title": "خطط التدخل",
                    "desc": "خطط دعم وتدخل فردية.",
                    "route": "/guidance/social/interventions",
                    "roles": ["social_counselor"],
                    "perms": ["guidance.interventions"],
                    "icon": "https://cdn.lordicon.com/qtqvorle.json",
                    "bi": "bi-list-check",
                },
                {
                    "id": "parents_comm",
                    "title": "التواصل مع أولياء الأمور",
                    "desc": "اتصالات منظمة ومؤرشفة.",
                    "route": "/guidance/social/parents",
                    "roles": ["social_counselor"],
                    "perms": ["communications.parents"],
                    "icon": "https://cdn.lordicon.com/megfpqpw.json",
                    "bi": "bi-chat-dots",
                },
            ],
        },
        {
            "role": "المرشد الأكاديمي",
            "role_key": "academic_counselor",
            "color": "#1e88e5",
            "items": [
                {
                    "id": "academic_advising",
                    "title": "الإرشاد الأكاديمي",
                    "desc": "خطط دراسية وإرشاد تحصيلي.",
                    "route": "/guidance/academic/advising",
                    "roles": ["academic_counselor"],
                    "perms": ["guidance.academic"],
                    "icon": "https://cdn.lordicon.com/oezixobx.json",
                    "bi": "bi-mortarboard",
                },
                {
                    "id": "remedial_programs",
                    "title": "برامج علاجية",
                    "desc": "خطط تعويض الفجوات الأكاديمية.",
                    "route": "/guidance/academic/remedial",
                    "roles": ["academic_counselor"],
                    "perms": ["guidance.remedial"],
                    "icon": "https://cdn.lordicon.com/krmfspeu.json",
                    "bi": "bi-bar-chart",
                },
                {
                    "id": "risk_alerts",
                    "title": "تنبيهات التراجع",
                    "desc": "رصد الطلبة المعرضين للخطر.",
                    "route": "/guidance/academic/alerts",
                    "roles": ["academic_counselor"],
                    "perms": ["analytics.risk"],
                    "icon": "https://cdn.lordicon.com/lywgqtim.json",
                    "bi": "bi-exclamation-triangle",
                },
            ],
        },
    ]
    # Merge base catalog with extra sections
    catalog += extra_sections

    # Place teacher evaluation items under their respective roles (instead of a separate section)
    eval_by_coordinator = {
        "id": "eval_by_coordinator",
        "title": "تقييم المنسق",
        "desc": "زيارات صفية وروبرك.",
        "route": "/evaluations/coordinator",
        "roles": ["subject_coordinator"],
        "perms": ["evaluations.write"],
        "icon": "https://cdn.lordicon.com/qtqvorle.json",
        "bi": "bi-clipboard2-data",
    }
    eval_by_academic = {
        "id": "eval_by_academic",
        "title": "تقييم النائب الأكاديمي",
        "desc": "متابعة الأداء الأكاديمي.",
        "route": "/evaluations/academic",
        "roles": ["academic_deputy"],
        "perms": ["evaluations.manage"],
        "icon": "https://cdn.lordicon.com/oezixobx.json",
        "bi": "bi-clipboard2-check",
    }
    eval_by_principal = {
        "id": "eval_by_principal",
        "title": "تقييم المدير",
        "desc": "تقييم شامل وملاحظات.",
        "route": "/evaluations/principal",
        "roles": ["principal"],
        "perms": ["evaluations.approve"],
        "icon": "https://cdn.lordicon.com/vspbqszr.json",
        "bi": "bi-clipboard2-pulse",
    }

    for _sec in catalog:
        rk = _sec.get("role_key")
        if rk == "subject_coordinator":
            _sec.setdefault("items", []).append(eval_by_coordinator)
        elif rk == "academic_deputy":
            _sec.setdefault("items", []).append(eval_by_academic)
        elif rk == "principal":
            _sec.setdefault("items", []).append(eval_by_principal)

    use_lordicon = not bool(request.GET.get("no_lordicon"))
    # Best-practice guidance per role (daily/weekly/monthly/KPIs/permissions hints)
    best_practices = {
        "teacher": {
            "daily": [
                "تسجيل حضور الحصة فور بدايتها (≤ 5 دقائق)",
                "مراجعة التأخر وإضافة ملاحظة تربوية مختصرة عند الحاجة",
                "نشر واجب واحد على الأقل لكل مادة نشطة",
            ],
            "weekly": [
                "تحديث خطط الدروس للأسبوع القادم",
                "مراجعة نتائج الكويزات ومعالجة فجوات الإتقان",
            ],
            "monthly": [
                "مزامنة دفتر العلامات والتقارير القصيرة مع ولي الأمر",
            ],
            "kpis": ["نسبة الحضور", "إنجاز الواجبات", "تقدّم الأهداف"],
            "perms": ["attendance.session.write", "academics.lesson_plans", "academics.grades"],
        },
        "wing_supervisor": {
            "daily": [
                "مراقبة الغياب اليومي وإغلاق اليوم للجناح",
                "متابعة حصص بلا إدخال وإشعار المعلمين",
                "إدارة أذونات الخروج ومتابعة العائدين",
            ],
            "weekly": [
                "تحليل الصفوف الأكثر غيابًا ووضع خطة تدخل",
                "مراجعة البلاغات السلوكية المفتوحة",
            ],
            "monthly": [
                "تقرير شهري لنسبة الحضور والتأخر على مستوى الجناح",
            ],
            "kpis": ["absentToday", "exit_events_open", "present_pct"],
            "perms": [
                "attendance.wing.read",
                "attendance.exits.manage",
                "behavior.incidents.manage",
            ],
        },
        "subject_coordinator": {
            "daily": ["دعم المعلمين بالمحتوى والأسئلة"],
            "weekly": ["توحيد التقييمات القصيرة", "مراجعة بنك الأسئلة"],
            "monthly": ["تقارير مقارنة بين الصفوف والمعلمين"],
            "kpis": ["اتساق التقييم", "تقدم المنهج"],
            "perms": ["academics.question_bank", "academics.rubrics", "analytics.subject.read"],
        },
        "principal": {
            "daily": ["مراجعة لوحة القيادة والتنبيهات المهمة"],
            "weekly": ["اجتماع متابعة مع النواب"],
            "monthly": ["تقارير رسمية للوزارة/المالك"],
            "kpis": ["حضور المدرسة", "تحصيل أكاديمي", "انضباط"],
            "perms": ["analytics.kpi", "reports.official", "it.users"],
        },
        "academic_deputy": {
            "daily": ["حل تعارضات الجدول"],
            "weekly": ["ضبط جداول الاختبارات", "متابعة تنفيذ المنهج"],
            "monthly": ["تقييم جودة التقييمات والروبركات"],
            "kpis": ["اكتمال الجداول", "معدلات التقييم"],
            "perms": [
                "academics.timetable.manage",
                "academics.assessments.manage",
                "academics.curricula",
            ],
        },
        "admin_deputy": {
            "daily": ["موافقات الإجازات والطلبات التشغيلية"],
            "weekly": ["مراجعة أوامر الشراء والصيانة"],
            "monthly": ["موازنة الرسوم والتحصيل"],
            "kpis": ["طلبات صيانة مفتوحة", "تحصيل الرسوم"],
            "perms": [
                "admin_ops.hr",
                "admin_ops.finance",
                "admin_ops.facilities",
                "admin_ops.procurement",
            ],
        },
        "student_affairs": {
            "daily": ["معالجة طلبات القبول والتحويل"],
            "weekly": ["تحديث السجلات والوثائق"],
            "monthly": ["تقارير الحضور والانضباط للطلاب"],
            "kpis": ["زمن إنجاز الطلبات"],
            "perms": ["student_affairs.admissions", "student_affairs.records"],
        },
        "finance": {
            "daily": ["إصدار الفواتير ومعالجة المدفوعات"],
            "weekly": ["مطابقات مالية"],
            "monthly": ["تقارير مالية"],
            "kpis": ["التحصيل", "المتأخرات"],
            "perms": ["finance.billing"],
        },
        "library": {
            "daily": ["إعارات وإرجاعات الكتب"],
            "weekly": ["تحديث الفهرس"],
            "monthly": ["تقارير الإعارة"],
            "kpis": ["نشاط المكتبة"],
            "perms": ["library.catalog"],
        },
        "it": {
            "daily": ["دعم المستخدمين وإدارة الأدوار"],
            "weekly": ["مراجعة السجلات والتكاملات"],
            "monthly": ["نسخ احتياطية واختبارات استرداد"],
            "kpis": ["طلبات الدعم", "نجاح النسخ"],
            "perms": ["it.users", "it.integrations"],
        },
        "health": {
            "daily": ["تسجيل الزيارات والإسعافات"],
            "weekly": ["متابعة الحالات"],
            "monthly": ["تقارير صحية"],
            "kpis": ["حالات طارئة", "زيارات"],
            "perms": ["health.clinic"],
        },
        "guidance": {
            "daily": ["جلسات إرشاد ومتابعة"],
            "weekly": ["خطط تدخل"],
            "monthly": ["تقارير متابعة"],
            "kpis": ["حالات مفتوحة", "خطة لكل طالب"],
            "perms": ["guidance.cases"],
        },
        "social_counselor": {
            "daily": [
                "استقبال الحالات السلوكية وتوثيقها",
                "جلسات إرشاد فردية قصيرة",
                "اتصال بولي الأمر عند الحاجة",
            ],
            "weekly": ["مراجعة خطط التدخل النشطة", "اجتماع تنسيقي مع مشرف الجناح/المرشد الأكاديمي"],
            "monthly": [
                "تقرير شهري عن الحالات المفتوحة والمغلقة",
                "تحليل أنماط السلوك ووضع توصيات",
            ],
            "kpis": ["عدد الحالات المفتوحة", "نسبة الحالات المغلقة", "متوسط زمن المعالجة"],
            "perms": ["guidance.cases", "guidance.interventions", "communications.parents"],
        },
        "academic_counselor": {
            "daily": [
                "متابعة الطلبة ذوي التراجع الأكاديمي",
                "إرشاد حول تنظيم الدراسة ووضع أهداف قصيرة المدى",
            ],
            "weekly": [
                "خطة علاجية فردية بالتنسيق مع المعلم والمنسق",
                "تحليل نتائج الكويزات/الاختبارات لرصد الفجوات",
            ],
            "monthly": ["تقرير تحصيلي لفئات الخطر", "تقييم فاعلية البرامج العلاجية"],
            "kpis": ["تحسن الدرجات", "نسبة حضور الجلسات الأكاديمية", "عدد الطلبة في خطر"],
            "perms": ["guidance.academic", "guidance.remedial", "analytics.risk"],
        },
    }
    # Attach best practices per role to each section for straightforward template access
    try:
        for _sec in catalog:
            rk = _sec.get("role_key")
            _sec["best"] = best_practices.get(rk, {})
    except Exception:
        pass
    # User roles (groups) for role-based filtering in template
    try:
        user_roles = list(request.user.groups.values_list("name", flat=True))
    except Exception:
        user_roles = []
    # ---- Dynamic enhancement from database staff job titles ----
    # Build role counts and title mapping using Staff.job_title to strengthen catalog based on real workforce
    try:
        title_rows = (
            Staff.objects.exclude(job_title__isnull=True)
            .exclude(job_title="")
            .values("job_title")
            .annotate(cnt=Count("id"))
            .order_by("job_title")
        )
    except Exception:
        title_rows = []

    # Mapping rules from Arabic job titles to platform role keys
    ROLE_MAP_RULES: list[tuple[re.Pattern, str]] = [
        (re.compile(r"مدير|المدير|قائد"), "principal"),
        (re.compile(r"نائب\s*أكاديمي|وكيل\s*أكاديمي|الشؤون\s*الأكاديمية"), "academic_deputy"),
        (re.compile(r"نائب\s*إداري|وكيل\s*إداري|الشؤون\s*الإدارية"), "admin_deputy"),
        (re.compile(r"منسق|منسّق|منسق مادة|رئيس قسم"), "subject_coordinator"),
        (re.compile(r"مشرف\s*جناح|مشرف مرحلة|رائد|head of wing"), "wing_supervisor"),
        (re.compile(r"معلم|معلّم|teacher"), "teacher"),
        (re.compile(r"سكرتير|سكرتارية|secretary"), "secretary"),
        (
            re.compile(r"امتحان|الامتحانات|اختبار|تقويم|امتحانات|امتحاني|امتحانيه|exams"),
            "exams_officer",
        ),
        (re.compile(r"مناوب|احتياط|تغطية|بديل"), "substitutions_officer"),
        (re.compile(r"انضباط|سلوك|تأديب|discipline"), "discipline_officer"),
        (re.compile(r"حضور|الغياب|Attendance"), "attendance_office"),
        (re.compile(r"سلامة|أمن|طوارئ|سلامه"), "safety_officer"),
        (re.compile(r"سجلات|مسجل|Registrar"), "registrar"),
        (re.compile(r"نقل|مواصلات|حافلات|باص"), "transport_officer"),
        (re.compile(r"مقصف|كانتين|مطعم مدرسي"), "canteen_manager"),
        (re.compile(r"مكتبة|أمين مكتبة|librar"), "library"),
        (re.compile(r"تقنية المعلومات|نظم|مبرمج|شبكات|حاسوب|IT|تقني"), "it"),
        (re.compile(r"ممرض|عيادة|صحة|تمريض|nurse|clinic"), "health"),
        (re.compile(r"مرشد\s*اجتماعي|اجتماعي|إرشاد اجتماعي"), "social_counselor"),
        (re.compile(r"مرشد\s*أكاديمي|إرشاد أكاديمي|ارشاد اكاديمي"), "academic_counselor"),
        (re.compile(r"شؤون\s*الطلاب|سجلات الطلاب|قبول|تسجيل"), "student_affairs"),
        (re.compile(r"مالية|محاسب|رسوم|تحصيل|Finance"), "finance"),
    ]

    def map_title_to_role(title: str) -> str | None:
        t = (title or "").strip()
        for rx, rk in ROLE_MAP_RULES:
            try:
                if rx.search(t):
                    return rk
            except Exception:
                continue
        return None

    role_counts: dict[str, int] = {}
    job_titles_map: list[dict] = []
    for row in title_rows:
        jt = str(row.get("job_title") or "").strip()
        cnt = int(row.get("cnt") or 0)
        rk = map_title_to_role(jt)
        job_titles_map.append({"job_title": jt, "count": cnt, "role_key": rk})
        if rk:
            role_counts[rk] = role_counts.get(rk, 0) + cnt

    # Strengthen teacher count using authoritative sources: Staff.role/group + TeachingAssignment + TimetableEntry(current term)
    try:
        from .models import Staff as _Staff  # type: ignore
        from .models import TeachingAssignment as _TA
        from .models import Term as _Term
        from .models import TimetableEntry as _TE

        teacher_ids: set[int] = set()
        # Staff.role = 'teacher'
        try:
            teacher_ids.update(int(i) for i in _Staff.objects.filter(role="teacher").values_list("id", flat=True))
        except Exception:
            pass
        # Group membership 'teacher'
        try:
            teacher_ids.update(
                int(i)
                for i in _Staff.objects.filter(user__groups__name="teacher").values_list("id", flat=True).distinct()
            )
        except Exception:
            pass
        # TeachingAssignment teacher ids
        try:
            teacher_ids.update(int(i) for i in _TA.objects.values_list("teacher_id", flat=True).distinct() if i)
        except Exception:
            pass
        # TimetableEntry teacher ids for current term (if any)
        try:
            _term = _Term.objects.filter(is_current=True).first()
            if _term:
                teacher_ids.update(
                    int(i) for i in _TE.objects.filter(term=_term).values_list("teacher_id", flat=True).distinct() if i
                )
        except Exception:
            pass
        if teacher_ids:
            role_counts["teacher"] = len({i for i in teacher_ids if i})
    except Exception:
        pass

    # Attach staff_count to each catalog section
    for _sec in catalog:
        rk = _sec.get("role_key")
        _sec["staff_count"] = role_counts.get(rk, 0)

    # Determine if current user can edit (UI designer roles)
    try:
        is_designer = _is_ui_designer(request.user)  # type: ignore[name-defined]
    except Exception:
        is_designer = False
    context = {
        "title": "منظومة الأيقونات والصلاحيات",
        "catalog": catalog,
        "use_lordicon": use_lordicon,
        "best_practices": best_practices,
        "user_roles": user_roles,
        "role_counts": role_counts,
        "job_titles_map": job_titles_map,
        "is_designer": is_designer,
    }
    return render(request, "data/icons_catalog.html", context)


@user_passes_test(_staff_only)
@login_required
def data_relations(request):
    """Visual ER-style diagram of the core app (school) database relations.
    Produces a Mermaid ER diagram string from Django model metadata.
    Adds live database info (engine/name/version) — handy for PostgreSQL environments.
    Enhanced with model categorization and comprehensive statistics.
    """
    try:
        from django.apps import apps
        from django.db.models import ForeignKey, ManyToManyField, OneToOneField
    except Exception:
        return render(
            request,
            "data/relations.html",
            {
                "title": "العلاقات بين الجداول",
                "mermaid": "erDiagram\n",
                "db_info": None,
            },
        )

    # Collect DB info (engine/vendor, name, version if supported)
    db_info = None
    try:
        vendor = connection.vendor  # 'postgresql', 'sqlite', 'mysql', ...
        settings_dict = getattr(connection, "settings_dict", {}) or {}
        db_name = settings_dict.get("NAME")
        db_ver = None
        if vendor == "postgresql":
            with connection.cursor() as cursor:
                cursor.execute("SELECT version();")
                row = cursor.fetchone()
                db_ver = row[0] if row else None
        elif vendor == "mysql":
            with connection.cursor() as cursor:
                cursor.execute("SELECT VERSION();")
                row = cursor.fetchone()
                db_ver = row[0] if row else None
        else:
            # Some backends expose server_version via connection
            db_ver = getattr(connection, "server_version", None)
        db_info = {"vendor": vendor, "name": db_name, "version": db_ver}
    except Exception:
        db_info = None

    app_label = "school"
    models = [m for m in apps.get_app_config(app_label).get_models()]

    # Categorize models for better organization
    model_categories = {
        "core": ["Class", "Student", "Staff", "Subject"],
        "academic": [
            "ClassSubject",
            "TeachingAssignment",
            "TimetableEntry",
            "AcademicYear",
            "Term",
        ],
        "attendance": [
            "AttendanceRecord",
            "AttendanceDaily",
            "AttendancePolicy",
            "ExitEvent",
        ],
        "scheduling": ["Wing", "PeriodTemplate", "TemplateSlot"],
        "assessment": ["AssessmentPackage", "SchoolHoliday"],
    }

    # Build entities and relations
    entities = []  # list of (ModelName, [fields], category)
    rels = []  # list of (left, left_card, right_card, right, label)
    model_stats = {"total": 0, "with_fk": 0, "with_m2m": 0, "with_o2o": 0}

    def model_alias(m):
        return m.__name__

    def get_category(model_name):
        for cat, names in model_categories.items():
            if model_name in names:
                return cat
        return "other"

    for m in models:
        model_stats["total"] += 1
        model_name = model_alias(m)
        category = get_category(model_name)

        # Collect main fields (exclude many-to-many auto tables)
        flds = []
        has_fk = has_m2m = has_o2o = False

        for f in m._meta.get_fields():
            try:
                if hasattr(f, "many_to_many") and f.many_to_many and not f.concrete:
                    continue
                if getattr(f, "auto_created", False) and not getattr(f, "concrete", False):
                    continue
                name = getattr(f, "name", getattr(f, "attname", str(f)))

                # Track field types
                if isinstance(f, ForeignKey):
                    has_fk = True
                    flds.append(f"{name}_FK")
                elif isinstance(f, OneToOneField):
                    has_o2o = True
                    flds.append(f"{name}_O2O")
                elif isinstance(f, ManyToManyField):
                    has_m2m = True
                    flds.append(f"{name}_M2M")
                else:
                    flds.append(name)
            except Exception:
                continue

        if has_fk:
            model_stats["with_fk"] += 1
        if has_m2m:
            model_stats["with_m2m"] += 1
        if has_o2o:
            model_stats["with_o2o"] += 1

        entities.append((model_name, flds, category))

        # Relations: FK and O2O from this model to target
        for f in m._meta.get_fields():
            try:
                if isinstance(f, ForeignKey):
                    rels.append((model_name, "}o", "||", model_alias(f.related_model), f.name))
                elif isinstance(f, OneToOneField):
                    rels.append((model_name, "||", "||", model_alias(f.related_model), f.name))
                elif isinstance(f, ManyToManyField):
                    # many-to-many (skip auto-created through models displayed as entities anyway)
                    rels.append((model_name, "}o", "o{", model_alias(f.related_model), f.name))
            except Exception:
                continue

    # Compose Mermaid erDiagram with enhanced formatting
    lines = ["erDiagram"]

    # Helper to sanitize field names for Mermaid (remove special chars)
    def sanitize_field(fld):
        return fld.replace("_FK", "").replace("_O2O", "").replace("_M2M", "").replace("&", "and")

    lines.append("  %% Core Models Foundation")
    for name, flds, cat in sorted(entities, key=lambda x: (x[2], x[0])):
        if cat == "core":
            lines.append(f"  {name} {{")
            # Show up to 15 fields for better visibility
            for fld in flds[:15]:
                clean_fld = sanitize_field(fld)
                lines.append(f"    string {clean_fld}")
            if len(flds) > 15:
                lines.append("    string more_fields")
            lines.append("  }")

    lines.append("")
    lines.append("  %% Academic Models Curriculum Teaching")
    for name, flds, cat in sorted(entities, key=lambda x: (x[2], x[0])):
        if cat == "academic":
            lines.append(f"  {name} {{")
            for fld in flds[:15]:
                clean_fld = sanitize_field(fld)
                lines.append(f"    string {clean_fld}")
            if len(flds) > 15:
                lines.append("    string more_fields")
            lines.append("  }")

    lines.append("")
    lines.append("  %% Attendance Exit Tracking")
    for name, flds, cat in sorted(entities, key=lambda x: (x[2], x[0])):
        if cat == "attendance":
            lines.append(f"  {name} {{")
            for fld in flds[:15]:
                clean_fld = sanitize_field(fld)
                lines.append(f"    string {clean_fld}")
            if len(flds) > 15:
                lines.append("    string more_fields")
            lines.append("  }")

    lines.append("")
    lines.append("  %% Scheduling Infrastructure")
    for name, flds, cat in sorted(entities, key=lambda x: (x[2], x[0])):
        if cat in ["scheduling", "assessment"]:
            lines.append(f"  {name} {{")
            for fld in flds[:15]:
                clean_fld = sanitize_field(fld)
                lines.append(f"    string {clean_fld}")
            if len(flds) > 15:
                lines.append("    string more_fields")
            lines.append("  }")

    lines.append("")
    lines.append("  %% Other Models")
    for name, flds, cat in sorted(entities, key=lambda x: (x[2], x[0])):
        if cat == "other":
            lines.append(f"  {name} {{")
            for fld in flds[:15]:
                clean_fld = sanitize_field(fld)
                lines.append(f"    string {clean_fld}")
            if len(flds) > 15:
                lines.append("    string more_fields")
            lines.append("  }")

    lines.append("")
    lines.append("  %% Relations")
    # Relations (deduplicate)
    seen = set()
    for a, lc, rc, b, label in rels:
        key = (a, lc, rc, b, label)
        if key in seen:
            continue
        seen.add(key)
        # Sanitize label (remove underscores and special chars)
        clean_label = label.replace("_", "")
        lines.append(f"  {a} {lc}--{rc} {b} : {clean_label}")

    mermaid_src = "\n".join(lines)

    # Optional: collect constraints and indexes to reflect latest updates
    constraints = None
    indexes = None
    try:
        if connection.vendor == "postgresql":
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT t.relname AS table,
                           c.conname AS name,
                           pg_get_constraintdef(c.oid) AS def
                    FROM pg_constraint c
                    JOIN pg_class t ON t.oid = c.conrelid
                    JOIN pg_namespace n ON n.oid = t.relnamespace
                    WHERE n.nspname='public' AND t.relname LIKE 'school_%'
                    ORDER BY t.relname, c.conname
                    """
                )
                constraints = [{"table": r[0], "name": r[1], "def": r[2]} for r in cursor.fetchall() or []]
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT tablename, indexname, indexdef
                    FROM pg_indexes
                    WHERE schemaname='public' AND tablename LIKE 'school_%'
                    ORDER BY tablename, indexname
                    """
                )
                indexes = [{"table": r[0], "name": r[1], "def": r[2]} for r in cursor.fetchall() or []]
    except Exception:
        constraints = None
        indexes = None

    # Also compute table stats (like data_overview) limited to school_* tables
    table_stats = []
    try:
        with connection.cursor() as cursor:
            introspection = connection.introspection
            tables = introspection.table_names()
            visible = [t for t in tables if t.startswith("school_")]
            for t in sorted(visible):
                try:
                    cursor.execute(f'SELECT COUNT(*) FROM "{t}"')
                    cnt = cursor.fetchone()[0]
                except Exception:
                    cnt = None
                table_stats.append({"name": t, "count": cnt})
    except Exception:
        table_stats = []

    context = {
        "title": "العلاقات بين الجداول",
        "mermaid": mermaid_src,
        "db_info": db_info,
        "constraints": constraints,
        "indexes": indexes,
        "model_stats": model_stats,
        "total_models": len(entities),
        "total_relations": len(rels),
        "tables": table_stats,
    }
    return render(request, "data/relations.html", context)


@user_passes_test(_staff_only)
@login_required
def data_table_detail(request, table):
    # Security: allow only public tables and apply same deny rules
    deny_exact = {"django_session"}
    deny_prefixes = ("auth_", "django_", "admin_")

    with connection.cursor() as cursor:
        introspection = connection.introspection
        all_tables = set(introspection.table_names())

    if table not in all_tables or table in deny_exact or any(table.startswith(p) for p in deny_prefixes):
        raise Http404("Table not found")

    # Registry of editable tables
    MODEL_REGISTRY = {
        Class._meta.db_table: (Class, ClassForm),
        Student._meta.db_table: (Student, StudentForm),
        Staff._meta.db_table: (Staff, StaffForm),
        Subject._meta.db_table: (Subject, SubjectForm),
        ClassSubject._meta.db_table: (ClassSubject, ClassSubjectForm),
        TeachingAssignment._meta.db_table: (TeachingAssignment, TeachingAssignmentForm),
    }

    editable = table in MODEL_REGISTRY

    # Common query params
    q = (request.GET.get("q") or "").strip()
    order = request.GET.get("order") or ""
    page = int(request.GET.get("page") or 1)
    per_page = int(request.GET.get("per_page") or 25)

    if editable:
        Model, FormClass = MODEL_REGISTRY[table]
        # Build columns, include pk first for actions
        fields = [f for f in Model._meta.fields if f.editable or f.primary_key]
        columns = [Model._meta.pk.name] + [f.name for f in fields if f.name != Model._meta.pk.name]

        # Handle POST actions
        if request.method == "POST":
            # Helper to rebuild querystring preserving filters
            def _qs_keep(req, drop_keys=None):
                drop_keys = set(drop_keys or [])
                from urllib.parse import parse_qsl, urlencode

                pairs = [
                    (k, v)
                    for k, v in parse_qsl(req.META.get("QUERY_STRING", ""), keep_blank_values=True)
                    if k not in drop_keys
                ]
                return ("?" + urlencode(pairs, doseq=True)) if pairs else ""

            action = request.POST.get("action")
            if action == "create":
                form = FormClass(request.POST)
                if form.is_valid():
                    try:
                        form.save()
                        messages.success(request, "تمت الإضافة بنجاح.")
                    except IntegrityError as e:
                        messages.error(request, f"تعذر الحفظ بسبب قيود البيانات: {e}")
                        return redirect(request.path + _qs_keep(request))
                    return redirect(request.path + _qs_keep(request, drop_keys=["add"]))
            elif action == "update":
                pk = request.POST.get("id")
                instance = Model.objects.filter(pk=pk).first()
                if instance is not None:
                    form = FormClass(request.POST, instance=instance)
                    if form.is_valid():
                        try:
                            form.save()
                            messages.success(request, "تم حفظ التعديلات.")
                        except IntegrityError as e:
                            messages.error(request, f"تعذر الحفظ بسبب قيود البيانات: {e}")
                            return redirect(request.path + _qs_keep(request))
                        # Preserve current page filters
                        return redirect(request.path + _qs_keep(request, drop_keys=["edit"]))
            elif action == "delete":
                pk = request.POST.get("id")
                obj = Model.objects.filter(pk=pk).first()
                if obj is not None:
                    try:
                        obj.delete()
                        messages.success(request, "تم حذف السجل.")
                    except (ProtectedError, IntegrityError) as e:
                        messages.error(request, f"لا يمكن حذف السجل لارتباطه بسجلات أخرى: {e}")
                return redirect(request.path + _qs_keep(request))

        # Build queryset (use select_related for FK fields and keep objects for better display)
        qs = Model.objects.all()
        # Select related for FK fields to optimize display
        fk_fields = [
            f.name for f in Model._meta.fields if getattr(f, "is_relation", False) and getattr(f, "many_to_one", False)
        ]
        if fk_fields:
            qs = qs.select_related(*fk_fields)
        if q:
            # Apply search across textual fields
            text_fields = [
                f.name
                for f in Model._meta.fields
                if getattr(f, "get_internal_type", lambda: "")() in ("CharField", "TextField")
            ]
            query = Q()
            for name in text_fields:
                query |= Q(**{f"{name}__icontains": q})
            if query:
                qs = qs.filter(query)
        if order:
            if order.lstrip("-") in columns:
                qs = qs.order_by(order)
        else:
            # Ensure deterministic ordering for pagination to avoid UnorderedObjectListWarning
            qs = qs.order_by(Model._meta.pk.name)
        total_count = qs.count()
        paginator = Paginator(qs, per_page)
        page_obj = paginator.get_page(page)

        add_mode = request.GET.get("add") == "1"
        edit_id = request.GET.get("edit")
        create_form = FormClass() if add_mode else None
        edit_form = None
        if edit_id:
            inst = Model.objects.filter(pk=edit_id).first()
            if inst is not None:
                edit_form = FormClass(instance=inst)

        context = {
            "title": f"جدول: {table}",
            "table": table,
            "columns": columns,
            "page_obj": page_obj,
            "q": q,
            "order": order,
            "per_page": per_page,
            "per_page_options": [10, 25, 50, 100],
            "total_count": total_count,
            "editable": True,
            "create_form": create_form,
            "edit_form": edit_form,
            "edit_id": edit_id,
        }
        return render(request, "data/table_detail.html", context)

    # Fallback read-only for non-registered tables (raw SQL path)
    # Read columns (vendor-agnostic)
    with connection.cursor() as cursor:
        description = connection.introspection.get_table_description(cursor, table)
        columns = [col.name for col in description]

    order_sql = ""
    if order and order.lstrip("-") in columns:
        col = order.lstrip("-")
        direction = "DESC" if order.startswith("-") else "ASC"
        order_sql = f' ORDER BY "{col}" {direction}'

    where_sql = ""
    params = []
    if q:
        like_params = []
        for col in columns:
            like_params.append(f'CAST("{col}" AS TEXT) ILIKE %s')
            params.append(f"%{q}%")
        where_sql = " WHERE " + " OR ".join(like_params)

    sql = f'SELECT * FROM "{table}"' + where_sql + order_sql
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()

    paginator = Paginator(rows, per_page)
    page_obj = paginator.get_page(page)

    context = {
        "title": f"جدول: {table}",
        "table": table,
        "columns": columns,
        "page_obj": page_obj,
        "q": q,
        "order": order,
        "per_page": per_page,
        "per_page_options": [10, 25, 50, 100],
        "total_count": len(rows),
        "editable": False,
    }
    return render(request, "data/table_detail.html", context)


@user_passes_test(_staff_only)
@login_required
def export_table_csv(request, table):
    # Registry of exportable tables (same as editable set)
    MODEL_REGISTRY = {
        Class._meta.db_table: Class,
        Student._meta.db_table: Student,
        Staff._meta.db_table: Staff,
        Subject._meta.db_table: Subject,
        ClassSubject._meta.db_table: ClassSubject,
        TeachingAssignment._meta.db_table: TeachingAssignment,
    }
    Model = MODEL_REGISTRY.get(table)
    if not Model:
        raise Http404("Table not found")

    fields = [f.name for f in Model._meta.concrete_fields]

    def row_iter():
        # Header
        yield ",".join(fields) + "\n"
        # Stream rows to avoid memory blowup
        for obj in Model.objects.all().iterator():
            vals = []
            for f in fields:
                v = getattr(obj, f)
                if v is None:
                    vals.append("")
                else:
                    s = str(v)
                    # Basic CSV escaping for commas and quotes
                    if any(ch in s for ch in [",", '"', "\n", "\r"]):
                        s = '"' + s.replace('"', '""') + '"'
                    vals.append(s)
            yield ",".join(vals) + "\n"

    resp = StreamingHttpResponse(row_iter(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{table}.csv"'
    return resp


# Timetable and calendar features have been fully removed from this module.
# Any previous views or utilities related to calendars or weekly timetables were
# deleted to ensure a safe and final cleanup as requested.


@login_required
def export_assignments_xlsx(request):
    """Export all teaching assignments to Excel with Arabic-friendly headers and professional styling (A3 fit, RTL, header/footer, wrap)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "الأنصبة"

    # RTL layout for Arabic
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    # Page setup: A3 landscape, fit to width
    try:
        ws.page_setup.paperSize = ws.PAPERSIZE_A3  # type: ignore[attr-defined]
    except Exception:
        pass
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # Margins (narrow)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    headers = [
        "المعلم",
        "الصف",
        "الشعبة",
        "المادة",
        "حصص/أسبوع",
        "ملاحظات",
    ]

    # Styles
    header_fill = PatternFill("solid", fgColor="F7F2F1")
    odd_fill = PatternFill("solid", fgColor="D7EEE5")  # matches #d7eee5
    even_fill = PatternFill("solid", fgColor="D9E7FB")  # matches #d9e7fb
    thin = Side(style="thin", color="8A8F99")  # darker grid
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    # Banner header row (colored like site theme)
    ws.append(["لوحة الأنصبة — جدول التكليفات"] + [None] * (len(headers) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    banner = ws.cell(row=1, column=1)
    banner.fill = PatternFill("solid", fgColor="7B1E1E")  # maroon
    banner.font = white_bold
    banner.alignment = center

    # Table header (row 2)
    ws.append(headers)

    # Style header row (row 2)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = bold
        c.alignment = center
        c.border = border

    # Freeze panes: keep banner+header and first column visible
    ws.freeze_panes = ws["B3"]

    # Order similar to dashboard default and include category for coloring
    qs = (
        TeachingAssignment.objects.select_related("teacher", "classroom", "subject")
        .annotate(
            teacher_min_category=Subquery(
                Staff.objects.filter(pk=OuterRef("teacher_id"))
                .annotate(all_min=Min(_subject_category_order_expr("assignments__subject__name_ar")))
                .values("all_min")[:1]
            )
        )
        .order_by(
            "teacher_min_category",
            "teacher__full_name",
            "classroom__grade",
            "classroom__section",
        )
    )

    # Write data rows with styling starting from row 3
    row_idx = 3
    for a in qs.iterator():
        ws.append(
            [
                a.teacher.full_name if a.teacher else "",
                a.classroom.grade if a.classroom else "",
                (a.classroom.section if a.classroom and a.classroom.section else ""),
                a.subject.name_ar if a.subject else "",
                int(a.no_classes_weekly or 0),
                a.notes or "",
            ]
        )
        # Apply fill based on category parity (match page colors)
        cat = int(a.teacher_min_category or 999)
        row_fill = odd_fill if (cat % 2 == 1) else even_fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = right if col in (1, 6) else center  # teacher+notes alignment
            cell.fill = row_fill
        row_idx += 1

    # Footer banner (merged) after data
    ws.append(["© 2025 — جميع الحقوق محفوظة - مدرسة الشحانية الاعدادية الثانوية"] + [None] * (len(headers) - 1))
    footer_row = row_idx
    ws.merge_cells(
        start_row=footer_row,
        start_column=1,
        end_row=footer_row,
        end_column=len(headers),
    )
    fcell = ws.cell(row=footer_row, column=1)
    fcell.fill = PatternFill("solid", fgColor="F7F2F1")
    fcell.font = bold
    fcell.alignment = center

    # Autosize columns (simple heuristic)
    for col in range(1, len(headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=2):
            v = row[0].value
            val_len = len(str(v)) if v is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(12, max_len + 2))

    # Header/Footer text (Excel print header/footer)
    try:
        ws.oddHeader.right.text = "&D &T"
        ws.oddHeader.center.text = "&Bلوحة الأنصبة — جدول التكليفات&B"
        ws.oddFooter.left.text = ""
        ws.oddFooter.right.text = "&P / &N"
    except Exception:
        try:
            ws.header_footer.oddHeader.center.text = "&Bلوحة الأنصبة — جدول التكليفات&B"  # type: ignore
            ws.header_footer.oddHeader.right.text = "&D &T"  # type: ignore
            ws.header_footer.oddFooter.left.text = ""  # type: ignore
            ws.header_footer.oddFooter.right.text = "&P / &N"  # type: ignore
        except Exception:
            pass

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="teaching_assignments.xlsx"'
    return resp


@login_required
def export_matrix_xlsx(request):
    """Export teacher-class matrix to Excel with Arabic support, A3 fit, colors, RTL, banner header/footer, wrapped cells, and blank empty cells."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Reuse ordering logic from teacher_class_matrix
    classes = list(Class.objects.all().order_by("grade", "section", "id"))
    cat_expr_for_staff = _subject_category_order_expr("assignments__subject__name_ar")
    teachers = (
        Staff.objects.filter(assignments__isnull=False)
        .distinct()
        .annotate(
            life_skills_total=Sum(
                "assignments__no_classes_weekly",
                filter=Q(assignments__subject__name_ar__icontains="مهارات"),
            ),
            non_ls_min=Min(
                Case(
                    When(
                        assignments__subject__name_ar__icontains="مهارات",
                        then=Value(999),
                    ),
                    default=cat_expr_for_staff,
                    output_field=IntegerField(),
                )
            ),
            job_title_cat=_job_title_category_order_expr("job_title"),
            all_min=Min(cat_expr_for_staff),
        )
        .annotate(
            has_social=Sum(
                Case(
                    When(
                        Q(assignments__subject__name_ar__icontains="جغرافيا")
                        | Q(assignments__subject__name_ar__icontains="تاريخ")
                        | Q(assignments__subject__name_ar__icontains="اجتماع")
                        | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                        then=Value(1),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            ),
            # Align ordering with teacher_class_matrix view
            override_cat=Case(
                When(full_name__icontains="وجدي", then=Value(14)),
                When(full_name__icontains="منير", then=Value(12)),
                When(
                    Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                    then=Value(12),
                ),
                When(full_name__icontains="السيد محمد", then=Value(9)),
                When(full_name__icontains="محمد عدوان", then=Value(9)),
                # Specific placement for Social Studies/History/Geography teachers
                When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                When(full_name__icontains="علاء محمد عبد الهادي القضاه", then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            prefer_social=Case(
                When(has_social__gt=0, then=Value(9)),
                default=Value(None),
                output_field=IntegerField(),
            ),
            min_category=Case(
                When(
                    Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                    then=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("job_title_cat"),
                        F("all_min"),
                        Value(999),
                    ),
                ),
                default=Coalesce(
                    F("override_cat"),
                    F("prefer_social"),
                    F("all_min"),
                    F("non_ls_min"),
                    F("job_title_cat"),
                    Value(999),
                ),
                output_field=IntegerField(),
            ),
        )
        .order_by("min_category", "full_name", "id")
    )

    # Build matrix data
    totals_by_class = {c.id: 0 for c in classes}
    matrix = {
        (a["teacher_id"], a["classroom_id"]): int(a["total"] or 0)
        for a in TeachingAssignment.objects.values("teacher_id", "classroom_id").annotate(
            total=Sum("no_classes_weekly")
        )
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "مصفوفة الأنصبة"

    # RTL layout for Arabic
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    # Page setup: A3 landscape, fit
    try:
        ws.page_setup.paperSize = ws.PAPERSIZE_A3  # type: ignore[attr-defined]
    except Exception:
        pass
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    # Styles
    header_fill = PatternFill("solid", fgColor="F7F2F1")
    odd_fill = PatternFill("solid", fgColor="D7EEE5")
    even_fill = PatternFill("solid", fgColor="D9E7FB")
    thin = Side(style="thin", color="8A8F99")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    def cls_label(c: Class) -> str:  # type: ignore
        sec = (c.section or "").strip()
        return f"{c.grade}-{sec}" if sec else str(c.grade)

    header = ["المعلّم"] + [cls_label(c) for c in classes] + ["المجموع"]

    # Banner header row (maroon)
    ws.append(["مصفوفة الأنصبة — المعلّم × الصف"] + [None] * (len(header) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    banner = ws.cell(row=1, column=1)
    banner.fill = PatternFill("solid", fgColor="7B1E1E")
    banner.font = white_bold
    banner.alignment = center

    # Table header at row 2
    ws.append(header)

    # Style header row (row 2)
    for col in range(1, len(header) + 1):
        hc = ws.cell(row=2, column=col)
        hc.fill = header_fill
        hc.font = bold
        hc.alignment = center if col != 1 else right
        hc.border = border

    # Freeze panes below header and after first column
    ws.freeze_panes = ws["B3"]

    # Data rows
    grand_total = 0
    row_idx = 3
    for t in teachers:
        row_vals = [t.full_name]
        s = 0
        for c in classes:
            v = int(matrix.get((t.id, c.id), 0) or 0)
            row_vals.append("" if v == 0 else v)  # blank for empty cells
            s += v
            totals_by_class[c.id] += v
        row_vals.append(s)
        ws.append(row_vals)
        # Styling row
        row_fill = odd_fill if (int(t.min_category or 999) % 2 == 1) else even_fill
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = right if col == 1 else center
            cell.fill = row_fill
        row_idx += 1
        grand_total += s

    # Footer totals row (styled)
    ws.append(["مجموع الأعمدة"] + [totals_by_class[c.id] for c in classes] + [grand_total])
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center if col != 1 else right
        cell.border = border

    # Header/Footer text in print header/footer
    try:
        ws.oddHeader.right.text = "&D &T"
        ws.oddHeader.center.text = "&Bمصفوفة الأنصبة — المعلّم × الصف&B"
        ws.oddFooter.left.text = ""
        ws.oddFooter.right.text = "&P / &N"
    except Exception:
        try:
            ws.header_footer.oddHeader.center.text = "&Bمصفوفة الأنصبة — المعلّم × الصف&B"  # type: ignore
            ws.header_footer.oddHeader.right.text = "&D &T"  # type: ignore
            ws.header_footer.oddFooter.left.text = ""  # type: ignore
            ws.header_footer.oddFooter.right.text = "&P / &N"  # type: ignore
        except Exception:
            pass

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="teacher_class_matrix.xlsx"'
    return resp


@login_required
def export_assignments_pdf(request):
    """Export assignments to PDF.
    Preferred: WeasyPrint (high-fidelity). Fallback: ReportLab with Arabic shaping.
    """
    try:
        from weasyprint import CSS, HTML  # type: ignore

        # Prepare data similar to dashboard table, include category for coloring
        qs = (
            TeachingAssignment.objects.select_related("teacher", "classroom", "subject")
            .annotate(
                teacher_min_category=Subquery(
                    Staff.objects.filter(pk=OuterRef("teacher_id"))
                    .annotate(all_min=Min(_subject_category_order_expr("assignments__subject__name_ar")))
                    .values("all_min")[:1]
                )
            )
            .order_by(
                "teacher_min_category",
                "teacher__full_name",
                "classroom__grade",
                "classroom__section",
                "subject__name_ar",
            )
        )
        context = {
            "assignments": qs,
            "kpi_weekly_total": TeachingAssignment.objects.aggregate(total=Sum("no_classes_weekly")).get("total") or 0,
            "title": "تقرير الأنصبة",
        }
        html = render(request, "school/export_assignments_pdf.html", context)
        document = HTML(
            string=html.content.decode("utf-8"),
            base_url=request.build_absolute_uri("/"),
        )
        css = CSS(
            string=(
                "@page { size: A3 landscape; margin: 8mm } "
                'body { font-family: "Noto Kufi Arabic", "Amiri", sans-serif; } '
                "table { width: 100%; border-collapse: collapse } "
                "th, td { border: 1px solid #8a8f99; padding: 3px } "
                "th { background: #f7f2f1 } "
                'tbody tr[data-cat="odd"] { background: #d7eee5; } '
                'tbody tr[data-cat="even"] { background: #d9e7fb; }'
            )
        )
        pdf_bytes = document.write_pdf(stylesheets=[css])
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="teaching_assignments.pdf"'
        return resp
    except Exception:
        # Fallback to ReportLab (pure Python) so export works without WeasyPrint system deps
        try:
            from io import BytesIO
            from pathlib import Path

            import arabic_reshaper  # type: ignore
            from bidi.algorithm import get_display  # type: ignore
            from reportlab.lib.pagesizes import A4, landscape  # type: ignore
            from reportlab.pdfbase import pdfmetrics  # type: ignore
            from reportlab.pdfbase.ttfonts import TTFont  # type: ignore
            from reportlab.pdfgen import canvas  # type: ignore
        except Exception:
            return HttpResponse(
                "لا يمكن إنشاء PDF حالياً لعدم توفر WeasyPrint ولا مكتبات ReportLab/Arabic. رجاءً استخدم Excel.",
                status=503,
                content_type="text/plain; charset=utf-8",
            )

        # Load font (Amiri from assets)
        repo_root = Path(__file__).resolve().parents[2]
        font_path = repo_root / "assets" / "fonts" / "Amiri" / "Amiri-Regular.ttf"
        try:
            pdfmetrics.registerFont(TTFont("Amiri", str(font_path)))
            font_name = "Amiri"
        except Exception:
            font_name = "Helvetica"

        # Fetch data
        qs = TeachingAssignment.objects.select_related("teacher", "classroom", "subject").order_by(
            "teacher__full_name",
            "classroom__grade",
            "classroom__section",
            "subject__name_ar",
        )

        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=landscape(A4))
        width, height = landscape(A4)

        def A(txt: str) -> str:
            try:
                return get_display(arabic_reshaper.reshape(txt or ""))
            except Exception:
                return txt or ""

        # Title
        c.setFont(font_name, 16)
        c.drawRightString(width - 36, height - 36, A("تقرير الأنصبة"))

        # Table header
        cols = [
            (A("المعلّم"), 220),
            (A("الصف"), 60),
            (A("الشعبة"), 60),
            (A("المادة"), 200),
            (A("حصص/أسبوع"), 80),
        ]
        x_right = width - 36
        y = height - 60
        c.setFont(font_name, 11)
        # header row
        x = x_right
        for label, w in cols:
            c.drawRightString(x, y, label)
            x -= w
        y -= 14
        c.line(36, y + 6, width - 36, y + 6)

        # rows with pagination
        line_h = 13
        rows_per_page = int((y - 40) // line_h)
        count_on_page = 0
        total_weekly = 0
        for a in qs.iterator():
            if count_on_page >= rows_per_page:
                c.showPage()
                c.setFont(font_name, 16)
                c.drawRightString(width - 36, height - 36, A("تقرير الأنصبة (متابعة)"))
                c.setFont(font_name, 11)
                y = height - 60
                x = x_right
                for label, w in cols:
                    c.drawRightString(x, y, label)
                    x -= w
                y -= 14
                c.line(36, y + 6, width - 36, y + 6)
                count_on_page = 0

            vals = [
                A(a.teacher.full_name if a.teacher else ""),
                str(a.classroom.grade if a.classroom else ""),
                A(a.classroom.section or "") if a.classroom else "",
                A(a.subject.name_ar if a.subject else ""),
                str(int(a.no_classes_weekly or 0)),
            ]
            x = x_right
            for (label, w), v in zip(cols, vals):
                c.drawRightString(x, y, v)
                x -= w
            y -= line_h
            count_on_page += 1
            total_weekly += int(a.no_classes_weekly or 0)

        # Footer total
        if y - 20 < 40:
            c.showPage()
            y = height - 60
        c.setFont(font_name, 12)
        c.drawRightString(x_right, y - 10, A(f"المجموع الإجمالي للحصص: {total_weekly}"))

        c.save()
        pdf = buf.getvalue()
        buf.close()
        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="teaching_assignments.pdf"'
        return resp


@login_required
def export_matrix_pdf(request):
    """Export teacher-class matrix to PDF.
    Preferred: WeasyPrint; fallback to ReportLab with Arabic shaping if WeasyPrint isn't available.
    """
    try:
        from weasyprint import CSS, HTML  # type: ignore

        classes = list(Class.objects.all().order_by("grade", "section", "id"))
        cat_expr_for_staff = _subject_category_order_expr("assignments__subject__name_ar")
        teachers = (
            Staff.objects.filter(assignments__isnull=False)
            .distinct()
            .annotate(
                life_skills_total=Sum(
                    "assignments__no_classes_weekly",
                    filter=Q(assignments__subject__name_ar__icontains="مهارات"),
                ),
                non_ls_min=Min(
                    Case(
                        When(
                            assignments__subject__name_ar__icontains="مهارات",
                            then=Value(999),
                        ),
                        default=cat_expr_for_staff,
                        output_field=IntegerField(),
                    )
                ),
                job_title_cat=_job_title_category_order_expr("job_title"),
                all_min=Min(cat_expr_for_staff),
            )
            .annotate(
                has_social=Sum(
                    Case(
                        When(
                            Q(assignments__subject__name_ar__icontains="جغرافيا")
                            | Q(assignments__subject__name_ar__icontains="تاريخ")
                            | Q(assignments__subject__name_ar__icontains="اجتماع")
                            | Q(assignments__subject__name_ar__icontains="علوم اجتماعية"),
                            then=Value(1),
                        ),
                        default=Value(0),
                        output_field=IntegerField(),
                    )
                ),
                # Align ordering with teacher_class_matrix view
                override_cat=Case(
                    When(full_name__icontains="وجدي", then=Value(14)),
                    When(full_name__icontains="منير", then=Value(12)),
                    When(
                        Q(full_name__icontains="أحمد المنصف") | Q(full_name__icontains="احمد المنصف"),
                        then=Value(12),
                    ),
                    When(full_name__icontains="السيد محمد", then=Value(9)),
                    When(full_name__icontains="محمد عدوان", then=Value(9)),
                    # Specific placement for Social Studies/History/Geography teachers
                    When(full_name__icontains="محمد عبدالعزيز يونس عدوان", then=Value(9)),
                    When(full_name__icontains="محمد عبدالله عارف العجلوني", then=Value(9)),
                    When(full_name__icontains="على ضيف الله حمد على", then=Value(9)),
                    When(
                        full_name__icontains="علاء محمد عبد الهادي القضاه",
                        then=Value(9),
                    ),
                    default=Value(None),
                    output_field=IntegerField(),
                ),
                prefer_social=Case(
                    When(has_social__gt=0, then=Value(9)),
                    default=Value(None),
                    output_field=IntegerField(),
                ),
                min_category=Case(
                    When(
                        Q(life_skills_total__lte=2) & Q(non_ls_min__isnull=True),
                        then=Coalesce(
                            F("override_cat"),
                            F("prefer_social"),
                            F("job_title_cat"),
                            F("all_min"),
                            Value(999),
                        ),
                    ),
                    default=Coalesce(
                        F("override_cat"),
                        F("prefer_social"),
                        F("all_min"),
                        F("non_ls_min"),
                        F("job_title_cat"),
                        Value(999),
                    ),
                    output_field=IntegerField(),
                ),
            )
            .order_by("min_category", "full_name", "id")
        )
        agg = TeachingAssignment.objects.values("teacher_id", "classroom_id").annotate(total=Sum("no_classes_weekly"))
        matrix = {(a["teacher_id"], a["classroom_id"]): int(a["total"] or 0) for a in agg}

        def cls_label(c: Class) -> str:  # type: ignore
            sec = (c.section or "").strip()
            return f"{c.grade}-{sec}" if sec else str(c.grade)

        rows = []
        col_totals = [0] * len(classes)
        grand_total = 0
        for t in teachers:
            vals = []
            s = 0
            for idx, c in enumerate(classes):
                v = int(matrix.get((t.id, c.id), 0) or 0)
                vals.append(v)
                s += v
                col_totals[idx] += v
            grand_total += s
            rows.append(
                {
                    "teacher": t.full_name,
                    "vals": vals,
                    "total": s,
                    "cat": int(t.min_category or 999),
                }
            )
        context = {
            "title": "مصفوفة الأنصبة",
            "classes": classes,
            "cls_label": cls_label,
            "rows": rows,
            "col_totals": col_totals,
            "grand_total": grand_total,
        }
        html = render(request, "school/export_matrix_pdf.html", context)
        css = CSS(
            string=(
                "@page { size: A3 landscape; margin: 8mm } "
                'body { font-family: "Noto Kufi Arabic", "Amiri", sans-serif; } '
                "table { width: 100%; border-collapse: collapse; table-layout: fixed; } "
                "th, td { border: 1px solid #8a8f99; padding: 2px } "
                "th { background: #f7f2f1 } "
                'tbody tr[data-cat-parity="odd"] { background: #d7eee5; } '
                'tbody tr[data-cat-parity="even"] { background: #d9e7fb; }'
            )
        )
        document = HTML(
            string=html.content.decode("utf-8"),
            base_url=request.build_absolute_uri("/"),
        )
        pdf_bytes = document.write_pdf(stylesheets=[css])
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="teacher_class_matrix.pdf"'
        return resp
    except Exception:
        # Fallback ReportLab implementation
        try:
            from io import BytesIO
            from pathlib import Path

            import arabic_reshaper  # type: ignore
            from bidi.algorithm import get_display  # type: ignore
            from reportlab.lib.pagesizes import A3, A4, landscape  # type: ignore
            from reportlab.pdfbase import pdfmetrics  # type: ignore
            from reportlab.pdfbase.ttfonts import TTFont  # type: ignore
            from reportlab.pdfgen import canvas  # type: ignore
        except Exception:
            return HttpResponse(
                "لا يمكن إنشاء PDF حالياً لعدم توفر WeasyPrint ولا مكتبات ReportLab/Arabic. رجاءً استخدم Excel.",
                status=503,
                content_type="text/plain; charset=utf-8",
            )

        classes = list(Class.objects.all().order_by("grade", "section", "id"))
        cat_expr_for_staff = _subject_category_order_expr("assignments__subject__name_ar")
        teachers = (
            Staff.objects.filter(assignments__isnull=False)
            .distinct()
            .annotate(all_min=Min(cat_expr_for_staff))
            .order_by("all_min", "full_name", "id")
        )
        agg = TeachingAssignment.objects.values("teacher_id", "classroom_id").annotate(total=Sum("no_classes_weekly"))
        matrix = {(a["teacher_id"], a["classroom_id"]): int(a["total"] or 0) for a in agg}

        use_a3 = len(classes) > 10
        pagesize = landscape(A3 if use_a3 else A4)
        width, height = pagesize

        # Font
        repo_root = Path(__file__).resolve().parents[2]
        font_path = repo_root / "assets" / "fonts" / "Amiri" / "Amiri-Regular.ttf"
        try:
            pdfmetrics.registerFont(TTFont("Amiri", str(font_path)))
            font_name = "Amiri"
        except Exception:
            font_name = "Helvetica"

        def A(txt: str) -> str:
            try:
                return get_display(arabic_reshaper.reshape(txt or ""))
            except Exception:
                return txt or ""

        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=pagesize)

        # Title
        c.setFont(font_name, 16)
        c.drawRightString(width - 36, height - 36, A("مصفوفة الأنصبة"))

        # Column labels
        c.setFont(font_name, 9 if use_a3 else 10)
        x_right = width - 36
        y = height - 60
        # Teacher header at far right
        c.drawRightString(x_right, y, A("المعلّم"))
        # Compute column width (try to fit)
        col_w = max(40, int((width - 120) / (len(classes) + 1)))
        x = x_right - 100  # reserve ~100px for teacher name
        # Class headers to the left
        for cls in classes:
            sec = (cls.section or "").strip()
            lbl = f"{cls.grade}-{sec}" if sec else str(cls.grade)
            c.drawRightString(x, y, A(lbl))
            x -= col_w
        c.drawRightString(x, y, A("المجموع"))
        y -= 14
        c.line(36, y + 6, width - 36, y + 6)

        # Rows
        line_h = 12
        rows_per_page = int((y - 40) // line_h)
        count_on_page = 0
        col_totals = [0] * len(classes)
        grand_total = 0
        for t in teachers:
            if count_on_page >= rows_per_page:
                # footer partial page
                c.showPage()
                c.setFont(font_name, 16)
                c.drawRightString(width - 36, height - 36, A("مصفوفة الأنصبة (متابعة)"))
                c.setFont(font_name, 9 if use_a3 else 10)
                y = height - 60
                c.drawRightString(x_right, y, A("المعلّم"))
                x = x_right - 100
                for cls in classes:
                    sec = (cls.section or "").strip()
                    lbl = f"{cls.grade}-{sec}" if sec else str(cls.grade)
                    c.drawRightString(x, y, A(lbl))
                    x -= col_w
                c.drawRightString(x, y, A("المجموع"))
                y -= 14
                c.line(36, y + 6, width - 36, y + 6)
                count_on_page = 0

            # Row values
            row_sum = 0
            x = x_right
            c.drawRightString(x, y, A(t.full_name))
            x -= 100
            for idx, cls in enumerate(classes):
                v = int(matrix.get((t.id, cls.id), 0) or 0)
                row_sum += v
                col_totals[idx] += v
                c.drawRightString(x, y, str(v) if v else "")
                x -= col_w
            grand_total += row_sum
            c.drawRightString(x, y, str(row_sum))
            y -= line_h
            count_on_page += 1

        # Totals row
        if y - 20 < 40:
            c.showPage()
            y = height - 60
        c.setFont(font_name, 10)
        x = x_right
        c.drawRightString(x, y, A("مجموع الأعمدة"))
        x -= 100
        for idx, cls in enumerate(classes):
            c.drawRightString(x, y, str(col_totals[idx]))
            x -= col_w
        c.drawRightString(x, y, str(grand_total))

        c.save()
        pdf = buf.getvalue()
        buf.close()
        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="teacher_class_matrix.pdf"'
        return resp


# --- Attendance entry page and APIs (teacher) ---


@login_required
def teacher_attendance_page(request):
    # التأكد من أن المستخدم مرتبط بكادر ومعه دور معلم
    staff = Staff.objects.filter(user=request.user, role="teacher").first()
    if not staff:
        messages.error(request, "هذه الصفحة مخصصة للمعلمين فقط.")
        return redirect("data_overview")

    # الصفوف المكلّف بها هذا المعلم (مميّزة حسب الصف فقط)
    assignments = (
        TeachingAssignment.objects.filter(teacher=staff)
        .select_related("classroom")
        .order_by("classroom", "classroom__name")
        .distinct("classroom")
    )
    selected_class_id = request.GET.get("class_id")
    if not selected_class_id and assignments:
        selected_class_id = str(assignments[0].classroom_id)
    selected_date_str = request.GET.get("date") or timezone.localdate().isoformat()
    try:
        selected_date = _date.fromisoformat(selected_date_str)
    except Exception:
        selected_date = timezone.localdate()

    students = []
    if selected_class_id:
        students = list(Student.objects.filter(class_fk_id=selected_class_id, active=True).order_by("full_name"))

    # حساب الحصص المسموحة اليوم ونافذة التحرير
    editable_periods = []
    close_at_map = {}
    allowed_periods_today = []
    server_now_iso = None
    try:
        if selected_class_id:
            editable_set, close_map, allowed = get_editable_periods_for_teacher(
                staff, int(selected_class_id), selected_date
            )
            editable_periods = sorted(list(int(p) for p in editable_set))
            # serialize datetimes to isoformat strings
            close_at_map = {str(k): v.isoformat() for k, v in close_map.items()}
            allowed_periods_today = sorted(list(int(p) for p in allowed))
        server_now_iso = timezone.localtime().isoformat()
    except Exception:
        pass

    ctx = {
        "title": "إدخال الغياب (المعلم)",
        "teacher_assignments": assignments,
        "students": students,
        "today": selected_date,
        "selected_class_id": selected_class_id,
        "editable_periods": editable_periods,
        "close_at_map": close_at_map,
        "allowed_periods_today": allowed_periods_today,
        "server_now": server_now_iso,
        "staff_role": getattr(staff, "role", ""),
    }
    return render(request, "school/teacher_attendance.html", ctx)


@login_required
def api_attendance_records(request):
    staff = Staff.objects.filter(user=request.user, role="teacher").first()
    if not staff:
        return JsonResponse({"error": "forbidden"}, status=403)
    try:
        class_id = int(request.GET.get("class_id") or "0")
        dt = _date.fromisoformat(request.GET.get("date"))
    except Exception:
        return JsonResponse({"error": "bad_request"}, status=400)

    # تحقّق صلاحية الوصول للصف
    if not TeachingAssignment.objects.filter(teacher=staff, classroom_id=class_id).exists():
        return JsonResponse({"error": "forbidden"}, status=403)

    term = get_active_term(dt)
    qs = AttendanceRecord.objects.filter(classroom_id=class_id, date=dt, term=term)
    data = [
        {
            "student_id": r.student_id,
            "period_number": r.period_number,
            "status": r.status,
            "runaway_reason": r.runaway_reason,
            "excuse_type": r.excuse_type,
        }
        for r in qs
    ]
    return JsonResponse({"records": data})


@login_required
def api_attendance_students(request):
    """إرجاع قائمة طلاب الصف المحدد للمعلم الحالي (للواجهة الديناميكية)."""
    staff = Staff.objects.filter(user=request.user, role="teacher").first()
    if not staff:
        return JsonResponse({"error": "forbidden"}, status=403)
    try:
        class_id = int(request.GET.get("class_id") or "0")
    except Exception:
        return JsonResponse({"error": "bad_request"}, status=400)

    if not class_id:
        return JsonResponse({"students": []})

    # تحقق أن المعلم مكلف بهذا الصف
    if not TeachingAssignment.objects.filter(teacher=staff, classroom_id=class_id).exists():
        return JsonResponse({"error": "forbidden"}, status=403)

    st_qs = Student.objects.filter(class_fk_id=class_id).order_by("full_name")
    students = [{"id": s.id, "full_name": s.full_name, "active": bool(getattr(s, "active", True))} for s in st_qs]
    return JsonResponse({"students": students})


@require_POST
@login_required
def api_attendance_bulk_save(request):
    import json as _json

    staff = Staff.objects.filter(user=request.user, role="teacher").first()
    if not staff:
        return JsonResponse({"error": "forbidden"}, status=403)

    try:
        payload = _json.loads(request.body or "{}")
        class_id = int(payload.get("class_id") or 0)
        dt = _date.fromisoformat(payload.get("date"))
        items = payload.get("records", [])
        submit_and_lock = bool(payload.get("submit_and_lock", False))
    except Exception:
        return JsonResponse({"error": "bad_request"}, status=400)

    if not class_id or not items:
        return JsonResponse({"error": "empty"}, status=400)

    # تحقّق صلاحية المعلم على الصف
    if not TeachingAssignment.objects.filter(teacher=staff, classroom_id=class_id).exists():
        return JsonResponse({"error": "forbidden"}, status=403)

    term = get_active_term(dt)
    period_times = get_period_times(dt)
    subject_map = get_subject_per_period(class_id, dt)

    # حصص المعلم المسموحة اليوم + نافذة التحرير الحالية
    editable_periods, _close_map, allowed_today = get_editable_periods_for_teacher(staff, class_id, dt)
    # إن كان المستخدم مشرف جناح، اسمح له بالتحرير لجميع الحصص بدون تقييد نافذة الزمن
    try:
        staff_role = getattr(staff, "role", "")
    except Exception:
        staff_role = ""
    if staff_role == "wing_supervisor":
        # periods 1..12 كحد أعلى منطقي
        editable_periods = set(range(1, 13))
        allowed_today = set(range(1, 13))

    saved = 0
    for item in items:
        try:
            st_id = int(item["student_id"])
            p = int(item["period_number"])
            status = item["status"]
        except Exception:
            continue

        # Skip inactive students: do not record attendance for them
        try:
            if not Student.objects.filter(id=st_id, active=True).exists():
                continue
        except Exception:
            pass

        # 1) يجب أن تكون الحصة ضمن جدول المعلم اليوم لهذا الصف
        if p not in allowed_today:
            continue
        # 2) يجب أن تكون ضمن النافذة الزمنية للتحرير
        if p not in editable_periods:
            continue

        # احترام السجلات المقفلة (ما عدا مشرف الجناح)
        existing = AttendanceRecord.objects.filter(student_id=st_id, date=dt, period_number=p, term=term).first()
        if existing and existing.locked and staff_role != "wing_supervisor":
            continue

        defaults = {
            "classroom_id": class_id,
            "teacher": staff,
            "term": term,
            "date": dt,
            "day_of_week": iso_to_school_dow(dt),
            "period_number": p,
            "start_time": period_times.get(p, (_time(0, 0), _time(0, 1)))[0],
            "end_time": period_times.get(p, (_time(0, 0), _time(0, 1)))[1],
            "status": status,
            "late_minutes": item.get("late_minutes", 0),
            "early_minutes": item.get("early_minutes", 0),
            "runaway_reason": item.get("runaway_reason", ""),
            "excuse_type": item.get("excuse_type", ""),
            "source": ("supervisor" if staff_role == "wing_supervisor" else "teacher"),
        }
        subj_id = subject_map.get(p) or get_default_subject_id()
        if subj_id:
            defaults["subject_id"] = subj_id
        try:
            AttendanceRecord.objects.update_or_create(
                student_id=st_id,
                date=dt,
                period_number=p,
                term=term,
                defaults=defaults,
            )
            saved += 1
        except Exception:
            continue

    if submit_and_lock:
        try:
            # اقفل السجلات التي أنشأها هذا المعلم لهذا الصف/اليوم ضمن حصصه
            AttendanceRecord.objects.filter(
                classroom_id=class_id,
                date=dt,
                term=term,
                period_number__in=list(allowed_today),
            ).update(locked=True)
        except Exception:
            pass

    try:
        update_daily_aggregates(class_id, dt, term)
    except Exception:
        pass

    return JsonResponse({"saved": saved})


# --- Helpers ---


def get_active_term(dt: _date):
    # أبسط اختيار: أول ترم يغطي التاريخ أو آخر ترم مضاف
    t = Term.objects.filter(start_date__lte=dt, end_date__gte=dt).order_by("-start_date").first()
    if not t:
        t = Term.objects.order_by("-start_date").first()
    return t


def get_period_times(dt: _date):
    """Return period start/end times for a given date using PeriodTemplate/TemplateSlot when available.
    Falls back to static defaults if no templates are defined.
    Returns: dict[int, (time, time)] mapping period_number -> (start_time, end_time)
    """
    from .models import PeriodTemplate, TemplateSlot  # local import to avoid circulars

    dow = iso_to_school_dow(dt)
    res = {}
    try:
        tpl_ids = list(PeriodTemplate.objects.filter(day_of_week=dow).values_list("id", flat=True))
        if tpl_ids:
            slots = (
                TemplateSlot.objects.filter(template_id__in=tpl_ids, kind="lesson")
                .exclude(number__isnull=True)
                .order_by("number", "start_time")
            )
            for s in slots:
                # last one wins if duplicates; templates should be consistent
                res[int(s.number)] = (s.start_time, s.end_time)
    except Exception:
        # if DB not migrated/available, ignore and use fallback
        res = {}

    if res:
        return res

    # Fallback static schedule (7 periods)
    base = [
        (8, 0),
        (8, 50),
        (8, 50),
        (9, 40),
        (9, 50),
        (10, 40),
        (10, 50),
        (11, 40),
        (11, 50),
        (12, 40),
        (12, 50),
        (13, 40),
        (13, 50),
        (14, 40),
    ]
    fallback = {}
    for i in range(7):
        s_h, s_m = base[i * 2]
        e_h, e_m = base[i * 2 + 1]
        fallback[i + 1] = (_time(s_h, s_m), _time(e_h, e_m))
    return fallback


def get_subject_per_period(class_id: int, dt: _date):
    term = get_active_term(dt)
    mp = {}
    dow = iso_to_school_dow(dt)
    if dow > 5:
        return mp
    for e in TimetableEntry.objects.filter(classroom_id=class_id, day_of_week=dow, term=term):
        mp[int(e.period_number)] = e.subject_id
    return mp


def get_editable_periods_for_teacher(staff: Staff, class_id: int, dt: _date):
    """Return (editable_set, close_at_map, allowed_periods_today)
    - allowed_periods_today: periods the teacher teaches for that class on dt
    - editable_set: subset of allowed that is currently within [start, start+lock_minutes]
    - close_at_map: period -> close datetime (timezone-aware)
    """
    from datetime import datetime, timedelta

    from django.utils import timezone

    # Teacher's periods for this class today
    term = get_active_term(dt)
    dow = iso_to_school_dow(dt)
    allowed = set(
        TimetableEntry.objects.filter(teacher=staff, classroom_id=class_id, day_of_week=dow, term=term).values_list(
            "period_number", flat=True
        )
    )
    period_times = get_period_times(dt)
    # Get policy
    try:
        from .models import AttendancePolicy

        pol = AttendancePolicy.objects.filter(term=term).order_by("-id").first()
        lock_min = pol.lesson_lock_after_minutes if pol else 120
    except Exception:
        lock_min = 120
    now = timezone.localtime()
    editable = set()
    close_at = {}
    for p in allowed:
        p = int(p)
        if p not in period_times:
            continue
        start_t, _end_t = period_times[p]
        start_dt = timezone.make_aware(datetime.combine(dt, start_t), now.tzinfo)
        close_dt = start_dt + timedelta(minutes=lock_min)
        close_at[str(p)] = close_dt
        if start_dt <= now <= close_dt:
            editable.add(p)
    return editable, close_at, allowed


def get_default_subject_id():
    # إن لم تتوفر مادة من الجدول، اختر أول مادة فعّالة
    s = Subject.objects.filter(is_active=True).order_by("id").first()
    return s.id if s else None


def update_daily_aggregates(class_id: int, dt: _date, term: Term):
    # Placeholder لعملية التجميع اليومي لاحقاً
    return None


# --- Auth utilities (GET-friendly logout) ---


@never_cache
def logout_to_login(request):
    """Logout via GET and redirect to unified login page.
    Provides a GET-friendly logout endpoint to avoid 405 errors from Django's default LogoutView.
    """
    try:
        logout(request)
    except Exception:
        # Never block logout on unexpected errors
        pass
    return redirect("login")


@never_cache
def root_router(request):
    """Root router: avoid LoginView at "/" to prevent redirect loops.
    - Authenticated users → portal_home
    - Anonymous users → login
    """
    if request.user.is_authenticated:
        return redirect("portal_home")
    return redirect("login")


# ================= Wings pages =================
@login_required
def wings_overview(request):
    wings = Wing.objects.all().order_by("id")
    return render(request, "school/wings.html", {"wings": wings})


@login_required
def wing_dashboard(request):
    """Redirect legacy Wing dashboard to the new SPA dashboard.
    Keeps the access check (role) and returns 403 if not a wing supervisor.
    """
    from apps.common.roles import normalize_roles  # type: ignore

    roles_raw = set(request.user.groups.values_list("name", flat=True))  # type: ignore
    staff = Staff.objects.filter(user=request.user).first()
    if staff and getattr(staff, "role", None):
        roles_raw.add(staff.role)
    roles = normalize_roles(roles_raw)
    if "wing_supervisor" not in roles and not getattr(request.user, "is_superuser", False):
        return HttpResponse("مسموح فقط لمشرف الجناح", status=403)

    # Redirect to SPA route (works in dev/prod when the frontend is served at root)
    # Using a hash route avoids server-side routing issues.
    return redirect("/#/wing/dashboard")


@login_required
def wing_detail(request, wing_id: int):
    try:
        wing = Wing.objects.get(pk=wing_id)
    except Wing.DoesNotExist:
        raise Http404("الجناح غير موجود")

    if request.method == "POST":
        sup_id = request.POST.get("supervisor_id")
        if sup_id == "" or sup_id is None:
            wing.supervisor = None
            wing.save(update_fields=["supervisor"])
            messages.success(request, "تم إزالة تعيين المشرف لهذا الجناح")
            return redirect("wing_detail", wing_id=wing.id)
        try:
            staff_obj = Staff.objects.get(pk=sup_id)
            wing.supervisor = staff_obj
            wing.save(update_fields=["supervisor"])
            messages.success(request, "تم تحديث المشرف المسؤول عن الجناح")
            return redirect("wing_detail", wing_id=wing.id)
        except Staff.DoesNotExist:
            messages.error(request, "المعلم المحدد غير موجود")

    staff_list = Staff.objects.all().order_by("full_name", "id")
    return render(
        request,
        "school/wing_detail.html",
        {
            "wing": wing,
            "staff_list": staff_list,
        },
    )


@login_required
def assignments_vs_timetable(request):
    """Compare teacher TeachingAssignment weekly loads vs actual weekly counts in TimetableEntry.
    Shows only differences (shortage/surplus or missing/extra) grouped by teacher.
    """
    # Current term for timetable comparison (if missing, treat timetable as empty)
    term = Term.objects.filter(is_current=True).first()

    # Aggregated assignments per teacher/class/subject
    a_qs = TeachingAssignment.objects.values("teacher_id", "classroom_id", "subject_id").annotate(
        weekly=Sum("no_classes_weekly")
    )
    a_map = {(r["teacher_id"], r["classroom_id"], r["subject_id"]): int(r["weekly"] or 0) for r in a_qs}

    # Aggregated timetable per teacher/class/subject for the current term
    if term:
        from django.db.models import Count

        t_qs = (
            TimetableEntry.objects.filter(term=term)
            .values("teacher_id", "classroom_id", "subject_id")
            .annotate(weekly=Count("id"))
        )
    else:
        t_qs = []
    t_map = {(r["teacher_id"], r["classroom_id"], r["subject_id"]): int(r["weekly"] or 0) for r in t_qs}

    # Union keys to detect extras/missing
    keys = set(a_map.keys()) | set(t_map.keys())

    # Load name maps
    teacher_ids = sorted({k[0] for k in keys})
    class_ids = sorted({k[1] for k in keys})
    subject_ids = sorted({k[2] for k in keys})
    teachers = {s.id: s for s in Staff.objects.filter(id__in=teacher_ids)}
    classes = {c.id: c for c in Class.objects.filter(id__in=class_ids)}
    subjects = {s.id: s for s in Subject.objects.filter(id__in=subject_ids)}

    # Build differences grouped by teacher
    rows_by_teacher = {}
    teacher_totals = {}
    teachers_with_diff = set()

    for k in sorted(keys):
        tid, cid, sid = k
        a = a_map.get(k, 0)
        t = t_map.get(k, 0)
        if a != t:
            teachers_with_diff.add(tid)
            diff = t - a
            rows_by_teacher.setdefault(tid, []).append(
                {
                    "class": classes.get(cid),
                    "subject": subjects.get(sid),
                    "assign_weekly": a,
                    "timetable_weekly": t,
                    "diff": diff,
                }
            )
        # accumulate totals per teacher for summary
        tot = teacher_totals.setdefault(tid, {"assign": 0, "time": 0})
        tot["assign"] += a
        tot["time"] += t

    # Build list of teachers for template (only those with any differences)
    teacher_items = []
    for tid in sorted(
        teachers_with_diff,
        key=lambda i: (teachers.get(i).full_name if teachers.get(i) else ""),
    ):
        t = teachers.get(tid)
        diffs = rows_by_teacher.get(tid, [])
        # sort diffs by class then subject name
        diffs.sort(
            key=lambda r: (
                ((r["class"].grade, r["class"].section, r["class"].name) if r["class"] else (0, "", "")),
                r["subject"].name_ar if r["subject"] else "",
            )
        )
        totals = teacher_totals.get(tid, {"assign": 0, "time": 0})
        totals["diff"] = totals["time"] - totals["assign"]
        teacher_items.append(
            {
                "teacher": t,
                "diffs": diffs,
                "totals": totals,
            }
        )

    # KPIs
    kpi_total_teachers = len(set(teacher_ids))
    kpi_with_diff = len(teachers_with_diff)
    kpi_no_diff = kpi_total_teachers - kpi_with_diff

    context = {
        "title": "مقارنة التكليفات مع الجداول",
        "term": term,
        "teacher_items": teacher_items,
        "kpi_total_teachers": kpi_total_teachers,
        "kpi_with_diff": kpi_with_diff,
        "kpi_no_diff": kpi_no_diff,
    }

    return render(request, "school/assignments_vs_timetable.html", context)


# === DB Audit / Best Practices ===


@user_passes_test(_staff_only)
@login_required
def data_db_audit(request):
    """Lightweight DB best-practices audit for the 'school' app.
    Runs introspection-based checks and suggests constraints/indexes improvements.
    Safe: read-only; no migrations applied automatically.
    """
    # DB info
    try:
        vendor = connection.vendor
        settings_dict = getattr(connection, "settings_dict", {}) or {}
        db_name = settings_dict.get("NAME")
        db_ver = None
        if vendor == "postgresql":
            with connection.cursor() as cursor:
                cursor.execute("SELECT version();")
                row = cursor.fetchone()
                db_ver = row[0] if row else None
        elif vendor == "mysql":
            with connection.cursor() as cursor:
                cursor.execute("SELECT VERSION();")
                row = cursor.fetchone()
                db_ver = row[0] if row else None
        else:
            db_ver = getattr(connection, "server_version", None)
        host = settings_dict.get("HOST") or "127.0.0.1"
        port = settings_dict.get("PORT") or "5432"
        user = settings_dict.get("USER") or "postgres"
        db_info = {
            "vendor": vendor,
            "name": db_name,
            "version": db_ver,
            "host": host,
            "port": str(port),
            "user": user,
        }
    except Exception:
        db_info = {
            "vendor": None,
            "name": None,
            "version": None,
            "host": None,
            "port": None,
            "user": None,
        }

    # Resolve table names
    term_tbl = Term._meta.db_table
    tten_tbl = TimetableEntry._meta.db_table
    atre_tbl = AttendanceRecord._meta.db_table
    tasg_tbl = TeachingAssignment._meta.db_table
    stud_tbl = Student._meta.db_table

    passes = []
    suggestions_core = []
    suggestions_opt = []

    def add_pass(title, detail):
        passes.append({"title": title, "detail": detail})

    def add_sug(title, detail, hint=None, optional=False):
        item = {"title": title, "detail": detail, "hint": hint, "optional": optional}
        if optional:
            suggestions_opt.append(item)
        else:
            suggestions_core.append(item)

    # Fetch constraints and indexes per table
    introspection = connection.introspection
    constraints = {}
    with connection.cursor() as cursor:
        for tbl in [term_tbl, tten_tbl, atre_tbl, tasg_tbl, stud_tbl]:
            try:
                constraints[tbl] = introspection.get_constraints(cursor, tbl)
            except Exception:
                constraints[tbl] = {}

    def has_named(tbl, name_substr):
        for cname in constraints.get(tbl, {}).keys():
            if name_substr.lower() in cname.lower():
                return True
        return False

    def has_index_with_columns(tbl, cols_exact_order):
        cols_exact_order = [c.lower() for c in cols_exact_order]
        for cname, cinfo in constraints.get(tbl, {}).items():
            try:
                if cinfo.get("index"):
                    cols = [c.lower() for c in (cinfo.get("columns") or [])]
                    if cols == cols_exact_order:
                        return True
            except Exception:
                continue
        return False

    # 1) Term: one current term per academic year (conditional unique)
    if has_named(term_tbl, "one_current_term_per_year"):
        add_pass(
            "ترم واحد حالي لكل سنة دراسية",
            "قيد فريد مشروط موجود (one_current_term_per_year)",
        )
    else:
        add_sug(
            "ترم واحد حالي لكل سنة دراسية",
            "يفضل إضافة UniqueConstraint مشروط لضمان وجود ترم حالي واحد لكل سنة.",
            hint=(
                "from django.db import migrations, models\n\n"
                "migrations.AddConstraint(\n"
                "    model_name='term',\n"
                "    constraint=models.UniqueConstraint(\n"
                "        fields=['academic_year', 'is_current'],\n"
                "        condition=models.Q(is_current=True),\n"
                "        name='one_current_term_per_year',\n"
                "    ),\n"
                ")"
            ),
        )

    # 2) TimetableEntry: bounds checks and composite indexes
    if has_named(tten_tbl, "tt_day_between_1_5"):
        add_pass("قيد CHECK على اليوم", "day_of_week بين 1 و 5")
    else:
        add_sug(
            "قيد CHECK على اليوم",
            "ضبط المجال ليكون اليوم بين الأحد والخميس (1..5).",
            hint=(
                "migrations.AddConstraint(\n"
                "    model_name='timetableentry',\n"
                "    constraint=models.CheckConstraint(\n"
                "        check=models.Q(day_of_week__gte=1) & models.Q(day_of_week__lte=5),\n"
                "        name='tt_day_between_1_5',\n"
                "    ),\n"
                ")"
            ),
        )
    if has_named(tten_tbl, "tt_period_between_1_7"):
        add_pass("قيد CHECK على رقم الحصة", "period_number بين 1 و 7")
    else:
        add_sug(
            "قيد CHECK على رقم الحصة",
            "ضبط المجال ليكون رقم الحصة بين 1 و 7.",
            hint=(
                "migrations.AddConstraint(\n"
                "    model_name='timetableentry',\n"
                "    constraint=models.CheckConstraint(\n"
                "        check=models.Q(period_number__gte=1) & models.Q(period_number__lte=7),\n"
                "        name='tt_period_between_1_7',\n"
                "    ),\n"
                ")"
            ),
        )

    if has_index_with_columns(tten_tbl, ["term_id", "teacher_id", "day_of_week", "period_number"]):
        add_pass(
            "فهرس مركب (term,teacher,day,period)",
            "يوجد فهرس يغطي نمط الاستعلام الشائع للمعلم",
        )
    else:
        add_sug(
            "فهرس مركب لجدول الحصص (حسب المعلم)",
            "يسّرع استعلامات مصفوفة الجدول والتقارير.",
            hint=(
                "migrations.AddIndex(\n"
                "    model_name='timetableentry',\n"
                "    index=models.Index(fields=['term','teacher','day_of_week','period_number'], name='tt_term_teacher_day_period_idx'),\n"
                ")"
            ),
        )

    if has_index_with_columns(tten_tbl, ["term_id", "classroom_id", "day_of_week", "period_number"]):
        add_pass(
            "فهرس مركب (term,class,day,period)",
            "يوجد فهرس يغطي نمط الاستعلام الشائع للصف",
        )
    else:
        add_sug(
            "فهرس مركب لجدول الحصص (حسب الصف)",
            "يسّرع التحقق من التعارضات واستعلامات العرض حسب الصف.",
            hint=(
                "migrations.AddIndex(\n"
                "    model_name='timetableentry',\n"
                "    index=models.Index(fields=['term','classroom','day_of_week','period_number'], name='tt_term_class_day_period_idx'),\n"
                ")"
            ),
        )

    # 3) AttendanceRecord: bounds and composite indexes
    if has_named(atre_tbl, "att_day_between_1_7"):
        add_pass("قيد CHECK على اليوم (الحضور)", "day_of_week بين 1 و 7")
    else:
        add_sug(
            "قيد CHECK على اليوم (الحضور)",
            "تحسين جودة البيانات لسجلات الحضور.",
            hint=(
                "migrations.AddConstraint(\n"
                "    model_name='attendancerecord',\n"
                "    constraint=models.CheckConstraint(\n"
                "        check=models.Q(day_of_week__gte=1) & models.Q(day_of_week__lte=7),\n"
                "        name='att_day_between_1_7',\n"
                "    ),\n"
                ")"
            ),
        )
    if has_named(atre_tbl, "att_period_between_1_7"):
        add_pass("قيد CHECK على رقم الحصة (الحضور)", "period_number بين 1 و 7")
    else:
        add_sug(
            "قيد CHECK على رقم الحصة (الحضور)",
            "ضبط المجال لرقم الحصة.",
            hint=(
                "migrations.AddConstraint(\n"
                "    model_name='attendancerecord',\n"
                "    constraint=models.CheckConstraint(\n"
                "        check=models.Q(period_number__gte=1) & models.Q(period_number__lte=7),\n"
                "        name='att_period_between_1_7',\n"
                "    ),\n"
                ")"
            ),
        )

    if has_index_with_columns(atre_tbl, ["classroom_id", "date", "period_number", "term_id"]):
        add_pass("فهرس مركب حضور (class,date,period,term)", "جاهز")
    else:
        add_sug(
            "فهرس مركب للحضور (class,date,period,term)",
            "يسّرع تقارير اليوميات وإدخالات الحضور.",
            hint=(
                "migrations.AddIndex(\n"
                "    model_name='attendancerecord',\n"
                "    index=models.Index(fields=['classroom','date','period_number','term'], name='attrec_class_date_period_term_idx'),\n"
                ")"
            ),
        )

    if has_index_with_columns(atre_tbl, ["student_id", "date", "term_id"]):
        add_pass("فهرس مركب حضور (student,date,term)", "جاهز")
    else:
        add_sug(
            "فهرس مركب للحضور (student,date,term)",
            "يسّرع استعلامات الطالب عبر التاريخ.",
            hint=(
                "migrations.AddIndex(\n"
                "    model_name='attendancerecord',\n"
                "    index=models.Index(fields=['student','date','term'], name='attrec_student_date_term_idx'),\n"
                ")"
            ),
        )

    # 4) TeachingAssignment already has unique_together; suggest composite (teacher,classroom) if frequent
    # We cannot know workload pattern here; provide soft suggestion only if index missing and vendor=postgresql
    if not has_index_with_columns(tasg_tbl, ["teacher_id", "classroom_id"]):
        add_sug(
            "فهرس مركب للتكليفات (teacher,classroom)",
            "اختياري بحسب نمط الاستعلام؛ مفيد للجداول الكبيرة.",
            hint=(
                "migrations.AddIndex(\n"
                "    model_name='teachingassignment',\n"
                "    index=models.Index(fields=['teacher','classroom'], name='ta_teacher_class_idx'),\n"
                ")"
            ),
            optional=True,
        )
    else:
        add_pass("فهرس مركب للتكليفات", "(teacher,classroom) موجود")

    # 5) Optional search improvements (Arabic names)
    add_sug(
        "فهرس Trigram لأسماء الطلاب (اختياري)",
        "لتحسين البحث التقريبي عن الأسماء العربية (PostgreSQL + pg_trgm).",
        hint=(
            "-- PostgreSQL\nCREATE EXTENSION IF NOT EXISTS pg_trgm;\n"
            "CREATE INDEX CONCURRENTLY IF NOT EXISTS student_name_trgm\n"
            "  ON %s USING GIN (full_name gin_trgm_ops);" % stud_tbl
        ),
        optional=True,
    )

    from datetime import datetime as _dt

    suggestions_all = suggestions_core + suggestions_opt

    context = {
        "title": "تدقيق قاعدة البيانات",
        "db_info": db_info,
        "now": _dt.now().strftime("%Y-%m-%d %H:%M"),
        "passes": passes,
        "suggestions": suggestions_all,
        "suggestions_core_count": len(suggestions_core),
        "suggestions_opt_count": len(suggestions_opt),
        "checks_total": len(passes) + len(suggestions_all),
        "checks_ok": len(passes),
        "checks_warn": len(suggestions_all),
    }
    return render(request, "data/audit.html", context)


import json

# === UI Tiles persistence (file-based, minimal backend) ===
from django.views.decorators.http import require_http_methods

_UI_RUNTIME_DIR = (
    os.path.join(settings.BASE_DIR, "backend", ".runtime")
    if hasattr(settings, "BASE_DIR")
    else os.path.join(os.path.dirname(__file__), "..", ".runtime")
)
_UI_TILES_FILE = os.path.join(_UI_RUNTIME_DIR, "ui_tiles.json")


def _ensure_runtime_dir():
    try:
        os.makedirs(_UI_RUNTIME_DIR, exist_ok=True)
    except Exception:
        pass


def _is_ui_designer(user) -> bool:
    try:
        if getattr(user, "is_superuser", False):
            return True
        groups = set(user.groups.values_list("name", flat=True))  # type: ignore
        return bool(groups.intersection({"principal", "it", "admin_deputy"}))
    except Exception:
        return False


@require_http_methods(["GET"])
@login_required
def api_ui_tiles_effective(request):
    """Return saved tiles (draft/effective). If missing, return empty array.
    This endpoint is read-only for any authenticated user to allow SPA to render.
    """
    try:
        _ensure_runtime_dir()
        if os.path.exists(_UI_TILES_FILE):
            with open(_UI_TILES_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # Validate basic shape
            if isinstance(data, dict) and "tiles" in data and isinstance(data["tiles"], list):
                return JsonResponse(data, safe=False)
            if isinstance(data, list):
                return JsonResponse({"tiles": data}, safe=False)
    except Exception:
        # fall through to empty
        pass
    return JsonResponse({"tiles": []})


@require_http_methods(["POST"])
@login_required
def api_ui_tiles_save(request):
    """Persist tiles JSON to the runtime file. Restricted to UI designers.
    Body: { tiles: [ { id, title, subtitle?, icon, color?, to?, href?, roles?, permissions?, kpiKey? } ], version?: int }
    """
    if not _is_ui_designer(request.user):
        return JsonResponse({"detail": "forbidden"}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse({"detail": "invalid json"}, status=400)
    tiles = payload.get("tiles")
    if not isinstance(tiles, list):
        return JsonResponse({"detail": "tiles must be a list"}, status=400)
    # Basic validation and normalization
    seen = set()
    norm_tiles = []
    for t in tiles:
        if not isinstance(t, dict):
            continue
        tid = str(t.get("id") or "").strip()
        title = str(t.get("title") or "").strip()
        icon = str(t.get("icon") or "").strip()
        if not tid or not title or not icon:
            continue
        if tid in seen:
            continue
        seen.add(tid)
        item = {
            "id": tid,
            "title": title,
            "subtitle": (t.get("subtitle") or "").strip() or None,
            "icon": icon,
            "color": (t.get("color") or "").strip() or None,
            "to": (t.get("to") or "").strip() or None,
            "href": (t.get("href") or "").strip() or None,
            "roles": t.get("roles") if isinstance(t.get("roles"), list) else [],
            "permissions": t.get("permissions") if isinstance(t.get("permissions"), list) else [],
            "kpiKey": (t.get("kpiKey") or None),
            "enabled": bool(t.get("enabled", True)),
        }
        norm_tiles.append(item)
    data = {
        "version": int(payload.get("version") or 1),
        "updated_by": getattr(request.user, "username", None),
        "tiles": norm_tiles,
    }
    try:
        _ensure_runtime_dir()
        with open(_UI_TILES_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return JsonResponse({"saved": len(norm_tiles), "version": data["version"]})
    except Exception as e:
        return JsonResponse({"detail": f"failed to save: {e}"}, status=500)


# --- Icons API (filesystem-backed) ---
# Minimal viable implementation to support sc:* icons with caching.
from typing import List, Tuple

ICON_DIRS: List[str] = [
    os.path.join(settings.BASE_DIR, "assets", "icons"),
    os.path.join(settings.BASE_DIR, "backend", "static", "data", "icons"),
]


def _iter_icon_files() -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    seen = set()
    for d in ICON_DIRS:
        if not os.path.isdir(d):
            continue
        try:
            for fn in os.listdir(d):
                if not fn.lower().endswith(".svg"):
                    continue
                name = fn[:-4]
                if name in seen:
                    continue
                seen.add(name)
                out.append((name, os.path.join(d, fn)))
        except Exception:
            continue
    return out


@require_GET
def icons_manifest(request):
    data = {}
    files = _iter_icon_files()
    for name, path in files:
        try:
            stat = os.stat(path)
            ver = int(stat.st_mtime)
            data[name] = {"v": ver}
        except Exception:
            data[name] = {"v": 1}
    return JsonResponse(data)


def _read_icon_file(name: str) -> Tuple[bytes, str, int]:
    name_norm = name.strip().lower()
    for d in ICON_DIRS:
        if not os.path.isdir(d):
            continue
        # try exact
        cand = os.path.join(d, f"{name}.svg")
        if os.path.isfile(cand):
            with open(cand, "rb") as f:
                b = f.read()
            st = os.stat(cand)
            return b, cand, int(st.st_mtime)
        # try case-insensitive scan
        try:
            for fn in os.listdir(d):
                if fn.lower() == f"{name_norm}.svg":
                    full = os.path.join(d, fn)
                    with open(full, "rb") as f:
                        b = f.read()
                    st = os.stat(full)
                    return b, full, int(st.st_mtime)
        except Exception:
            pass
    raise Http404("Icon not found")


@require_GET
def icon_svg(request, name: str):
    try:
        content, path, ver = _read_icon_file(name)
    except Http404:
        return HttpResponse(status=404)
    # ETag from sha1 of content
    import hashlib as _hl

    etag = '"' + _hl.sha1(content).hexdigest() + '"'
    inm = request.META.get("HTTP_IF_NONE_MATCH")
    if inm and inm.strip() == etag:
        return HttpResponseNotModified()

    # Normalize basic headers
    resp = HttpResponse(content, content_type="image/svg+xml; charset=utf-8")
    resp["ETag"] = etag
    # Cache aggressively if client passes correct version
    v = request.GET.get("v")
    try:
        v_int = int(v) if v else None
    except Exception:
        v_int = None
    if v_int and v_int == ver:
        resp["Cache-Control"] = "public, max-age=31536000, immutable"
    else:
        resp["Cache-Control"] = "public, max-age=60"
    return resp
